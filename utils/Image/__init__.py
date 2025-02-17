import os
import re
import shutil
import numpy as np
import pandas as pd
from typing import Dict, List, Literal, Union
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from PIL import Image as PILImage
from PIL import ImageDraw, ImageFont
from IPython.display import display, HTML

def list_directories(path=None):
    '''### 현재 디렉토리의 하위 디렉토리 목록을 반환
- `path` : 디렉토리의 경로 (기본값 : None = 현재 작업 디렉토리)
    '''
    if path is None:
        path = os.getcwd()
    
    directories = [item for item in os.listdir(path) if os.path.isdir(os.path.join(path, item))]
    return directories


def list_image_files(path=None):
    '''### 지정된 경로 또는 현재 경로에서 모든 이미지 파일을 리스트로 반환
    - `.jpg`, `.jpeg`, `.png`, `.gif`, `.bmp`, `.jfif'`
    - `path`: 이미지 파일을 검색할 디렉토리 경로 (기본값: None, 현재 작업 디렉토리 사용)
    '''
    # path가 None이면 현재 작업 디렉토리를 사용합니다.
    if path is None:
        path = os.getcwd()

    # 지원되는 이미지 확장자 목록
    supported_extensions = ('.jpg', '.jpeg', '.png', '.gif', '.bmp', '.jfif')
    
    # 해당 경로에서 모든 파일을 검색하고, 이미지 확장자를 가진 파일만 필터링합니다.
    image_files = [os.path.join(path, file) for file in os.listdir(path)
                   if os.path.isfile(os.path.join(path, file)) and file.lower().endswith(supported_extensions)]
    
    return image_files


def image_rename(image_path, save_folder=None, name_format="%s", file_format='png', display_print=True):
    '''### 이미지 파일 이름을 코드로 변환
    - `image_path`: 이미지 경로 및 이름
    - `save_folder`: 저장할 폴더 이름 (기본값: None)
    - `name_format`: 저장할 파일 이름 형식 (기본값: "%s") / %s 부분이 변환된 파일 이름으로 대체
    - `file_format`: 저장할 이미지 확장자 (기본값: png)
    - `display_print`: 파일 이름 변환 결과 출력 여부 (기본값: True)

    파일명에 숫자가 포함되어 있는 경우에만 작동
    파일명에 숫자가 포함되어 있지 않은 경우, 파일명을 `[NONE_CODE_이미지명]`으로 변경
    포함된 숫자가 2개 이상인 경우 '_'로 구분되어서 파일명이 변경됨 (예: 1_2.png)
    이미지 확장자는 `png`(기본값)으로 변경됨
    '''
    # 경로에서 파일 이름을 분리
    image_name = os.path.split(image_path)[-1]

    if not os.path.exists(image_path):
        if display_print :
            display(HTML(f"""❓ <b style="color: #e7046f"><i>The file does not exist</i></b> : {image_path}"""))

        return {'original_name': image_name, 'new_name': None, 'none_check': None, 'duplicate_check': None}
    
    dir_path = os.path.split(image_path)[:-1]

    file_name, image_format = os.path.splitext(image_name)
    
    # 파일명에서 숫자 추출
    find_numbers = re.findall(r'\d+', file_name)
    find_numbers = [str(int(x)) for x in find_numbers]
    new_name = None

    # 숫자가 없는 경우
    none_check = False

    if len(find_numbers) == 0:
        new_name = f'NONE_CODE_{image_name}'
        none_check = True
    else:
        new_name = '_'.join(find_numbers) + f'.{file_format}'

    new_name = name_format%new_name

    version = 1    
    # save_folder가 제공된 경우 폴더가 없으면 생성
    if save_folder is not None:
        check_dir = os.path.join(*dir_path, save_folder)
        if not os.path.exists(check_dir):
            os.makedirs(os.path.join(*dir_path, save_folder))
        # 새로운 파일 경로 조합
        new_path = os.path.join(*dir_path, save_folder, new_name)

        while True :
            if os.path.exists(new_path) :
                version += 1
                new_name = new_name.split('.')[0]
                new_name = f'{new_name}_v{version}.{file_format}'
                new_path = os.path.join(*dir_path, save_folder, new_name)
            else :
                break
        
        new_path = os.path.join(*dir_path, save_folder, new_name)

        # 파일 복사
        shutil.copy(image_path, new_path)
    else:
        # save_folder가 제공되지 않은 경우, 현재 위치에 파일 이름 변경

        while True :
            if os.path.exists(os.path.join(*dir_path, new_name)) :
                version += 1
                new_name = new_name.split('.')[0]
                new_name = f'{new_name}_v{version}.{file_format}'
            else :
                break
        
        os.rename(os.path.join(*dir_path, image_name), os.path.join(*dir_path, new_name))
    
    if len(find_numbers) >= 2 :
        if display_print :
            display(HTML(f"""⚠️ <b style="color: #e7046f">{image_name}</b> : <i>More than one number found in the regular expression</i>"""))

    if display_print : 
        display(HTML(f"""✔️ <b><i>Rename</i></b> : {image_name} → <b style="color: #2d6df6">{new_name}</b>"""))

    # 파일명 중복으로 인한 처리
    dup_check = True if version > 1 else False

    return {'original_name': image_name, 'new_name': new_name, 'none_check': none_check, 'duplicate_check': dup_check}



def image_resize(image_path, width, height, save_folder=None, display_print=True):
    '''### 이미지를 캔버스에 맞게 크기를 조정하고 투명 배경의 PNG로 저장하는 함수
    - 기본적으로 `png`로 저장되며, 원본 이미지를 제외한 배경은 투명이 된다.
    - `image_path`: 이미지 파일 경로
    - `width`: 이미지의 width
    - `height`: 이미지의 height
    - `save_folder`: 이미지 저장 폴더 (기본값: None) / None이면 원본 파일 위치에 저장
    - `display_print`: 변환 결과 출력 여부 (기본값: True)
    '''
    # 파일 존재 여부 확인
    if not os.path.exists(image_path):
        if display_print:
            display(HTML(f"""❓ <b style="color: #e7046f"><i>The file does not exist</i></b>: <span>{image_path}</span>"""))
        
        return {'image_name': os.path.basename(image_path), 'resize_image_name': None, 'original_width': None, 'new_width': None, 'original_height': None, 'new_height': None}
        
    # 이미지 로드
    img = Image.open(image_path).convert('RGBA')  # 이미지를 RGBA 모드로 변환
    img_width, img_height = img.size

    # 캔버스의 비율과 이미지의 비율을 계산
    canvas_ratio = width / height
    image_ratio = img_width / img_height

    # 이미지가 캔버스에 최대한 크게 들어가도록 크기를 조정
    if image_ratio > canvas_ratio:
        # 이미지의 가로가 캔버스의 가로에 비해 크면 가로 기준으로 조정
        new_width = width
        new_height = int(new_width / image_ratio)
    else:
        # 이미지의 세로가 캔버스의 세로에 비해 크거나 같으면 세로 기준으로 조정
        new_height = height
        new_width = int(new_height * image_ratio)

    # 이미지 크기를 조정
    resized_img = img.resize((new_width, new_height), Image.Resampling.LANCZOS)

    # 새 캔버스 생성 (RGBA 모드, 투명 배경)
    canvas = Image.new('RGBA', (width, height), (255, 255, 255, 0))

    # 이미지를 캔버스 중앙에 배치하기 위해 시작 좌표 계산
    start_x = (width - new_width) // 2
    start_y = (height - new_height) // 2

    # 캔버스에 이미지 삽입
    canvas.paste(resized_img, (start_x, start_y), resized_img)  # resized_img를 마스크로 사용하여 투명 배경 유지

    # 결과 이미지 파일명 생성 (원본 파일명 사용)
    orignal_filename = os.path.basename(image_path)
    if save_folder is None:
        save_folder = os.path.dirname(image_path)
        new_filename = os.path.splitext(orignal_filename)[0] + '.png'
    else:
        # 지정된 폴더가 존재하지 않는 경우, 폴더 생성
        original_path = os.path.split(image_path)[:-1]
        os.makedirs(os.path.join(*original_path, save_folder), exist_ok=True)
        
        save_folder = os.path.join(*original_path, save_folder)
        new_filename = os.path.splitext(orignal_filename)[0] + '.png'

    # 최종 저장 경로
    save_path = os.path.join(*os.path.split(save_folder), new_filename)

    # 파일 저장
    canvas.save(save_path)
    if display_print :
        display(HTML(f"""✔️ <b><i>Resize and save file complete ({width}x{height})</i> : <b style="color: #2d6df6;">{save_path}</b></b>"""))

    return {'image_name': os.path.basename(image_path), 'resize_image_name': new_filename, 'original_width': img_width, 'new_width': width, 'original_height': img_height, 'new_height': height}


def get_mean_width(path=None) :
    '''### 경로에 있는 이미지의 평균 너비를 반환
- `path` : 경로 (기본값 : None) / None인 경우 현재 경로에서 진행
    '''
    images = list_image_files(path)
    if not images :
        display(HTML('''⚠️ <b style="color: #e7046f;"><i>No images found</i></b>'''))
        return
    widths = [Image.open(img).convert('RGBA').size[0] for img in images]
    return int(np.mean(widths))


def image_re_all(image_path=None, save_folder=None, name_format="%s", file_format='png', width=500, height=500, display_print=True) :
    """### `image_rename` / `image_resize` 모두 실행
- `image_path` : 이미지 경로 (기본값 : None)
- `save_folder` : 저장 경로 (기본값 : None)
- `name_format`: 저장할 파일 이름 형식 (기본값: "%s") / %s 부분이 변환된 파일 이름으로 대체
- `file_format` : 저장 파일 포맷 (기본값 : 'png')
- `width` : `resize` 시 너비 (기본값 : 500)
- `height` : `resize` 시 높이 (기본값 : 500)
- `display_print` : 변환 결과 출력 여부 (기본값 : True)
    """
    data = []

    if image_path is None :
        image_path = os.getcwd()

    parent_path = os.path.split(image_path)
    for img in list_image_files(image_path) :
        rename_data = image_rename(img, save_folder, name_format, file_format, display_print=False)
        new_name = rename_data['new_name']
        if save_folder is None :
            rename_path = os.path.join(*parent_path, new_name)
        else :
            rename_path = os.path.join(*parent_path, save_folder, new_name)
            
        resize_data = image_resize(image_path=rename_path, width=width, height=height, save_folder=None, display_print=False)
        
        data.append({**rename_data, **resize_data})
    
    if display_print :
        for d in data :
            orginal_name = d['original_name']
            new_name = d['new_name']
            none_check = d['none_check']
            duplicate_check = d['duplicate_check']

            if none_check :
                display(HTML(f"""⚠️ <b style="color: #e7046f">{orginal_name}</b> : <i>The code (number) does not exist in the filename.</i>"""))

            if duplicate_check :
                display(HTML(f"""⚠️ <b style="color: #e7046f">{orginal_name}</b> : <i>The code (number) is duplicated.</i> > <b style="color: #2d6df6;">{new_name}</b>"""))

            display(HTML(f"""✔️ <b><i>Rename/Resize Complete (Size: {width}x{height})</i> : {orginal_name} → <b style="color: #2d6df6">{new_name}</b>"""))

    return pd.DataFrame(data)



def create_dummy_img(save_name, save_path=None, file_format='png', width=500, height=500, background_color=(255, 255, 255, 0)) :
    '''### 더미 이미지 생성
    - `save_name` : 이미지 이름
    - `save_path` : 저장할 경로 (기본값 : None) / 경로가 지정되지 않으면 현재 경로를 저장 경로로 사용
    - `width` : 이미지 너비 (기본값 : 500)
    - `height` : 이미지 높이 (기본값 : 500)
    '''

    # 이미지 생성
    image = Image.new('RGBA', (width, height), background_color)

    # 드로잉 객체 생성
    draw = ImageDraw.Draw(image)

    # 목표 텍스트 너비 설정 (이미지 너비의 65%)
    target_text_width = width * 0.65

    # 폰트 크기 결정을 위한 초기 값 설정
    font_size = 10
    font = ImageFont.truetype("arial.ttf", font_size)
    text_width = 0

    # 적절한 폰트 크기 찾기
    while text_width < target_text_width:
        font_size += 1
        font = ImageFont.truetype("arial.ttf", font_size)
        img_name = f'{save_name}.{file_format}'
        text_bbox = draw.textbbox((0, 0), img_name, font=font)
        text_width = text_bbox[2] - text_bbox[0]
        if text_width > target_text_width:
            font_size -= 1
            break

    # 최종 폰트 설정
    font = ImageFont.truetype("arial.ttf", font_size)
    text_bbox = draw.textbbox((0, 0), img_name, font=font)
    text_width = text_bbox[2] - text_bbox[0]
    text_height = text_bbox[3] - text_bbox[1]

    # 텍스트 중앙 배치
    x = (width - text_width) / 2
    y = (height - text_height) / 2

    # 텍스트 그리기
    draw.text((x, y), img_name, fill="black", font=font)

    # 저장 경로 설정
    if save_path is None:
        img_save_path = os.getcwd()  # 현재 작업 디렉토리 사용
    else:
        img_save_path = os.path.join(*os.path.split(save_path))  # 주어진 경로의 폴더 부분 사용
        os.makedirs(img_save_path, exist_ok=True) # 폴더 생성

    path = os.path.join(img_save_path, img_name)
    image.save(path)




# Image Insert to Excel 
def insert_img_in_ws(ws, image_path: str, col_letter: str, row_idx: int, target_width: int, row_height: int) -> None:
    """
    이미 열려있는 worksheet(ws)에 이미지를 삽입합니다.
    """
    if not os.path.exists(image_path):
        raise FileNotFoundError(f"Image file not found: {image_path}")
    
    img_for_excel = Image(image_path)
    img = PILImage.open(image_path)
    
    if img.format == 'MPO':
        img = img.convert('RGB')
        img.save(image_path)
        img_for_excel = Image(image_path)
    
    width_percent = target_width / img.width
    new_height = int(img.height * width_percent)
    img_for_excel.width = target_width
    img_for_excel.height = new_height

    # 해당 셀의 열 및 행 크기를 조정
    ws.column_dimensions[col_letter].width = (target_width // 7) - 1
    ws.row_dimensions[row_idx].height = row_height

    ws.add_image(img_for_excel, f'{col_letter}{row_idx}')

def sort_key(key: str):
    match = re.match(r"([A-Z]+)(\d+)", key)
    return (match.group(1), int(match.group(2))) if match else (key, 0)

def process_image_insertion(
    data_file: str, 
    sheet: str = None,
    platform: Literal['decipher', 'stg'] = 'decipher',
    images_path: str = 'images', 
    img_variables: Union[List[str], None] = None,
    row_height: int = 205, 
    target_width: int = 110,
    mkdir: bool = True,
    dir_name: str = 'insert_img',
) -> None:
    """엑셀 파일에 이미지를 삽입합니다.
    
    Args:
        data_file: 이미지를 삽입할 엑셀 파일 경로
        sheet: 작업할 시트 이름 (None인 경우 첫 번째 시트 사용)
        platform: 데이터 플랫폼 유형 ('decipher' 또는 'stg')
        images_path: 이미지 파일이 있는 디렉토리 경로
        img_variables: 이미지를 삽입할 변수명 리스트 (None인 경우 자동 감지)
        row_height: 이미지가 들어갈 행의 높이
        target_width: 삽입할 이미지의 목표 너비
        mkdir: 결과 파일을 저장할 새 디렉토리 생성 여부
        dir_name: 결과 파일을 저장할 디렉토리 이름
    """
    def print_status(message: str, end='\r', flush=True):
        print(message, end=end, flush=flush)
    
    print_status('📝 Starting image insertion process...')
    
    if platform is None:
        raise ValueError('platform value error')
    
    index_col_dict = {'decipher': 'record', 'stg': 'SbjNum'}
    index_col = index_col_dict[platform]
    
    # Excel 데이터를 읽어 record 리스트 생성
    if sheet is None:
        df = pd.read_excel(data_file, index_col=index_col)
    else:
        df = pd.read_excel(data_file, sheet_name=sheet, index_col=index_col)
    df.index = df.index.astype(str)
    records = list(df.index)
    
    print_status('📊 Loading Excel workbook...')
    # 새 파일 생성 (중복 피하기)
    if mkdir:
        if not isinstance(dir_name, str):
            raise ValueError('dir_name must be str')
        os.makedirs(dir_name, exist_ok=True)
    version = 1
    base_filename = os.path.basename(data_file)
    base_name = f"v{version}_Img_{base_filename}"
    new_name = os.path.join(dir_name, base_name) if mkdir else base_name
    while os.path.exists(new_name):
        version += 1
        base_name = f"v{version}_Img_{base_filename}"
        new_name = os.path.join(dir_name, base_name) if mkdir else base_name

    # 원본 파일을 새 파일로 복사
    wb_temp = load_workbook(data_file)
    wb_temp.save(new_name)
    wb_temp.close()
    
    print_status('🔍 Analyzing image variables...')
    # 새 파일을 한 번만 열어 작업 (Workbook은 메모리상에서 처리)
    wb = load_workbook(new_name)
    sheet_name = sheet or wb.sheetnames[0]
    ws = wb[sheet_name]
    max_cols = ws.max_column
    
    images = os.listdir(images_path)
    
    # 이미지 변수 설정: 자동 감지 혹은 인자 사용
    if img_variables is None:
        match platform:
            case 'stg':
                survey_info = [
                    [img.split('_--_')[0].split('_')[0], '_'.join(img.split('_--_')[0].split('_')[1:])]
                    for img in images
                ]
                variables = set([s[-1] for s in survey_info])
                summary_var: Dict[str, List[str]] = {}
                for v in variables:
                    if '_' not in v:
                        summary_var[v] = [v]
                    else:
                        base, row = v.split('_')
                        if base not in summary_var:
                            rows = sorted(set(
                                int(re.findall(r'R(\d+)', i.split('_')[-1])[0])
                                for i in variables if base in i
                            ))
                            summary_var[base] = [f'{base}_R{r}' for r in rows]
                sorted_summary_var = dict(sorted(summary_var.items(), key=lambda item: sort_key(item[0])))
                img_variables = sum(sorted_summary_var.values(), [])
            case 'decipher':
                blob = df.astype(str).apply(lambda col: col.map(lambda x: "blob" in x)).any()
                img_variables = blob[blob].index.tolist()

    print_status('📝 Preparing worksheet columns...')
    # 새 열을 기존 열 뒤에 추가하기
    match_col_index = {col: max_cols + idx + 1 for idx, col in enumerate(img_variables)}
    for col, col_index in match_col_index.items():
        ws.cell(row=1, column=col_index, value=col)
    

    # 이미지 파일들을 (record, variable) 튜플을 키로 갖는 딕셔너리로 미리 매핑 (파일명에 "record_variable" 패턴일 경우)
    image_lookup = {}
    for img_file in images:
        for rec in records:
            for var in img_variables:
                if f"{rec}_{var}" in img_file:
                    image_lookup[(rec, var)] = img_file
                    break  # 중복 매핑 방지
    
    print_status('🖼️ Inserting images...')
    total_records = len(records)
    for idx, rec in enumerate(records, start=2):
        for var, col_index in match_col_index.items():
            img_file = image_lookup.get((rec, var))
            if img_file:
                col_letter = get_column_letter(col_index)
                img_path = os.path.join(images_path, img_file)
                insert_img_in_ws(ws, img_path, col_letter, idx, target_width, row_height)
        progress = ((idx - 1) / total_records) * 100
        print_status(f'🖼️ Processing images... {progress:.1f}%')
    
    print_status('💫 Applying final formatting...')
    # 모든 셀에 대해 가운데 정렬 적용
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    wb.save(new_name)
    wb.close()
    print_status('✨ Image insertion complete!', end='\n')