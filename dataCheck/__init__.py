import pandas as pd
import numpy as np
import types
import openpyxl
import re
import nbformat as nbf
from collections import OrderedDict
import os
from zipfile import ZipFile
import json
from decipher.beacon import api
import time
from ..key import api_key, api_server
from decipherAutomatic.getFiles import *
from decipherAutomatic.utils import *
from pandas.io.formats import excel
from typing import Union, List, Tuple, Dict, Optional

def with_cols_check(with_cols: List) -> bool:
    """입력이 리스트인지 확인합니다."""
    if not isinstance(with_cols, list):
        print("❌ [ERROR] Type of with_cols must be list")
        return True
    return False

def df_err_check(df: bool, err: bool) -> bool:
    """입력 값이 boolean 타입인지 확인합니다."""
    if not isinstance(df, bool):
        print("❌ [ERROR] The type of df must be bool")
        return True
    if not isinstance(err, bool):
        print("❌ [ERROR] The type of err must be bool")
        return True
    return False

def df_err_return(df: bool, df_return, err: bool, err_return) -> Union[Dict[str, pd.DataFrame], pd.DataFrame, bool]:
    """`df`와 `err` 플래그에 따라 적절한 값을 반환합니다."""
    if df and not err:
        return df_return
    elif not df and err:
        return err_return
    elif df and err:
        return {"df": df_return, "err": err_return}
    else:
        return False

def sa_check(sa: Optional[str]) -> bool:
    """`sa`가 문자열인지 확인합니다."""
    if not sa or not isinstance(sa, str):
        print("❌ [ERROR] Please check variable name / Type must be str")
        print(" example) 'Q1'")
        return True
    return False

def ma_check(ma: Union[List, Tuple], cols: List[str], len_chk: bool = True) -> bool:
    """`ma`가 리스트나 튜플인지, 그리고 다른 조건들을 만족하는지 확인합니다."""
    if not ma:
        print("❌ [ERROR] Please check variable names")
        print(" example1) ['Q1r1', 'Q1r2', 'Q1r3']")
        print(" example2) ('Q1r1', 'Q1r3')")
        return True

    if not isinstance(ma, (list, tuple)):
        print("❌ [ERROR] Type of variable must be list or tuple")
        return True

    if len_chk and len(ma) < 2:
        print("❌ [ERROR] Variable must be 2 length or more")
        return True

    if isinstance(ma, tuple) and len(ma) != 2:
        print("❌ [ERROR] The variable must include 2 arguments")
        return True

    if isinstance(ma, tuple):
        first_index = cols.index(ma[0])
        last_index = cols.index(ma[1])
        if first_index > last_index:
            print(f"❌ [ERROR] Please check the column index / current index ( {first_index}-{last_index} )")
            return True
    return False

def ma_return(ma: Union[List, Tuple], cols: List[str]) -> List[str]:
    """`ma`에 지정된 열을 반환합니다."""
    if isinstance(ma, tuple):
        first_index = cols.index(ma[0])
        last_index = cols.index(ma[1]) + 1
        return cols[first_index:last_index]
    elif isinstance(ma, list):
        return ma

def cond_check(cond: Optional[pd.Series], add_text: str = None) -> Union[bool, None]:
    """`cond`가 pandas 시리즈인지 확인합니다."""
    if isinstance(cond, pd.Series):
        return True
    else:
        if cond is None:
            return None
        else:
            if add_text:
                print(f"❌ [ERROR] {add_text}")
            print("❌ [ERROR] Type of cond must be pandas.core.series.Series type")
            return False

def list_check(_list: Optional[List], add_text: str = "") -> bool:
    """`_list`가 리스트인지 확인합니다."""
    if _list is not None and not isinstance(_list, list):
        print(f"❌ [ERROR] Type of {add_text} must be list")
        return True
    return False

def int_check(_int: Optional[int], add_text: str = "") -> bool:
    """`_int`가 정수인지 확인합니다."""
    if _int is not None and not isinstance(_int, int):
        print(f"❌ [ERROR] Type of {add_text} must be int")
        return True
    return False

def str_check(_str: Optional[str], add_text: str = "") -> bool:
    """`_str`가 문자열인지 확인합니다."""
    if _str is not None and not isinstance(_str, str):
        print(f"❌ [ERROR] Type of {add_text} must be str")
        return True
    return False

def none_check(_none: Optional, add_text: str = "") -> bool:
    """`_none`이 None인지 확인합니다."""
    if _none is None:
        print(f"❌ [ERROR] Please check {add_text}")
        return True
    return False

def list_int_check(_check: Union[List, int, None], add_text: str = "") -> bool:
    """`_check`가 리스트 또는 정수인지 확인합니다."""
    if _check is not None:
        if not isinstance(_check, (list, int)):
            print(f"❌ [ERROR] Please check {add_text}")
            return True

        if isinstance(_check, list):
            for c in _check:
                if not isinstance(c, int):
                    print(f"❌ [ERROR] Please check value in {add_text}")
                    return True
    return False

def list_or_int_set(_check: Union[List, int, None]) -> List[int]:
    """리스트나 정수를 리스트 형태로 반환합니다."""
    if _check is None:
        return []
    if isinstance(_check, list):
        return _check
    if isinstance(_check, int):
        return [_check]

def sum_list(*args: List) -> List:
    """리스트의 요소들을 더합니다."""
    return sum([*args], [])

def key_id_check(base: List[str], var: str, var_type: str) -> Dict[str, Union[bool, str, List[str]]]:
    """`base`가 `qid`를 포함하는지 확인합니다."""
    qid = base[0]
    qids = list(qid)
    qids.reverse()
    for q in qids:
        if not q.isdigit():
            break
        else:
            qid = qid[:-1]

    for ma in base:
        if qid not in ma:
            print("❌ [ERROR] Please check multi question variable names")
            print(f"{var_type} variable name : {var}")
            print(f"Base MA variable key name : {qid}")
            return {"ok": False, "return": base}

    return {"ok": True, "return": qid}

def not_empty_cell(cell, row: int) -> bool:
    """셀 값이 비어 있지 않은지 확인합니다."""
    a = cell.cell(row, 1).value
    b = cell.cell(row, 2).value
    c = cell.cell(row, 3).value

    return bool(a or b or c)

def re_big(txt: str) -> Optional[str]:
    """대괄호 안의 내용을 추출합니다."""
    re_chk = re.search(r'\[(.*?)\]', txt)
    if re_chk:
        return re_chk.group(1).strip()
    return None

def colon_split(txt: str, num: int) -> Optional[str]:
    """콜론으로 텍스트를 나누고 지정된 부분을 반환합니다."""
    re_chk = txt.split(":")
    if re_chk:
        return re_chk[num].strip()
    return None

class Ready:
    def __init__(self, dataframe: pd.DataFrame, key_id: str = 'record', include_cols: List[str] = []):
        """
        데이터 변환 및 검증을 위한 클래스
        
        Args:
            `dataframe` (pd.DataFrame): 처리할 데이터프레임.
            `key_id` (str, optional): 데이터프레임의 키 식별자. 기본값은 'record'.
            `include_cols` (List[str], optional): 기본적으로 포함할 열 목록. 기본값은 [].
        """
        self.df: pd.DataFrame = dataframe
        self.back_df: pd.DataFrame = self.df.copy()
        self.key_id: str = key_id
        self.default_show_cols: List[str] = sum([[self.key_id], include_cols], [])
        self.rows: List[int] = list(self.df.index)
        self.cols: List[str] = list(self.df.columns)
        self.separator: str = "-"*10 + "\n\n"
        self.err_col: str = 'err'
        self.only_col: str = 'only_err'
        self.count_label: str = "count"
        self.exist_col: str = "exist"
        self.masa_label: str = "missing_col"
        self.ma_base: str = "ma_base"
        self.ma_answer: str = "ma_answer"

    def change_df(self, new_dataframe: pd.DataFrame) -> pd.DataFrame:
        """
        데이터프레임을 변경
        
        Args:
            `new_dataframe` (pd.DataFrame): 새 데이터프레임.
        
        Returns:
            pd.DataFrame: 변경된 데이터프레임.
        """
        self.df = new_dataframe
        return self.df

    def reset(self) -> pd.DataFrame:
        """
        데이터프레임을 초기 상태로 되돌립니다.
        
        Returns:
            pd.DataFrame: 초기화된 데이터프레임.
        """
        self.df = self.back_df.copy()
        return self.df

    def apply_col(self, col: str, apply: Optional[types.FunctionType] = None, with_cols: Optional[List[str]] = None, cond: Optional[pd.Series] = None, axis: str = "row") -> pd.DataFrame:
        """
        조건에 따라 데이터프레임 열에 함수를 적용
        
        Args:
            `col` (str): 함수를 적용할 열 이름.
            `apply` (types.FunctionType, optional): 적용할 함수. 기본값은 None.
            `with_cols` (List[str], optional): 추가로 포함할 열. 기본값은 None.
            `cond` (pd.Series, optional): 함수를 적용할 조건. 기본값은 None.
            `axis` (str, optional): 함수를 적용할 축. 기본값은 "row".
        
        Returns:
            pd.DataFrame: 변경된 데이터프레임.
        """
        show_cols = self.default_show_cols.copy()
        
        if none_check(col, add_text="column name"): return self.df
        if str_check(col, add_text="column name"): return self.df
        
        if axis not in ["row", "col"]:
            print("❌ [ERROR] axis is available only 'row' or 'col'")
            return self.df

        if with_cols is not None:
            if ma_check(with_cols, self.cols, len_chk=False): return self.df
            with_cols = ma_return(with_cols, self.cols)
        else:
            with_cols = []

        axis = 1 if axis == "row" else 0
        show_cols = sum_list(show_cols, [col], with_cols)

        curr_df = self.df.copy()
        cond_flag = cond_check(cond)
        if cond_flag == False: return self.df

        if apply is None:
            curr_df[col] = np.nan
        elif isinstance(apply, types.FunctionType):
            try:
                if cond_flag:
                    curr_df[col] = curr_df[cond].apply(apply, axis=axis)
                else:
                    curr_df[col] = curr_df.apply(apply, axis=axis)
            except:
                print("❌ [ERROR] The apply argument insert lambda function")
                print(" example) apply=lambda x: example_function(x.SQ1, x.SQ2)")
                return self.df
        else:
            if cond_flag:
                curr_df.loc[cond, col] = apply
            else:
                curr_df[col] = apply

        self.df = curr_df.copy()
        return self.df[show_cols]

    def count_col(self, col: str, variables: List[str], value: Optional[Union[int, List[int]]] = None, with_cols: Optional[List[str]] = None, cond: Optional[pd.Series] = None) -> pd.DataFrame:
        """
        지정된 열에서 값을 세어 `col`에 저장
        
        Args:
            `col` (str): 카운트를 저장할 열 이름.
            `variables` (List[str]): 값을 셀 열 목록.
            `value` (Union[int, List[int]], optional): 특정 값(들)을 셉니다. 기본값은 None.
            `with_cols` (List[str], optional): 추가로 포함할 열. 기본값은 None.
            `cond` (pd.Series, optional): 조건을 적용할 시리즈. 기본값은 None.
        
        Returns:
            pd.DataFrame: 변경된 데이터프레임.
        """
        show_cols = self.default_show_cols.copy()

        if none_check(col, add_text="column name"): return self.df
        if str_check(col, add_text="column name"): return self.df

        if ma_check(variables, self.cols): return self.df
        ma_cols = ma_return(variables, self.cols)

        if with_cols is not None:
            if ma_check(with_cols, self.cols, len_chk=False): return self.df
            with_cols = ma_return(with_cols, self.cols)
        else:
            with_cols = []

        if value is not None:
            if not isinstance(value, (list, int)):
                print("❌ [ERROR] The value must be list or int")
                return self.df

        cond_flag = cond_check(cond)
        if cond_flag == False: return self.df

        show_cols = sum_list(show_cols, [col], ma_cols, with_cols)

        curr_df = self.df.copy()
        curr_df[col] = np.nan
        if cond_flag:
            cond_idx = list(curr_df[cond].index)
        else:
            cond_idx = list(curr_df.index)

        for idx in cond_idx:
            answers = list(curr_df.loc[idx, ma_cols])

            answer_cnt = 0
            if isinstance(value, int):
                answer_cnt = answers.count(value)
            elif isinstance(value, list):
                for v in value:
                    if v in answers:
                        answer_cnt += 1
            else:
                for a in answers:
                    if not pd.isnull(a) and a != 0:
                        answer_cnt += 1

            curr_df.loc[idx, col] = answer_cnt

        self.df = curr_df.copy()
        return self.df[show_cols]

    def sum_col(self, col: str, variables: List[str], with_cols: Optional[List[str]] = None, cond: Optional[pd.Series] = None) -> pd.DataFrame:
        """
        지정된 열에서 값을 더하여 `col`에 저장합니다.
        
        Args:
            `col` (str): 합계를 저장할 열 이름.
            `variables` (List[str]): 값을 더할 열 목록.
            `with_cols` (List[str], optional): 추가로 포함할 열. 기본값은 None.
            `cond` (pd.Series, optional): 조건을 적용할 시리즈. 기본값은 None.
        
        Returns:
            pd.DataFrame: 변경된 데이터프레임.
        """
        show_cols = self.default_show_cols.copy()

        if none_check(col, add_text="column name"): return self.df
        if str_check(col, add_text="column name"): return self.df

        if ma_check(variables, self.cols): return self.df
        ma_cols = ma_return(variables, self.cols)

        if with_cols is not None:
            if ma_check(with_cols, self.cols, len_chk=False): return self.df
            with_cols = ma_return(with_cols, self.cols)
        else:
            with_cols = []

        cond_flag = cond_check(cond)
        if cond_flag == False: return self.df

        show_cols = sum_list(show_cols, [col], ma_cols, with_cols)

        curr_df = self.df.copy()
        curr_df[col] = np.nan
        if cond_flag:
            curr_df.loc[cond, col] = curr_df[ma_cols].sum(axis=1)
        else:
            curr_df.loc[:, col] = curr_df[ma_cols].sum(axis=1)

        self.df = curr_df.copy()
        return self.df[show_cols]

    def freq(self, qids: Union[str, List[str], Tuple[str]]) -> None:
        """
        지정된 열의 빈도표를 출력합니다.
        
        Args:
            `qids` (Union[str, List[str], Tuple[str]]): 빈도표를 출력할 열(들).
        """
        if isinstance(qids, str):
            print(self.df[qids].value_counts())
        elif isinstance(qids, (list, tuple)):
            ma_cols = ma_return(qids, self.cols)
            freq_list = [(qid, self.df[qid].value_counts()) for qid in ma_cols]
            for qid, fq in freq_list:
                print(f"💠 {qid}")
                print(fq)
                print(self.separator)
                print("")
        else:
            print("❌ [ERROR] Type of qid must be str or list or tuple")

    def crosstabs(self, *qids: str) -> pd.DataFrame:
        """
        교차표를 출력합니다.
        
        Args:
            `qids` (str): 교차표를 출력할 열(들).
        
        Returns:
            pd.DataFrame: 교차표 결과.
        """
        for qid in qids:
            if not isinstance(qid, str):
                print("❌ [ERROR] Type of qid must be str and max 3")

        if not 2 <= len(qids) <= 3:
            print("❌ [ERROR] The qids is atleast 2 and atmost 3")

        if len(qids) == 2:
            return pd.crosstab(self.df[qids[0]], self.df[qids[1]], margins=True)

        if len(qids) == 3:
            return pd.crosstab([self.df[qids[0]], self.df[qids[1]]], self.df[qids[2]], margins=True)

    def safreq(self, sa: Optional[str] = None, cond: Optional[pd.Series] = None, with_cols: Optional[List[str]] = None, only: Optional[Union[range, List, str, int]] = None, df: bool = False, err: bool = False) -> Union[pd.DataFrame, Dict[str, pd.DataFrame], bool]:
        """
        단수 응답(단일 변수) 데이터 체크 메서드.
        
        Args:
            `sa` (str, optional): 단수 응답(단일 변수) 열 이름. 기본값은 None.
            `cond` (pd.Series, optional): 조건을 적용할 시리즈. 기본값은 None.
            `with_cols` (List[str], optional): 추가로 포함할 열. 기본값은 None.
            `only` (Union[range, List, str, int], optional): 특정 값(들)만 확인. 기본값은 None.
            `df` (bool, optional): 데이터프레임을 반환할지 여부. 기본값은 False.
            `err` (bool, optional): 오류를 반환할지 여부. 기본값은 False.
        
        Returns:
            Union[pd.DataFrame, Dict[str, pd.DataFrame], bool]: 결과 데이터프레임 또는 오류 정보.
        """
        show_cols = self.default_show_cols.copy()
        if df_err_check(df, err): return self.df
        
        if sa_check(sa): return self.df

        if with_cols is not None:
            if ma_check(with_cols, self.cols, len_chk=False): return self.df
            with_cols = ma_return(with_cols, self.cols)
        else:
            with_cols = []
        
        curr_df = self.df.copy()
        err_col = self.err_col
        only_col = self.only_col
        curr_df[err_col] = np.nan
        curr_df[err_col] = curr_df[err_col].astype('object')
        curr_df[only_col] = np.nan
        curr_df[only_col] = curr_df[only_col].astype('object')
        
        sa_cols = [err_col, only_col, sa]
        show_cols = sum_list(show_cols, sa_cols, with_cols)
        
        cond_flag = cond_check(cond)        
        if cond_flag == False: return self.df

        only_text = ""
        if only is not None:
            if isinstance(only, range):
                only = list(only)
                only.append(only[-1]+1)
                only_filt = curr_df[sa].isin(only)
                only_text = f"Range {only[0]} THRU {only[-1]}"
            elif isinstance(only, list):
                only_filt = curr_df[sa].isin(only)
                if len(only) > 6:
                    only_text = f"List [{only[0]}, {only[1]}, ... , {only[-2]}, {only[-1]}]"
                else:
                    only_text = f"List {only}"
            elif isinstance(only, (str, int)):
                only_filt = curr_df[sa] == only
                only_text = only
            else:
                print("❌ [ERROR] Type of only must be range or list or str or int")
                return self.df
        
        print_str = f"📢 {sa} MISSING CHECK\n"

        if not cond_flag:
            print_str += "  💠 All base\n"
            ms_chk = list(curr_df[curr_df[sa].isnull()].index)
            if ms_chk:
                curr_df.loc[ms_chk, 'err'] = 'missing'
        else:
            print_str += "  💠 Condition\n"
            ms_chk = list(curr_df[(curr_df[sa].isnull()) & (cond)].index)
            if ms_chk:
                curr_df.loc[ms_chk, err_col] = 'missing'
            over_chk = list(curr_df[(~curr_df[sa].isnull()) & ~(cond)].index)
            if over_chk:
                curr_df.loc[over_chk, err_col] = 'base'
            resp_chk = list(curr_df[cond].index)
            if not resp_chk:
                print_str += "  ❓ No response to this condition\n"

        err_chk = list(curr_df[~curr_df[err_col].isnull()].index)
        if not err_chk:
            print_str += "  ✅ No error\n"
        else:
            print_str += f"  ❌ Error sample count : {len(err_chk)}\n"

        print_str += self.separator

        if only is not None:
            print_str += f"📢 {sa} ANSWER DATA CHECK\n"
            print_str += f"  💠 Answer only in {only_text}\n"
            only_chk = list(curr_df[~only_filt].index)
            if cond_flag:
                only_chk = list(curr_df[(~only_filt) & (cond)].index)
            if not only_chk:
                print_str += "  ✅ Only value check : No error\n"
            else:
                curr_df.loc[only_chk, only_col] = 'chk'
                print_str += f"  ❌ Only Error sample count : {len(only_chk)}\n"
            print_str += self.separator

        err_df = curr_df[ (~curr_df[err_col].isnull()) | (~curr_df[only_col].isnull()) ][show_cols].copy()

        curr_df[err_col] = curr_df[err_col].fillna('')
        curr_df[only_col] = curr_df[only_col].fillna('')
        err_df[err_col] = err_df[err_col].fillna('')
        err_df[only_col] = err_df[only_col].fillna('')

        return_df = curr_df[cond][show_cols] if cond_flag else curr_df[show_cols]
        outputs = df_err_return(df, return_df, err, err_df)
        if isinstance(outputs, bool) and not outputs:
            base = curr_df[cond][sa] if cond_flag else curr_df[sa]
            if not base.dtype == 'object':
                curr_desc = base.describe()
                if not curr_desc.dtype == 'object':
                    print_str += '🧮 Description\n'
                    print_str += f'  - Mean : {round(float(curr_desc["mean"]), 2)}\n'
                    print_str += f'  - Median : {curr_desc["50%"]}\n'
                    print_str += f'  - Max : {curr_desc["max"]}\n'
                    print_str += f'  - Min : {curr_desc["min"]}\n'
            print(print_str)
        else:
            return outputs

    def mafreq(self, ma: Union[List[str], Tuple[str]], cond: Optional[pd.Series] = None, with_cols: Optional[List[str]] = None, atleast: int = 1, atmost: Optional[int] = None, exactly: Optional[int] = None, df: bool = False, err: bool = False) -> Union[pd.DataFrame, Dict[str, pd.DataFrame], bool]:
        """
        복수 응답 데이터 체크 메서드.
        
        Args:
            `ma` (Union[List[str], Tuple[str]]): 복수 응답 열 변수명.
            `cond` (pd.Series, optional): 조건을 적용할 시리즈. 기본값은 None.
            `with_cols` (List[str], optional): 추가로 포함할 열. 기본값은 None.
            `atleast` (int, optional): 최소 응답 수. 기본값은 1.
            `atmost` (int, optional): 최대 응답 수. 기본값은 None.
            `exactly` (int, optional): 정확한 응답 수. 기본값은 None.
            `df` (bool, optional): 데이터프레임을 반환할지 여부. 기본값은 False.
            `err` (bool, optional): 오류를 반환할지 여부. 기본값은 False.
        
        Returns:
            Union[pd.DataFrame, Dict[str, pd.DataFrame], bool]: 결과 데이터프레임 또는 오류 정보.
        """
        show_cols = self.default_show_cols.copy()
        if df_err_check(df, err): return self.df
        
        if ma_check(ma, self.cols): return self.df
        ma_cols = ma_return(ma, self.cols)

        if with_cols is not None:
            if ma_check(with_cols, self.cols, len_chk=False): return self.df
            with_cols = ma_return(with_cols, self.cols)
        else:
            with_cols = []
        
        cnt_col = self.count_label

        curr_df = self.df.copy()
        
        err_col = self.err_col
        curr_df[err_col] = np.nan
        curr_df[err_col] = curr_df[err_col].astype('object')
        
        cond_flag = cond_check(cond)
        if cond_flag == False: return self.df

        err_col = self.err_col
        show_cols = sum_list(show_cols, [cnt_col, err_col], ma_cols, with_cols)
        
        for idx in list(curr_df.index):
            cnt = 0 
            values = list(curr_df.loc[idx, ma_cols])
            for v in values:
                if not pd.isnull(v) and v != 0:
                    cnt += 1
            curr_df.loc[idx, cnt_col] = cnt
        
        count_list = [
            {
                "type": "atleast",
                "value": atleast,
                "cond": curr_df[cnt_col] < atleast
            },
            {
                "type": "atmost",
                "value": atmost,
                "cond": curr_df[cnt_col] > atmost
            },
            {
                "type": "exactly",
                "value": exactly,
                "cond": curr_df[cnt_col] != exactly
            }
        ]
        
        for ck in count_list:
            if int_check(ck["value"], add_text=ck["type"]): return self.df
                
        start = ma[0]
        end = ma[-1]
        
        print_str = f"📢 '{start} - {end}' Answer Check\n"
        
        if not cond_flag:
            print_str += "  💠 All base\n"
        else:
            print_str += "  💠 Condition\n"
            
        for item in count_list:
            check_value = item["value"]
            if check_value is not None:
                check_type = item["type"]
                if not cond_flag:
                    err_chk = list(curr_df[item["cond"]].index)
                    if err_chk:
                        curr_df.loc[err_chk, err_col] = check_type
                else:
                    err_chk = list(curr_df[(item["cond"]) & (cond)].index)
                    if err_chk:
                        curr_df.loc[err_chk, err_col] = check_type
                        
                print_str += f"  💠 The {check_type} error check ({check_type} = {check_value})\n"
        
        if cond_flag:
            err_chk = list(curr_df[(curr_df[cnt_col] > 0) & ~(cond)].index)
            if err_chk:
                curr_df.loc[err_chk, err_col] = 'base'
            resp_chk = list(curr_df[cond].index)
            if not resp_chk:
                print_str += "  ❓ No response to this condition\n"
        
        err_chk = list(curr_df[~curr_df[err_col].isnull()].index)
        if err_chk:
            print_str += f"  ❌ Error sample count : {len(err_chk)}\n"
        else:
            base = curr_df[cond][cnt_col] if cond_flag else curr_df[cnt_col]
            desc = base.describe()
            print_str += "  🧮 Description\n"
            print_str += f"    - Mean Count : {round(float(desc['mean']), 2)}\n"
            print_str += f"    - Max  Count : {desc['max']}\n"
            print_str += f"    - Min  Count : {desc['min']}\n"
            print_str += f"  ✅ No error\n"
        print_str += self.separator

        err_df = curr_df[~curr_df[err_col].isnull()][show_cols].copy()
        curr_df[err_col] = curr_df[err_col].fillna('')
        err_df[err_col] = err_df[err_col].fillna('')
        
        return_df = curr_df[cond][show_cols] if cond_flag else curr_df[show_cols]
        outputs = df_err_return(df, return_df, err, err_df)
        if isinstance(outputs, bool) and not outputs:
            print(print_str)
        else:
            return outputs

    def dupchk(self, ma: Union[List[str], Tuple[str]], with_cols: Optional[List[str]] = None, okUnique: Optional[Union[List, range, int, str]] = None, df: bool = False, err: bool = False) -> Union[pd.DataFrame, Dict[str, pd.DataFrame], bool]:
        """
        중복 응답 데이터 체크 메서드 (순위 응답)
        
        Args:
            `ma` (Union[List[str], Tuple[str]]): 중복을 체크할 열들.
            `with_cols` (List[str], optional): 추가로 포함할 열. 기본값은 None.
            `okUnique` (Union[List, range, int, str], optional): 무시할 특정 값(들). 기본값은 None.
            `df` (bool, optional): 데이터프레임을 반환할지 여부. 기본값은 False.
            `err` (bool, optional): 오류를 반환할지 여부. 기본값은 False.
        
        Returns:
            Union[pd.DataFrame, Dict[str, pd.DataFrame], bool]: 결과 데이터프레임 또는 오류 정보.
        """
        show_cols = self.default_show_cols.copy()
        if df_err_check(df, err): return self.df
        
        if ma_check(ma, self.cols): return self.df
        rk_cols = ma_return(ma, self.cols)

        if with_cols is not None:
            if ma_check(with_cols, self.cols, len_chk=False): return self.df
            with_cols = ma_return(with_cols, self.cols)
        else:
            with_cols = []
            
        if okUnique is not None:
            if not isinstance(okUnique, (list, range, int, str)):
                print("❌ [ERROR] Type of okUnique must be list or int or str")
                return self.df
            if isinstance(okUnique, range):
                okUnique = list(okUnique)
                okUnique.append(okUnique[-1] + 1)
            elif isinstance(okUnique, (int, str)):
                okUnique = [okUnique]
        else:
            okUnique = []

        dup_df = self.df.copy()
        raw_index = list(dup_df.index)
        
        dup_col = 'dupchk'
        dup_df[dup_col] = np.nan
        dup_df[dup_col] = dup_df[dup_col].astype('object')
        
        show_cols = sum_list(show_cols, [dup_col], rk_cols, with_cols)
        
        for idx in raw_index:
            r = dup_df.loc[idx, rk_cols]
            answers = list(r)
            answers = [x for x in answers if x is not pd.NA]
            dup_del = set(answers)

            dup_values = []
            for dup in dup_del:
                if not pd.isnull(dup) and dup not in okUnique:
                    cnt = answers.count(dup)
                    if cnt > 1:
                        dup_values.append(dup)
            
            dup_df.loc[idx, [dup_col]] = str(dup_values) if dup_values else np.nan
        
        check_row = dup_df[~dup_df[dup_col].isnull()]
        check_row_index = list(check_row.index)
        
        print_str = f"📢 '{rk_cols[0]} - {rk_cols[-1]}' Duplicated value check\n"
        print_str += f"  💠 okUnique = {okUnique}\n"
        if not check_row_index:
            print_str += "  ✅ Answer is not duplicated\n"
        else:
            print_str += f"  ❌ Error sample count : {len(check_row_index)}\n"
        print_str += self.separator

        dup_df[dup_col] = dup_df[dup_col].fillna('')
        
        outputs = df_err_return(df, dup_df[show_cols], err, check_row[show_cols])
        if isinstance(outputs, bool) and not outputs:
            print(print_str)
        else:
            return outputs

    def logchk(self, base_cond: pd.Series, answer_cond: pd.Series, with_cols: Optional[List[str]] = None, df: bool = False, err: bool = False) -> Union[pd.DataFrame, Dict[str, pd.DataFrame], bool]:
        """
        특정 로직에 대한 응답 체크
        (`base_cond`가 `True`일 때, `answer_cond`도 `True`)
        
        Args:
            `base_cond` (pd.Series): 베이스 조건.
            `answer_cond` (pd.Series): 베이스 조건이 True일 때 응답 조건.
            `with_cols` (List[str], optional): 추가로 포함할 열. 기본값은 None.
            `df` (bool, optional): 데이터프레임을 반환할지 여부. 기본값은 False.
            `err` (bool, optional): 오류를 반환할지 여부. 기본값은 False.
        
        Returns:
            Union[pd.DataFrame, Dict[str, pd.DataFrame], bool]: 결과 데이터프레임 또는 오류 정보.
        """
        show_cols = self.default_show_cols.copy()
        if df_err_check(df, err): return self.df

        input_flag = cond_check(base_cond, 'base_cond') 
        output_flag = cond_check(answer_cond, 'answer_cond')
        if input_flag == False or output_flag == False: return self.df
        
        if with_cols is not None:
            if ma_check(with_cols, self.cols, len_chk=False): return self.df
            with_cols = ma_return(with_cols, self.cols)
        else:
            with_cols = []
            
        curr_df = self.df.copy()
        err_col = self.err_col
        curr_df[err_col] = np.nan
        curr_df[err_col] = curr_df[err_col].astype('object')
    
        check_index = list(curr_df[(base_cond) & ~(answer_cond)].index)
        
        show_cols = sum_list(show_cols, [err_col], with_cols)
        print_str = "📢 Punching Logic Check\n"

        resp_chk = list(curr_df[base_cond].index)
        if not resp_chk:
            print_str += "  ❓ No response to this condition\n"
        else:
            if len(check_index) == 0:
                print_str += f"  ✅ Punching Logic correct\n"
            else:
                curr_df.loc[check_index, err_col] = 'chk'
                print_str += f"  ❌ [ERROR] Punching Logic Error\n"
                print_str += f"  ❌ Error sample count : {len(check_index)}\n"
        print_str += self.separator
        err_df = curr_df[~curr_df[err_col].isnull()].copy()
        err_df[err_col] = err_df[err_col].fillna('')
        curr_df[err_col] = curr_df[err_col].fillna('')

        outputs = df_err_return(df, curr_df[base_cond][show_cols], err, err_df[show_cols])
        if isinstance(outputs, bool) and not outputs:
            print(print_str)
        else:
            return outputs

    def masa(self, ma: Union[List[str], Tuple[str]], sa: str, cond: Optional[pd.Series] = None, diff_value: Optional[Union[List, int]] = None, with_cols: Optional[List[str]] = None, df: bool = False, err: bool = False) -> Union[pd.DataFrame, Dict[str, pd.DataFrame], bool]:
        """
        `복수 응답`을 베이스로 하는 `단수 응답` 로직 체크.
         
        Args:
            `ma` (Union[List[str], Tuple[str]]): 복수 응답 열 목록.
            `sa` (str): 단수 응답 열.
            `cond` (pd.Series, optional): 조건을 적용할 시리즈. 기본값은 None.
            `diff_value` (Union[List, int], optional): 무시할 특정 값(들). 기본값은 None.
            `with_cols` (List[str], optional): 추가로 포함할 열. 기본값은 None.
            `df` (bool, optional): 데이터프레임을 반환할지 여부. 기본값은 False.
            `err` (bool, optional): 오류를 반환할지 여부. 기본값은 False.
        
        Returns:
            Union[pd.DataFrame, Dict[str, pd.DataFrame], bool]: 결과 데이터프레임 또는 오류 정보.
        """
        show_cols = self.default_show_cols.copy()
        if df_err_check(df, err): return self.df

        if ma_check(ma, self.cols): return self.df
        ma_cols = ma_return(ma, self.cols)
        
        key_id = key_id_check(ma_cols, sa, "SA")
        if not key_id["ok"]: return self.df

        ma_qid = key_id["return"]

        if sa_check(sa): return self.df

        cond_flag = cond_check(cond)
        if cond_flag == False: return self.df

        if with_cols is not None:
            if ma_check(with_cols, self.cols, len_chk=False): return self.df
            with_cols = ma_return(with_cols, self.cols)
        else:
            with_cols = []

        masa_cols = [sa] + ma_cols
        ms_col = self.masa_label
        exist = self.exist_col
        ma_base = self.ma_base

        show_cols = sum_list(show_cols, [ms_col, exist, ma_base], masa_cols, with_cols)
        curr_df = self.df.copy()
        if cond_flag:
            curr_df = self.df[cond].copy()

        curr_df[ms_col] = np.nan
        curr_df[ms_col] = curr_df[ms_col].astype('object')
        curr_df[exist] = np.nan
        curr_df[exist] = curr_df[exist].astype('object')
        curr_df[ma_base] = np.nan
        curr_df[ma_base] = curr_df[ma_base].astype('object')

        filt_df = curr_df[~curr_df[sa].isnull()].copy()
        filt_index = list(filt_df.index)

        print_str = "📢 Multi variable base Single variable Logic Check\n"

        if not list(curr_df.index):
            print_str += "❓ No response to this condition\n"
            print(print_str)
            return self.df

        print_str += f"  💠 SA : {sa}\n"
        print_str += f"  💠 MA : {ma_cols[0]} - {ma_cols[-1]} ({len(ma_cols)} columns)\n"

        diff_list = list_or_int_set(diff_value)
        if diff_list:
            print_str += f"  ❗ Do not check the code : {diff_list}\n"

        for idx in filt_index:
            v = int(filt_df.loc[idx, sa])
            base = f"{ma_qid}{v}"
            curr_base = [col.replace(ma_qid, '') for col in ma_cols if not pd.isnull(filt_df.loc[idx, col]) and filt_df.loc[idx, col] != 0]
            filt_df.loc[idx, ma_base] = str(curr_base)

            if v in diff_list:
                continue

            if base in ma_cols:
                base_v = filt_df.loc[idx, base]
                if pd.isnull(base_v) or base_v == 0:
                    filt_df.loc[idx, ms_col] = base
            else:
                filt_df.loc[idx, exist] = base

        err_index = list(filt_df[~filt_df[ms_col].isnull()].index)
        exist_check = list(filt_df[~filt_df[exist].isnull()].index)

        err_df = filt_df[~(filt_df[ms_col].isnull()) | ~(filt_df[exist].isnull())][show_cols]
        if err_index or exist_check:
            if err_index:
                print_str += f"  ❌ [ERROR] MA-SA Logic Error\n"
                print_str += f"  ❌ Error sample count : {len(err_index)}\n"
            if exist_check:
                print_str += f"   ❗ [WARNING] Exist Variable Error\n"
                print_str += f"  ❌ Error sample count : {len(exist_check)}\n"
        else:
            print_str += "  ✅ Logic correct\n"
        print_str += self.separator

        filt_df[ms_col] = filt_df[ms_col].fillna('')
        filt_df[exist] = filt_df[exist].fillna('')
        err_df[ms_col] = err_df[ms_col].fillna('')
        err_df[exist] = err_df[exist].fillna('')

        outputs = df_err_return(df, filt_df[show_cols], err, err_df)
        if isinstance(outputs, bool) and not outputs:
            print(print_str)
        else:
            return outputs

    def mama(self, base_ma: Union[List[str], Tuple[str]], ma: Union[List[str], Tuple[str]], cond: Optional[pd.Series] = None, diff_value: Optional[Union[List, int]] = None, with_cols: Optional[List[str]] = None, df: bool = False, err: bool = False) -> Union[pd.DataFrame, Dict[str, pd.DataFrame], bool]:
        """
        `복수 응답`을 베이스로 하는 `복수 응답` 로직 체크.
        
        Args:
            `base_ma` (Union[List[str], Tuple[str]]): 기준이 되는 복수 응답 열 목록.
            `ma` (Union[List[str], Tuple[str]]): 체크할 복수 응답 열 목록.
            `cond` (pd.Series, optional): 조건을 적용할 시리즈. 기본값은 None.
            `diff_value` (Union[List, int], optional): 무시할 특정 값(들). 기본값은 None.
            `with_cols` (List[str], optional): 추가로 포함할 열. 기본값은 None.
            `df` (bool, optional): 데이터프레임을 반환할지 여부. 기본값은 False.
            `err` (bool, optional): 오류를 반환할지 여부. 기본값은 False.
        
        Returns:
            Union[pd.DataFrame, Dict[str, pd.DataFrame], bool]: 결과 데이터프레임 또는 오류 정보.
        """
        show_cols = self.default_show_cols.copy()
        if df_err_check(df, err): return self.df

        if ma_check(base_ma, self.cols): return self.df
        base_ma_cols = ma_return(base_ma, self.cols)

        if ma_check(ma, self.cols): return self.df
        ma_cols = ma_return(ma, self.cols)

        cond_flag = cond_check(cond)
        if cond_flag == False: return self.df

        diff_list = list_or_int_set(diff_value)
        if with_cols is not None:
            if ma_check(with_cols, self.cols, len_chk=False): return self.df
            with_cols = ma_return(with_cols, self.cols)
        else:
            with_cols = []

        base_key_id = key_id_check(base_ma_cols, ma_cols, "MA")
        if not base_key_id["ok"]: return self.df

        base_ma_qid = base_key_id["return"]

        key_id = key_id_check(ma_cols, ma_cols, "MA")
        if not key_id["ok"]: return self.df

        ma_qid = key_id["return"]

        cols_order = []
        for ma_col in ma_cols:
            cols_order.append(ma_col)
            switch_qid = ma_col.replace(ma_qid, base_ma_qid)
            if switch_qid in base_ma_cols:
                cols_order.append(switch_qid)

        ms_col = self.masa_label
        exist = self.exist_col
        ma_base = self.ma_base
        ma_answer = self.ma_answer

        show_cols = sum_list(show_cols, [ms_col, exist, ma_base, ma_answer], cols_order, with_cols)
        curr_df = self.df.copy()
        if cond_flag:
            curr_df = self.df[cond].copy()
        curr_df[ms_col] = np.nan
        curr_df[ms_col] = curr_df[ms_col].astype('object')
        curr_df[exist] = np.nan
        curr_df[exist] = curr_df[exist].astype('object')
        curr_df[ma_base] = np.nan
        curr_df[ma_base] = curr_df[ma_base].astype('object')
        curr_df[ma_answer] = np.nan
        curr_df[ma_answer] = curr_df[ma_answer].astype('object')

        filt_df = curr_df[~(curr_df[ma_cols].isnull()).all(axis=1)].copy()
        filt_index = list(filt_df.index)

        print_str = "📢 Multi variable base Multi variable Logic Check\n"
        
        if not list(curr_df.index):
            print_str += "❓ No response to this condition\n"
            print(print_str)
            return self.df
        
        print_str += f"  💠 MA : {ma_cols[0]} - {ma_cols[-1]} ({len(ma_cols)} columns)\n"
        print_str += f"  💠 MA : {base_ma_cols[0]} - {base_ma_cols[-1]} ({len(base_ma_cols)} columns)\n"

        if diff_list:
            print_str += f"  ❗ Do not check the code : {diff_list}\n"

        for idx in filt_index:
            curr_base = [col.replace(base_ma_qid, '') for col in base_ma_cols if not pd.isnull(filt_df.loc[idx, col]) and filt_df.loc[idx, col] != 0]
            answers = [col.replace(ma_qid, '') for col in ma_cols if not pd.isnull(filt_df.loc[idx, col]) and filt_df.loc[idx, col] != 0]
            filt_df.loc[idx, ma_base] = str(curr_base)
            filt_df.loc[idx, ma_answer] = str(answers)
            
            err_vars = []
            exist_vars = []
            for answer in answers:
                if int(answer) in diff_list:
                    continue
                if not pd.isnull(answer):
                    v = int(answer)
                    base_id = base_key_id["return"]
                    base = f"{base_id}{v}"
                    if base in base_ma_cols:
                        base_v = filt_df.loc[idx, base]
                        if pd.isnull(base_v) or base_v == 0:
                            err_vars.append(base)
                    else:
                        exist_vars.append(base)
                        
            if err_vars:
                filt_df.loc[idx, ms_col] = str(err_vars)
            if exist_vars:
                filt_df.loc[idx, exist] = str(exist_vars)

        err_index = list(filt_df[~filt_df[ms_col].isnull()].index)
        exist_check = list(filt_df[~filt_df[exist].isnull()].index)

        err_df = filt_df[~(filt_df[ms_col].isnull()) | ~(filt_df[exist].isnull())][show_cols]

        if err_index or exist_check:
            if err_index:
                print_str += f"  ❌ [ERROR] MA-MA Logic Error\n"
                print_str += f"  ❌ Error sample count : {len(err_index)}\n"
            if exist_check:
                print_str += f"   ❗ [WARNING] Exist Variable Error\n"
                print_str += f"  ❌ Error sample count : {len(exist_check)}\n"
        else:
            print_str += "  ✅ Logic correct\n"
        print_str += self.separator

        filt_df[ms_col] = filt_df[ms_col].fillna('')
        filt_df[exist] = filt_df[exist].fillna('')
        err_df[ms_col] = err_df[ms_col].fillna('')
        err_df[exist] = err_df[exist].fillna('')

        outputs = df_err_return(df, filt_df[show_cols], err, err_df)
        if isinstance(outputs, bool) and not outputs:
            print(print_str)
        else:
            return outputs


def Setting(pid: str, 
            mode: str = 'auto', 
            cond: Optional[str] = None,
            use_variable: bool = False,
            key: str = api_key, 
            server: str = api_server, 
            json_export: bool = True, 
            data_layout: bool = False, 
            base_layout: str = 'DoNotDelete',
            datamap_name: str = 'Datamap',
            mkdir: bool = False,
            dir_name: Optional[str] = None) -> None:

    """
    데이터 체크 노트북 파일 및 데이터 세팅
    
    Args:
        `pid` (str): 프로젝트 ID.
        `mode` (str, optional): 모드 ('auto' 또는 'file'). 기본값은 'auto'.
        `cond` (str, optional): 데이터 필터링 조건. 기본값은 None.
        `use_variable` (bool, optional): 변수 파일 사용 여부. 기본값은 False.
        `key` (str, optional): API 키. 기본값은 api_key.
        `server` (str, optional): API 서버. 기본값은 api_server.
        `json_export` (bool, optional): JSON 내보내기 여부. 기본값은 True.
        `data_layout` (bool, optional): 데이터 레이아웃 내보내기 여부. 기본값은 False.
        `base_layout` (str, optional): 기본 레이아웃 이름. 기본값은 'DoNotDelete'.
        `datamap_name` (str, optional): 데이터 맵 이름. 기본값은 'Datamap'.
        `mkdir` (bool, optional): 디렉토리 생성 여부. 기본값은 False.
        `dir_name` (str, optional): 디렉토리 이름. 기본값은 None.
    """

    #pd.io.formats.excel.ExcelFormatter.header_style = None
    excel.ExcelFormatter.header_style = None
    
    if pid == '' or not pid :
        print('❌ Please enter pid')
        return

    if not mode in ['auto', 'file'] :
        print('❌ Please check the mode argument (auto or file)')
        return

    parent_path = os.getcwd()
    if mkdir :
        folder_name = pid
        if dir_name != None :
            folder_name = dir_name
        parent_path =  os.path.join(parent_path, folder_name)
        chk_mkdir(parent_path)

    if mode == 'file' :
        file_name = f'{pid}.xlsx'
        xl = openpyxl.load_workbook(file_name)
        map_sheet = datamap_name
        data_map = xl[map_sheet]
        print('📢 Read excel file (xlsx)')

    if mode == 'auto' :
        file_name = f'{pid}.csv'
        if cond != None :
            if cond.isdigit() :
                print('❌ [ERROR] : The cond argument can only be a string')
                return
        delivery_cond = 'qualified' if cond == None else f'qualified and {cond}'
        try :
            api.login(key, server)
        except :
            print('❌ Error : Decipher api login failed')
            return

        path = f'surveys/selfserve/548/{pid}'
        # get csv data
        try :
            csv_data = api.get(f'{path}/data', format='csv', cond=delivery_cond)
        except :
            print('❌ Error : Please check the cond argument')
            return

        csv_binary = f'binary_{pid}.csv'
        create_binary_file(parent_path, csv_binary, csv_data)
        create_ascii_file(parent_path, csv_binary, f'{pid}.csv')
        
        time.sleep(3)

        # get datamap xlsx
        map_xlsx = api.get(f'{path}/datamap', format='xlsx')
        create_binary_file(parent_path, f'mapsheet_{pid}.xlsx', map_xlsx)

        xl = openpyxl.load_workbook(os.path.join(parent_path, f'mapsheet_{pid}.xlsx'))
        map_sheet = 'datamap'
        data_map = xl[map_sheet]

        print('📢 Using Decipher REST API')

    mx_row = data_map.max_row
    mx_col = data_map.max_column

    key_ids = key_vars
    diff_vars = sys_vars
    all_diff = key_ids + diff_vars

    rank_chk = ['1순위', '2순위', '1st', '2nd']

    na = 'noanswer'
    eltxt = 'element'
    col_name = ["a", "b", "c"]
    curr_var = {col:[] for col in col_name }

    variables = []
    
    #print("  ❌ DataCheck Setting Start")
    for row in range(1, mx_row+1) :
        if not_empty_cell(data_map, row) :
            for idx, col in enumerate(range(1, mx_col+1)) :
                curr_col = col_name[idx]
                curr_dict = curr_var[curr_col]
                cell = data_map.cell(row, col)
                if cell.value or cell.value == 0: 
                    curr_dict.append(cell.value)
                    curr_var[curr_col] = curr_dict
                if cell.value == None or cell.value == "" :
                    curr_dict.append("")
                    curr_var[curr_col] = curr_dict
        else :
            variables.append(curr_var)
            curr_var = {col:[] for col in col_name }


    qids = OrderedDict()
    for variable in variables :
        # qid, type summary
        a_cell = variable['a']
        a_cell = [a for a in a_cell if a != '' and a != None]
        b_cell = variable['b']
        #b_cell = [b for b in b_cell if b != '' and b != None]
        c_cell = variable['c']
        #c_cell = [c for c in c_cell if c != '' and c != None]
        qid = a_cell[0] # qid

        # print(qid)
        # print(b_cell)
        # print(c_cell)

        type_chk = a_cell[1] if len(a_cell) >= 2 else None # type check

        # attribute
        main_qlabel = None
        qtype = None
        qelements = []
        qvalue = None
        qtitle = None

        # labels setting
        qlabes = {
            'values' : {b:c_cell[idx] for idx, b in enumerate(b_cell) if type(b) == int}  if c_cell else {},
            'texts' : {re_big(b):c_cell[idx] for idx, b in enumerate(b_cell) if type(b) == str or b == None} if c_cell else {}
        }
        #  find name in []

        # main q label check
        find_qname = re_big(qid.split(':')[0])
        if find_qname :
            main_qlabel = find_qname
            qelements.append(main_qlabel)
            qtitle = colon_split(qid, 1)
        else :
            main_qlabel = colon_split(qid, 0)
            qtitle = colon_split(qid, 1)

        # type check
        if type_chk :
            open_text = 'OPEN'
            if open_text.upper() in type_chk.upper() :
                qtype = 'OE'
            else :
                qtype = 'CE'

            # value check
            value_text = 'VALUES'
            if value_text.upper() in type_chk.upper() :
                qvalue = colon_split(type_chk, 1)
            else :
                qvalue = None


        else :
            qtype = 'OTHER'

        # other oe check
        oe_chk = 'oe'
        if oe_chk in main_qlabel :
            qtype = 'OTHER_OE'

        # elements setting
        for b in b_cell :
            b_chk = re_big(str(b))
            if b_chk :
                qelements.append(b_chk)

        if not 'OE' in qtype:
            # ma check
            unchk = 'Unchecked'
            c_chk = [c.upper() for c in c_cell if type(c) == str]
            if unchk.upper() in c_chk : 
                qtype = 'MA'
            else :
                # radio/number check
                int_chk = [b for b in b_cell if type(b) == int]
                if int_chk :
                    qtype = 'SA'
                else :
                    qtype = 'NUM'

        if len(qelements) >= 2 :
            if main_qlabel in qelements :
                qelements.remove(main_qlabel)

        el_labels = {key:value for key, value in qlabes['texts'].items() if key}
        
        qids[main_qlabel] = {
            'element' : qelements,
            'title' : qtitle,
            'type' : qtype,
            'value' : qvalue,
            'value_label' : qlabes['values'],
            'element_label' : el_labels,
        }

    if na in list(qids.keys()) :
        nas = qids[na]
        els = nas[eltxt]
        for el in els :
            na_name = el.split('_')[:-1]
            na_el = '_'.join(na_name).replace(na, '')
            if not na_el == '' :
                if qids[na_el] :
                    qel = qids[na_el][eltxt]
                    qel.append(el)
                    qids[na_el][eltxt] = qel

    # print(qids)

    # default setting
    nb = nbf.v4.new_notebook()
    ipynb_file_name = f'DataCheck_{pid}.ipynb'
    order_qid = list(qids.items())

    # json export
    if json_export :
        with open(os.path.join(parent_path, f'map_{pid}.json'), 'w', encoding='utf-8') as f :
            json.dump(qids, f, ensure_ascii=False, indent=4)

    # print(qids)
    # data layout export
    if data_layout :
        try :
            api.login(key, server)
            survey = f'selfserve/548/{pid}'
            url = f'surveys/{survey}/layouts'
            map = api.get(url)
        except :
            print('❌ Error : Decipher API failed')
        
        maps = [m for m in map if m['description'] == base_layout ]
        if not maps :
            print(f'❌ Error : The base layout({base_layout}) is null')
            return 
        base_map = maps[0]

        variables = base_map['variables']
        # print(variables)
        exactly_diff_vars = key_vars + sys_vars
        ce_vars = ['radio', 'checkbox', 'number', 'float', 'select']
        oe_vars = ['text', 'textarea']
        diff_label_names = ['vqtable', 'voqtable', 'dummy', 'DUMMY', 'Dummmy']
        
        ce = open(os.path.join(parent_path, f'CE_{pid}.txt'), 'w')
        oe = open(os.path.join(parent_path, f'OE_{pid}.txt'), 'w')

        for label, width in [ ('record', 7), ('uuid', 16) ]:
            write_text = f'{label},{label},{width}\n'
            ce.write(write_text)
            oe.write(write_text)

        resp_chk = [v for v in variables if v['label'] == 'RespStatus']
        if resp_chk :
            ce.write(f'RespStatus,RespStatus,8\n')

        for var in variables :
            label = var['label']
            qlabel = var['qlabel']
            qtype = var['qtype']
            fwidth = var['fwidth']
            altlabel = var['altlabel']
            shown = var['shown']
            if not shown :
                continue

            write_text = f'{label},{altlabel},{fwidth}\n'
            if (not label in exactly_diff_vars and not qlabel in exactly_diff_vars) :
                if [dl for dl in diff_label_names if (dl in label) or (dl in qlabel)] :
                    continue
                if qtype in ce_vars :
                    if qtype in ['number', 'float'] :
                        verify_check = [attr['value'].split('-')[1] for ql, attr in list(qids.items()) if (ql == qlabel) or (ql == label)]
                        if verify_check :
                            max_width = len(verify_check[0])
                                # print(label, verify_check, max_width)
                            if qtype == 'float' :
                                max_width += 4
                            write_text = f'{label},{altlabel},{max_width}\n'
                    ce.write(write_text)
                if qtype in oe_vars :
                    oe.write(write_text)

        oe.write(f'decLang,decLang,60\n')
        # if resp_chk :
        #     oe.write(f'RespStatus,RespStatus,8\n')

        ce.close()
        oe.close()

    # variable py file create
    variable_py_name = f'variables_{pid}.py'
    py_file = open(os.path.join(parent_path, variable_py_name), 'w')
    py_file.write(f'# {pid} variables\n')

    ipynb_cell = []


    # set_file_name = 'pd.read_excel(file_name)' if mode == 'file' else 'pd.read_csv(file_name, low_memory=False)'

    default = f'''import pandas as pd
import numpy as np
from variables_{pid} import * 

file_name = "{pid}.xlsx"
df = DataCheck(pd.read_excel(file_name, engine="openpyxl"), keyid="record")
comp = (df.status == 3)
'''

    ipynb_cell.append(nbf.v4.new_code_cell(default))
#     ipynb_cell.append(nbf.v4.new_code_cell('''lambda_count = lambda x: x.count() - (x==0).sum()

# def lp(print_word) :
#     print(f'🟢 {print_word}')

# def lchk() :
#     print(f'-------- 🔽 LIVE CHECK 🔽--------')

# def ep(err_df, err_qid) :
#     if len(err_df) >= 1 :
#         print(f'❌ {err_qid} has Error')
#     else :
#         print(f'✅ {err_qid} is OK')'''))


    # # Additional functions
    # functions = '''def scale_datacheck(qid, cond=None) :
    # print(f'📌 {qid} Check')
    # base = eval(qid)
    # print(f'- Variables : {base}')
    # print(f'- Length : {len(base)}')
    # err_flag = False
    # for x in base :
    #     err = df.safreq(x, cond=cond, err=True)
    #     err_idx = list(err.index)
    #     if err_idx :
    #         print(f'❌ {x} has Error')
    #         err_flag = True

    # if not err_flag :
    #     print(f'✅ {qid} confirm')
    # '''

    # ipynb_cell.append(nbf.v4.new_code_cell(functions))

    # qids setting
    for idx, attrs in enumerate(order_qid) :
        qid = attrs[0]
        els = attrs[1]
        if not qid in all_diff :
            qels = els['element']
            qtype = els['type']
            qval = els['value']
            qtitle = els['title']
            val_label = els['value_label']
            el_label = els['element_label']

            cell_texts = []
            cell_texts.append(f'# {qid} : {qtype}')
            # sa check #
            if qtype == 'SA' or (qtype == 'MA' and len(qels) == 1):
                if qtype == 'SA' :
                    if qval :
                        val_chk = f"# value : {qval}"
                        cell_texts.append(val_chk)

                    if len(qels) >= 2 :
                        diff_na = [q for q in qels if not na in q]
                        py_file.write(f"{qid} = {diff_na}\n")

                    for qel in qels :
                        if na in qel :
                            cell_texts.append(f'# The {qid} contains {qel}')
                        else :
                            safreq = f"df.safreq('{qel}')"
                            if use_variable : safreq = f"df.safreq({qel})"

                            py_file.write(f"{qel} = '{qel}'\n")
                            cell_texts.append(safreq)

                    if val_label :
                        values = [v for v in val_label.keys() if not int(v) == 0]
                        py_file.write(f'{qid}_value = {values}\n')
                        #values_txt = f'{qid}_values = {values}'
                        #cell_texts.append(values_txt)

                    # rank check
                    if len(qels) >= 2 :
                        labels = list(el_label.values())
                        rk = []
                        for rk_txt in rank_chk :
                            for label in labels :
                                mu_rk = rk_txt.strip().replace(' ','').upper()
                                mu_label = label.strip().replace(' ','').upper()
                                if mu_rk in mu_label :
                                    rk.append(label)
                        if len(rk) >= 2 :
                            dup_diff_na = [q for q in qels if not na in q]
                            set_qid = f"('{dup_diff_na[0]}', '{dup_diff_na[-1]}')"

                            py_file.write(f"{qid} = {dup_diff_na}\n")
                            #cell_texts.append(f'{qid} = {set_qid}')
                            dupchk = f"df.dupchk({set_qid})"
                            if use_variable : dupchk = f"df.dupchk({qid})"

                            cell_texts.append(dupchk)
                else :
                    if qval :
                        val_chk = f"# value : {qval}"
                        py_file.write(f"{qid} = '{qid}'\n")
                        py_file.write(f'{qid}_value = [0, 1]\n')
                        
                        cell_texts.append(val_chk)
                        safreq = f"df.safreq('{qels[0]}')"
                        if use_variable : safreq = f"df.safreq({qels[0]})"
                        cell_texts.append(safreq)
            ### sa end ###

            # ma check #
            elif qtype == 'MA' :
                if len(qels) > 1 :
                    diff_na = [q for q in qels if not na in q]
                    if not diff_na :
                        continue
                    nas = [q for q in qels if na in q]
                    first_el = diff_na[0]
                    last_el = diff_na[-1]
                    set_qid = f"('{first_el}', '{last_el}')"

                    for q in diff_na :
                        py_file.write(f"{q} = '{q}'\n")

                    py_file.write(f"{qid} = {diff_na}\n")

                    if val_label :
                        values = [v for v in val_label.keys() if not int(v) == 0]
                        if not values == [1] :
                            py_file.write(f'{qid}_value = {values}\n')
                            #values_txt = f'{qid}_values = {values}'
                            #cell_texts.append(values_txt)
                        else :
                            py_file.write(f'{qid}_value = [0, 1]\n')
                    # cell_texts.append(f'{qid} = {set_qid}')

                    mafreq = f"df.mafreq({set_qid})"
                    if use_variable : mafreq = f"df.mafreq({qid})"

                    cell_texts.append(mafreq)

                    if nas :
                        cell_texts.append(f'# The {qid} contains {nas}')
            ### ma end ###


            # num check #
            elif qtype == 'NUM' :
                range_set = None

                if len(qels) >=2 :
                    diff_na = [q for q in qels if not na in q]
                    py_file.write(f"{qid} = {diff_na}\n")

                if qval :
                    values = qval.split('-')
                    range_set = f"only=range({values[0]}, {values[1]})"

                for qel in qels :
                    if na in qel :
                        cell_texts.append(f'# The {qid} contains {qel}')
                    else :
                        if range_set :
                            safreq = f"df.safreq('{qel}', {range_set})"
                            if use_variable : safreq = f"df.safreq({qel}, {range_set})"
                        else :
                            safreq = f"df.safreq('{qel}')"
                            if use_variable : safreq = f"df.safreq({qel})"

                        py_file.write(f"{qel} = '{qel}'\n")
                        cell_texts.append(safreq)

            ### num end ###

            # text check #
            elif qtype == 'OE' :
                if len(qels) >=2 :
                    diff_na = [q for q in qels if not na in q]
                    py_file.write(f"{qid} = {diff_na}\n")

                for qel in qels :
                    if na in qel :
                        cell_texts.append(f'# The {qid} contains {qel}')
                    else :
                        safreq = f"df.safreq('{qel}')"
                        if use_variable : safreq = f"df.safreq({qel})"

                        py_file.write(f"{qel} = '{qel}'\n")
                        cell_texts.append(safreq)
            ### text end ###

            # other open check #
            elif qtype == 'OTHER_OE' :
                for qel in qels :
                    safreq = f"df.safreq('{qel}')"
                    if use_variable : safreq = f"df.safreq({qel})"

                    py_file.write(f"{qel} = '{qel}'\n")
                    cell_texts.append(safreq)
            ### other open end ###


            if cell_texts :
                cell = '\n'.join(cell_texts)
                ipynb_cell.append(nbf.v4.new_code_cell(cell))
            else :
                mark = f'The {qid} not cotains elements'
                ipynb_cell.append(nbf.v4.new_markdown_cell(mark))

            py_file.write(f'\n')

    py_file.close()
    #ipynb_cell
    nb['cells'] = ipynb_cell
    #print(nb)
    ipynb_file_path = os.path.join(parent_path, ipynb_file_name)
    if not os.path.isfile(ipynb_file_path) :
        with open(ipynb_file_path, 'w') as f:
            nbf.write(nb, f)
    else :
        print('❗ The DataCheck ipynb file already exists')
    
    print("✅ DataCheck Setting Complete")


def qset(qname: str, code: Union[List, Tuple]) -> List[str]:
    """
    코드 기준의 변수 리스트 생성
    
    Args:
        qname (str): 변수 구분자 텍스트 ('Q2r', 'Q3b').
        code (Union[List, Tuple]): 생성할 코드 리스트.
    
    Returns:
        List[str]: 코드 기준의 변수 리스트 ('Q2r1', 'Q2r2', 'Q2r3').
    """
    if not isinstance(code, (list, tuple)):
        print('❌ Code should only be entered as a list or tuple type.')
        return 
    return list(map(lambda x: f'{qname}{x}', code))