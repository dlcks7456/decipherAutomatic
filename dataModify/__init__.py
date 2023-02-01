from decipher.beacon import api
import pandas as pd
from datetime import datetime
from dateutil.relativedelta import relativedelta
from ..key import api_key, api_server
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Alignment, Color, fills, Side
import os
from dataclasses import dataclass
from decipherAutomatic.getFiles import *

def find_cells(
        file_name,
        key_variable = 'record', 
        start_row_num = 1, 
        start_col_num = 1, 
        find_text=['수정', '삭제'], 
        mkdir=False) : 
    
    if file_name == '' or file_name == None :
        print('❌ Please check the file_name')
        return
    
    if key_variable == '' or file_name == None :
        print('❌ Please check the key_variable')
        return
    
    if type(start_row_num) != int :
        print('❌ The start_row_num arg is a number only')
        return

    if start_row_num == None :
        print('❌ Please check the start_row_num')
        return

    if type(start_col_num) != int :
        print('❌ The start_col_num arg is a number only')
        return

    if start_col_num == None :
        print('❌ Please check the start_col_num')
        return

    if type(mkdir) != bool :
        print('❌ The mkdir arg is a boolean only')
        return

    # make files
    now = datetime.now()
    year = now.year
    month = '{:02}'.format(now.month)
    day = '{:02}'.format(now.day)
    file_verison = 1

    # find high light cell
    wb = load_workbook(file_name, data_only=True)
    st = wb[wb.sheetnames[0]]
    wb_rows = range(1, st.max_row+1)
    wb_cols = range(1, st.max_column+1)

    set_rows = [start_row_num] # 0 is name of variables 
    set_cols = [start_col_num]

    for row in wb_rows :
        # find not empty cells
        if row > 1 :
            first_col_cell = st.cell(row=row, column=start_col_num)
            first_value = first_col_cell.value
            if first_value != None :
                if type(first_value) == str :
                    if first_value.replace(' ', '').upper() in find_text :
                        set_rows.append(row)
        
        for col in wb_cols :
            curr_cell = st.cell(row=row, column=col)
            if not curr_cell.fill.start_color.index == '00000000' :
                set_rows.append(row)
                set_cols.append(col)
            
            if curr_cell.value == key_variable :
                set_cols.append(col)

    set_rows = list(set(set_rows))
    set_cols = list(set(set_cols))

    set_wb = openpyxl.Workbook()
    set_wb.active.title = 'update'
    ws = set_wb.active

    cell_color = PatternFill(start_color="ff9999", fill_type="solid")

    row_cnt = 0
    for row in wb_rows :
        col_cnt = 0
        if row in set_rows : 
            row_cnt += 1
            for col in wb_cols :
                if not col in set_cols :
                    continue
                col_cnt +=1 
                set_cell = ws.cell(row=row_cnt, column=col_cnt)
                chk_cell = st.cell(row=row, column=col)
                set_cell.value = chk_cell.value
                if not chk_cell.fill.start_color.index == '00000000' :
                    set_cell.fill = cell_color
    
    save_file_name = f'find_cells_v{file_verison}.xlsx'
    folder_name = f'{year}{month}{day}'

    if mkdir :
        chk_mkdir(folder_name)

    while True :
        if mkdir :
            save_file_name = f'find_cells_v{file_verison}.xlsx'
            path_join = os.path.join(folder_name, save_file_name)
            if not os.path.exists(path_join) :
                break
            else :
                file_verison += 1
        else :
            save_file_name = f'find_cells_{folder_name}_v{file_verison}.xlsx'
            if not os.path.exists(save_file_name) :
                break
            else :
                file_verison += 1

    save_path = path_join if mkdir else save_file_name
    set_wb.save(save_path)

    os.startfile(save_path)

    print(f'✅ Finding highlight cells is done')
    print(f' 📗 File name is \'{save_file_name}\'')


def get_file_names(
    dir='',  
    excel_only=True,
    recent=False) :

    if not type(dir) == str :
        print('❌ The dir arg is a string only')
        return
    
    if not type(excel_only) == bool :
        print('❌ The excel_only arg is a boolean only')
        return
    
    if not type(recent) == bool :
        print('❌ The recent arg is a boolean only')
        return
    
    file_list = []
    if dir == '' or dir == None :
        file_list = os.listdir()
    else :
        try :
            file_list = os.listdir(dir)
        except :
            print('❌ Path Error')

    if excel_only :
        file_list = [f for f in file_list if 'xlsx' in f or 'xls' in f]

    if not file_list :
        print('❗ This folder is empty')
        return 

    if recent :
        file_recent_chk = []
        for f_name in file_list :
            set_path = os.path.join(dir, f_name) if not dir == '' or dir == None else f_name
            written_time = os.path.getatime(set_path)
            file_recent_chk.append((os.path.join(dir, f_name), written_time))

        sorted_file_list = sorted(file_recent_chk, key=lambda x : x[1], reverse=True)
        return sorted_file_list[0][0]
    else :
        return file_list



@dataclass
class SetModify:
    pid: str
    file_name: str = None
    keyid: str = 'record'
    key: str = api_key
    server: str = api_server
    init_list: list = None
    modify_list: list = None
    delete_list: list = None
    requal_list: list = None

    def __post_init__(self) :
        if self.file_name != None :
            self.df = pd.read_excel(self.file_name, dtype=str)
        else :
            self.df = None
            print('❗ The dataframe is not setup')

        self.project_path = f'surveys/selfserve/548/{self.pid}'
        
        # api login
        api.login(self.key, self.server)

        # get variables
        datamap = api.get(f'{self.project_path}/datamap', format='json')
        self.variables = [(v['label'], v['qlabel'], v['vgroup']) for v in datamap['variables']]

        print('✅ It\'s ready to setup')
        # variables name check
        self.chk_variables = [(label, qlabel) for label, qlabel, vgroup in self.variables if (qlabel != None) and (label != qlabel) and (not qlabel in label)]
        if self.chk_variables :
            print('❗ You need to check the name of the variable')
            for label, qlabel in self.chk_variables :
                print(f'  🔹 {label} ▶ qlabel = {qlabel}' )
        
    def setup(self, 
            check_index=0,
            modi=['수정'],
            delete=['삭제'],
            respstatus='RespStatus',
            respstatus_value=99) :

        if not list(self.df.index) :
            print('❌ self.df is not defined')
            return

        # modifiy data
        self.df.iloc[:, check_index] = self.df.iloc[:, check_index].str.replace(' ', '')
        if not self.df.columns[check_index] in [label for label, qlabel, vgroup in self.variables] :
            self.modify_df = self.df[self.df.iloc[:, check_index].str.contains('|'.join(modi))].copy()
            self.modify_df.drop(self.modify_df.columns[check_index], axis=1, inplace=True)
        else :
            self.modify_df = self.df.copy()

        # delete data
        self.df.iloc[:, check_index] = self.df.iloc[:, check_index].str.replace(' ', '')
        if not self.df.columns[check_index] in [label for label, qlabel, vgroup in self.variables] :
            self.delete_df = self.df[self.df.iloc[:, check_index].str.contains('|'.join(delete))].copy()
            self.delete_df.drop(self.delete_df.columns[check_index], axis=1, inplace=True)
        else :
            self.delete_df = pd.DataFrame()
        
        print('✅ The modify/delete data found (modfiy_df/delte_df)')
        
        # modify
        if not self.modify_df.empty :
            modi_dict = self.modify_df.copy().to_dict('index')
            for idx, md in modi_dict.items() :
                for key, value in md.items() :
                    if pd.isna(value) :
                        md[key] = None
                modi_dict[idx] = md
                
            self.modify_list = list(modi_dict.values())
        
        # delete
        if not self.delete_df.empty :
            delete_dict = self.delete_df.to_dict('index')
            self.delete_list = list(delete_dict.values())
            self.delete_list = [{self.keyid:i[self.keyid], respstatus:respstatus_value} for i in self.delete_list]
        
        if self.modify_df.empty and self.delete_df.empty :
            modi_dict = self.df.to_dict('index')
            for idx, md in modi_dict.items() :
                for key, value in md.items() :
                    if pd.isna(value) :
                        md[key] = None
                modi_dict[idx] = md
            
            self.modify_list = list(modi_dict.values())
            print('  ❗ Only modify')
        
        print('✅ Setup complete')

    def set_delete(self, 
                id_list=[],
                respstatus='RespStatus',
                respstatus_value=99) :
        if not id_list :
            print('❌ The id_list is empty')
            print(f'✅ delete_list initialized.')
            self.delete_list = []
            return
        
        id_list = list(set(id_list))       
        self.delete_list = [{self.keyid:i, respstatus:respstatus_value} for i in id_list]

        print(f'✅ Complete making delete_list : total {len(self.delete_list)}\'s')

    def set_requal(self,
                id_list=[],
                respstatus='RespStatus',
                respstatus_value=1):
        if not id_list :
            print('❌ The id_list is empty')
            print(f'✅ requal_list initialized.')
            self.requal_list = []
            return

        # get markers data
        self.markers_data = api.get(f'{self.project_path}/data', format='json', fields=f'{self.keyid},markers')
        id_list = [str(i) for i in id_list]
        self.requal_list = [data for data in self.markers_data if str(data[self.keyid]) in id_list]
        if not self.requal_list :
            print('❗ There is no such ID')
            return

        bad_marker = 'bad:'
        for data in self.requal_list :
            markers = data['markers']
            markers = markers.split(',')
            if not markers :
                continue
            
            markers = [marker.replace(bad_marker, '') if bad_marker in marker else marker for marker in markers]
            markers = ','.join(markers)
            
            data['markers'] = markers
            data[respstatus] = respstatus_value
            
        print(f'✅ Complete making requal_list : total {len(self.requal_list)}\'s')

    
    def send(self, test=True, delete_mode='disqualify', delete_marker='delete_sample', delete_date=True, backup=True) :
        # Data backup
        backup_path = f'{self.project_path}/data'
        edit_api = f'{self.project_path}/data/edit'

        now = datetime.now()
        now_year = now.year
        now_month = '{0:02}'.format(now.month)
        now_day = '{0:02}'.format(now.day)
        set_date = f'{now_year}{now_month}{now_day}'
        print(f'📣 test mode is {test}')
        print('')
        if backup and not test :
            pd.io.formats.excel.ExcelFormatter.header_style = None

            print('📥 Data BackUp ... ⌛')
            data_path = os.path.join(os.getcwd(), 'BackUp')
            chk_mkdir(data_path)
            try :
                csv_data = api.get(backup_path, format='csv', cond='everything')

                binary_csv_name = f'{self.pid}_binary.csv'
                create_binary_file(data_path, binary_csv_name, csv_data)

                backup_version = 1

                while True :
                        backup_file_name_chk = f'{set_date}_v{backup_version}.xlsx'
                        path_join = os.path.join(data_path, backup_file_name_chk)
                        if not os.path.exists(path_join) :
                            break
                        else :
                            backup_version += 1

                create_ascii_file(data_path, binary_csv_name, f'{set_date}_v{backup_version}.csv')

                print(' 🔔 Data BackUp is done')
            except :
                print(' ❌ [ERROR] : Get Data API is Error')
            print('')

        if not delete_mode in ['disqualify', 'delete'] :
            print('❌ The delete_mode is only [\'disqualify\', \'delete\']')
            return

        # Modify
        try :
            if self.modify_list :
                print('🟦 Modify data')
                result = api.put(edit_api, key=self.keyid, data=self.modify_list, test=test)

                stats = result['stats']
                bad = stats['bad']
                
                if bad :
                    print(' ⛔ This found a bad samples in file. Please check modifed data.')
                    for num, label, value, err in bad :
                        print(f'  ❌ {err}')
                        print(f'      {label} = {value}')
                        print('')
                else :                
                    for key, value in stats.items() :
                        if key in ['unchanged', 'fieldsUpdated', 'created'] :
                            print(f' 🔹 {key} : {value}')
                    print('-'*8)
            else :
                print('❗ The modfiy_list is empty.')
        except :
            print('❌ Decipher API modify error')
        
        print('')

        # Delete
        try :
            if self.delete_list :
                print(f'🟦 Delete data (mode = {delete_mode})')
                if delete_date :
                    delete_marker = f'{delete_marker}_{set_date}'
                
                dt_result = api.delete(edit_api, key=self.keyid, data=self.delete_list, test=test, mode=delete_mode, disqualify=delete_marker)

                stats = dt_result['stats']
                bad = stats['bad']
                if bad :
                    print(' ⛔ This found a bad samples in file. Please check delete data.')
                    for num, label, value, err in bad :
                        print(f'  ❌ {err}')
                        print(f'      {label} = {value}')
                        print('')
                else :
                    if delete_mode == 'disqualify' :
                        print(f' 📣 marker = {delete_marker}')
                    for key, value in stats.items() :
                        if key in ['unchanged', 'fieldsUpdated', 'deleted', 'disqualified'] :
                            print(f' 🔹 {key} : {value}')
                    print('-'*8)
            else :
                print('❗ The delete_list is empty.')
        except :
            print('❌ Decipher API delete error')

        print('')

        # Requalify
        try :
            if self.requal_list :
                print('🟦 Requalify data')
                result = api.put(edit_api, key=self.keyid, data=self.requal_list, test=test)

                stats = result['stats']
                bad = stats['bad']
                
                if bad :
                    print(' ⛔ This found a bad samples in file. Please check requalify data.')
                    for num, label, value, err in bad :
                        print(f'  ❌ {err}')
                        print(f'      {label} = {value}')
                        print('')
                else :                
                    for key, value in stats.items() :
                        if key in ['unchanged', 'fieldsUpdated', 'created'] :
                            print(f' 🔹 {key} : {value}')
                    print('-'*8)
            else :
                print('❗ The requal_list is empty.')
        except :
            print('❌ Decipher API Requalify error')
        
        print('')

        if test :
            print('❗ It\'s a test mode. If you want to update at the Decipher, enter the \'False\' at the test argument ( .send(test=False) )')
        else :
            print('✅ Data update complete')        


