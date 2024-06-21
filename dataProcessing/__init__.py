import pandas as pd
from IPython.display import display, HTML
from typing import Union, List, Tuple, Dict, Optional, Literal, Callable, Any, NoReturn
import numpy as np
from dataclasses import dataclass, field
import contextlib
import os
import openpyxl
import re
import nbformat as nbf
from collections import OrderedDict
import json
from decipher.beacon import api
import time
from ..key import api_key, api_server
from decipherAutomatic.getFiles import *
from decipherAutomatic.utils import *
from pandas.io.formats import excel
import zipfile

def check_print(variables: Union[List[str], Tuple[str, ...], str], 
                error_type: Literal['SA', 'MA', 'LOGIC', 'DUP'], 
                df: pd.DataFrame, 
                warnings: Optional[List[str]] = None,
                alt: Optional[str] = None) -> str:
    qid = None
    
    if isinstance(variables, str): 
        qid = variables

    if isinstance(variables, list) or isinstance(variables, tuple) :
        list_vars = list(variables)
        if len(list_vars) == 1 :
            qid = list_vars[0]
        if len(list_vars) >= 2 :
            qid = f'{list_vars[0]} - {list_vars[-1]}'
    
    error_type_msg = {
        'SA': 'SA Variable Check',
        'MA': 'MA Variable Check',
        'LOGIC': 'If Base cond is True, then Answer cond is also True',
        'DUP': 'Answer Duplicate Check',
        'MASA': 'Multi Variable Base Single Variable Check',
        'MAMA': 'Multi Variable Base Multi Variable Check',
        'MARK': 'Multi Variable Base Rank Variable Check',
        'RATERANK' : 'Check if Rank is answered in order of Rate question score'
    }


    print_str = ''
    print_str += f"""<div class="datcheck-title">📢 <span class="title-type">{error_type}</span> <span class="title-msg">({error_type_msg[error_type]})</span></div>""" # Error Text Title

    # Result HTML
    correct = """<div class="datacheck-head check-correct">✅ {html_title}</div>"""
    fail    = """<div class="datacheck-head check-fail">❌ {html_title} : Error {err_cnt}'s</div>"""
    check   = """<div class="datacheck-check">📌 <span class="print-comment">{check_title}</span></div>"""
    warning  = """<div class="datacheck-warning check-warn">⚠️ {warn_title}</div>"""

    if warnings is not None :
        for warn in warnings :
            print_str += warning.format(warn_title=warn)

    # Base Check
    err_cols = df.columns
    
    ms_err = 'BASE'
    if ms_err in err_cols :
        err_cnt = len(df[df[ms_err]==1])
        print_str += f"""👨‍👩‍👧‍👦 <span class="print-comment">Check Sample : <span class="check-bold">{len(df)}'s</span></span>"""
        html_title = f"""Answer Base Check"""
        if err_cnt == 0 :
            print_str += correct.format(html_title=html_title)
        else :
            print_str += fail.format(html_title=html_title, err_cnt=err_cnt)

    # Cases responded to other than base 
    add_err = 'ADD'
    if add_err in err_cols :
        err_cnt = len(df[df[add_err]==1])
        html_title = "Other than Base Check"
        if err_cnt >= 1:
            print_str += fail.format(html_title=html_title, err_cnt=err_cnt)

    # Answer Able Value Check  / SA 
    only_err = 'ONLY'
    if only_err in err_cols :
        err_cnt = len(df[df[only_err]==1])
        html_title = "Answer Able Value Check"
        if err_cnt == 0 :
            print_str += correct.format(html_title=html_title)
        else :
            print_str += fail.format(html_title=html_title, err_cnt=err_cnt)
        
        err_answer = list(df[df[only_err]==1][variables].values)
        err_answer = ['NA' if pd.isna(x) else x for x in err_answer]
        err_answer = list(set(err_answer))
        if err_answer :
            print_str += f"""<div class="print-padding-left">🗒️ <span class="print-comment">Invalid response</span> : {list(err_answer)}</div>"""

    # Disable Value Check  / SA 
    disable_err = 'DISABLE'
    if disable_err in err_cols :
        err_cnt = len(df[df[disable_err]==1])
        html_title = "Answer Disable Value Check"
        if err_cnt == 0 :
            print_str += correct.format(html_title=html_title)
        else :
            print_str += fail.format(html_title=html_title, err_cnt=err_cnt)
        
        err_answer = list(set(list(df[df[disable_err]==1][variables].values)))
        if err_answer :
            print_str += f"""<div class="print-padding-left">🗒️ <span class="print-comment">Invalid response</span> : {list(err_answer)}</div>"""

    # MA Variable Answer Count Check
    if (error_type in ['MA']) :
        for lg in ['ATLEAST', 'ATMOST', 'EXACTLY'] :
            if not lg in err_cols :
                continue
            err_cnt = len(df[df[lg]==1])
            html_title = f"{lg} Check"
            if err_cnt == 0 :
                print_str += correct.format(html_title=html_title)
            else :
                print_str += fail.format(html_title=html_title, err_cnt=err_cnt)

    # Answer Description Print
    desc_table = None
    if (error_type in ['SA']) and (pd.api.types.is_numeric_dtype(df[qid])) :
        desc = df[qid].describe().round(1)
        desc_table = """
    <div class="datacheck-desc">📋 {qid} Describe</div>
    <table class="print-padding-left"">
        <tr><td><b>Count</b></td><td>{cnt}</td></tr>
        <tr><td><b>Mean</b></td><td>{mean}</td></tr>
        <tr><td><b>Min</b></td><td>{minv}</td></tr>
        <tr><td><b>Max</b></td><td>{maxv}</td></tr>
    </table>""".format(qid=qid, cnt=desc.loc['count'], mean=desc.loc['mean'], minv=desc.loc['min'], maxv=desc.loc['max'])

    if (error_type in ['MA']) :
        desc = df['ANS_CNT'].describe().round(1)
        desc_table = """
    <div class="datacheck-desc">📋 {qid} Answer Count Describe</div>
    <table class="print-padding-left">
        <tr><td><b>Mean</b></td><td>{mean}</td></tr>
        <tr><td><b>Min</b></td><td>{minv}</td></tr>
        <tr><td><b>Max</b></td><td>{maxv}</td></tr>
    </table>""".format(qid=qid, mean=desc.loc['mean'], minv=desc.loc['min'], maxv=desc.loc['max'])

    if desc_table is not None :
        print_str += desc_table


    # Logic Check
    if (error_type in ['LOGIC', 'MASA', 'MAMA', 'MARK', 'RATERANK']) :
        err_cnt = len(df[df['LOGIC']==1])
        base_cond = 'BASE_COND'
        if base_cond in list(df.columns) :
            base_cnt = len(df[df[base_cond]==1])
            print_str += check.format(check_title=f"Base Cond Answer Count : <b>{base_cnt}'s</b>")
        if err_cnt == 0 :
            print_str += correct.format(html_title=f"Logic Correct")
        else :
            print_str += fail.format(html_title="Logic has Error", err_cnt=err_cnt)

    # Duplicate Check
    if (error_type in ['DUP']) :
        err_cnt = len(df[df['DUP']==1])
        if err_cnt == 0 :
            print_str += correct.format(html_title="No Duplicate")
        else :
            dup_rows = df[df['DUP'] == 1]
            summary = []
            
            for index, row in dup_rows.iterrows():
                row_values = row[variables]
                duplicates = row_values[row_values.duplicated(keep=False)]
                summary.extend(duplicates.unique().tolist())
            
            summary = list(set(summary))
            print_str += fail.format(html_title=f"Duplicate Answer", err_cnt=err_cnt)
            print_str += f"""<div class="print-padding-left">🗒️ <span class="print-comment">Invalid response</span> : {summary}</div>"""


    print_type = "alt-main"
    if "check-fail" in print_str :
        print_type = "alt-fail"

    final_print = f"""
    <div class="datacheck-print {print_type}">
        <div class="datacheck-alt">{alt if alt is not None else qid}</div>
        <div class="datacheck-result">
            {print_str}
        </div>
    </div>
    """

    return final_print

def get_key_id(base: List[str]) -> Union[None, str]:
    """`base`가 `qid`를 포함하는지 확인합니다."""
    qid = base[0]
    qids = list(qid)
    qids.reverse()
    for q in qids:
        if not q.isdigit():
            break
        else:
            qid = qid[:-1]

    for ma in base:
        if qid not in ma:
            print_text = """<div class="check-bold check-fail">❌ [ERROR] Please check multi question variable names</div>"""
            print_text += f"""<div class="print-paiddng-left">Base MA variable key name : <span class="check-correct">{qid}</span></div>"""
            display(HTML(print_text))
            qid = None
            return qid
        
    return qid

def css_apply() -> None :
    module_path = os.path.dirname(__file__)
    css_file_path = os.path.join(module_path, 'styles.css')  # Assuming 'styles.css' is the CSS file in the module

    try:
        with open(css_file_path, 'r') as file:
            css_content = file.read()
        css = f"""
        <style>
        {css_content}
        </style>
        <div class="check-correct check-bold">❇️ DataCheck CSS Set UP</div>
        """
        display(HTML(css))
    except Exception as e:
        print(f"Failed to load CSS file: {e}")

def lambda_ma_to_list(row, qids) :
    qid_key = get_key_id(qids)
    return [int(x.replace(qid_key, '')) for x in qids if not (pd.isna(row[x]) or row[x] == 0)]

@dataclass
class PrintDataFrame:
    show_cols: List[str]
    df: pd.DataFrame

    def __call__(self, extra: Optional[List[str]] = None):
        if extra is None:
            return self.df[self.show_cols]
        return self.df[self.show_cols + extra] if extra else self.df

@dataclass
class ErrorDataFrame:
    chk_id: str
    qid_type: str
    show_cols: List[str]
    df: pd.DataFrame
    err_list: List[str]
    warnings: List[str] = field(default_factory=list)
    alt: Optional[str] = None

    def __post_init__(self):
        self.show_col_with_err = self.err_list + self.show_cols
        self.err_base = [x for x in self.err_list if x not in ['BASE_COND', 'ANSWER_COND']]
        err_df = self.df[(self.df[self.err_base]==1).any(axis=1)]
        self.err = PrintDataFrame(self.show_col_with_err, err_df)
        self.full = PrintDataFrame(self.show_col_with_err, self.df)
        self.chk_msg = check_print(self.chk_id, self.qid_type, self.full(), self.warnings, self.alt)
        self.extra_cols = []

    def __getitem__(self, key):
        extra_cols = [key] if isinstance(key, str) else key
        self.extra_cols = extra_cols
        return self.err()[self.base + extra_cols]

    def __repr__(self):
        return ''

class DataCheck(pd.DataFrame):
    _metadata = ['_keyid', '_spssmeta']

    def __init__(self, *args, **kwargs):
        self._keyid = kwargs.pop('keyid', None)
        self._spssmeta = kwargs.pop('spssmeta', None)
        
        super().__init__(*args, **kwargs)
        if self._keyid is not None:
            self[self._keyid] = self[self._keyid].astype(int)
            self.set_index(self._keyid, inplace=True)

        pd.set_option('display.max_columns', None)
        pd.set_option('display.max_rows', None)

        self._count_fnc: Callable[[pd.Series], int] = lambda x: x.count() - (x==0).sum()
        self.attrs['display_msg'] = 'all'
        self.attrs['default_filter'] = pd.Series([True] * len(self), index=self.index)
        self.attrs['result_html'] = []
        self.attrs['meta'] = self._spssmeta

    @property
    def _constructor(self) -> Callable[..., 'DataCheck']:
        return DataCheck

    def any(self, *args, **kwargs) -> pd.Series:
        if 'axis' not in kwargs:
            kwargs['axis'] = 1
        return super().any(*args, **kwargs)

    def all(self, *args, **kwargs) -> pd.Series:
        if 'axis' not in kwargs:
            kwargs['axis'] = 1
        return super().all(*args, **kwargs)

    @property
    def keyid(self) -> Optional[str]:
        return self._keyid

    @keyid.setter
    def keyid(self, value: Optional[str]) -> None:
        self._keyid = value

    # @property
    # def meta(self) -> Any :
    #     return self.attrs['meta']

    # @meta.setter
    # def meta(self, meta_data: Optional[Any]) -> None :
    #     self.attrs['meta'] = meta_data

    @property
    def display_msg(self) -> Optional[Literal['all', 'error', None]]:
        return self.attrs['display_msg']

    @display_msg.setter
    def display_msg(self, option: Optional[Literal['all', 'error', None]]) -> None:
        if not option in ['all', 'error', None] :
            display(HTML(f"""<div class="check-bold check-fail">❌ The argument option can only be a 'all', 'error', None</div>"""))
            return
        self.attrs['display_msg'] = option

    @property
    def default_filter(self) -> pd.Series :
        return self.attrs['default_filter']

    @property
    def count_fnc(self) -> Callable[[pd.Series], int]:
        return self._count_fnc

    @count_fnc.setter
    def count_fnc(self, fnc: Callable[[pd.Series], int]) -> None:
        if not callable(fnc):
            raise ValueError("The value must be a callable.")
        self._count_fnc = fnc

    @staticmethod
    def result_alt(qid: Union[str, List], alt: Optional[str]=None) -> str :
        alt_qid = qid
        if isinstance(qid, list) :
            alt_qid = f'{qid[0]}-{qid[-1]}'
        result_alt = alt_qid if alt is None else f'{alt_qid}: {alt}'
        return result_alt

    def result_html_update(self, **kwargs) :
        result_html = self.attrs['result_html'].copy()            
        key = 'alt'
        updated = False
        if key in kwargs :
            chk_alt = {idx: result[key].strip().replace(' ', '') for idx, result in enumerate(result_html) if key in result and isinstance(result[key], str)}
            curr = kwargs[key].strip().replace(' ', '')
            for idx, value in chk_alt.items() :
                if curr == value :
                    result_html[idx] = kwargs
                    updated = True
        if not updated :
            result_html.append(kwargs)
        self.attrs['result_html'] = result_html

    def comp(self) :
        """
        전체 데이터 기준 `Series`를 Return
        """
        return pd.Series([True] * len(self), index=self.index)

    def set_filter(self, filter_cond: Optional[pd.Series] = None) -> None:
        """
        데이터 체크 기본 필터를 변경
        `filter_cond` (pd.Series or None) : `None`이면 전체 샘플 기준으로 변경
        """
        if filter_cond is None :
            self.attrs['default_filter'] = self.comp()
            display(HTML(f"""🛠️ <span class="check-bold">Data Filter <span class="check-warn">Reset</span> : {len(self)}'s</span>"""))
        else :
            self.attrs['default_filter'] = filter_cond
            filt_data = self[filter_cond]
            display(HTML(f"""🛠️ <span class="check-bold">Data Filter <span class="check-warn">Setting</span> : {len(filt_data)}'s</span>"""))

    def col_name_check(self, *variables: str) -> bool:
        """`qid`에 지정된 열이 데이터프레임에 있는지 확인"""
        chk_qid = [qid for qid in variables if not qid in list(self.columns)]
        if chk_qid :
            display(HTML(f"""<div class="check-bold check-fail">❌ The variable {chk_qid} is not in the data frame</div>"""))
            return False
        return True

    @contextlib.contextmanager
    def preserve_display_msg(self):
        original_display_msg = self.attrs['display_msg']
        try:
            yield
        finally:
            self.attrs['display_msg'] = original_display_msg

    def display_descriptin(self, alt: str, title: str, desc: pd.DataFrame) -> None :
        desc_table = """
<table>
    <tr><td><b>Mean</b></td><td>{mean}</td></tr>
    <tr><td><b>Min</b></td><td>{minv}</td></tr>
    <tr><td><b>Max</b></td><td>{maxv}</td></tr>
</table>""".format(mean=desc.loc['mean'], minv=desc.loc['min'], maxv=desc.loc['max'])
        
        print_result = f"""
<div class="datacheck-apply alt-sub">
    <div class="datacheck-alt">{alt}</div>
    <div class="apply-title">📊 {title}</div>
    <div class="print-padding-left">{desc_table}</div>
</div>
"""
        if self.attrs['display_msg'] ==  'all' :
            display(HTML(print_result))

    def count_col(self, cnt_col_name: str, cols: Union[List[str], Tuple[str], str], value: Optional[Union[int, List[int]]] = None) -> None:
        if not self.col_name_check(*cols) : return
        
        cnt_col = []
        alt = ""
        if isinstance(cols, tuple) or isinstance(cols, list):
            cnt_col = self.ma_return(cols)
            alt = f"{cnt_col[0]}-{cnt_col[-1]} Answer Count"
        if isinstance(cols, str) :
            cnt_col = [cols]
            alt = f"{cols} Answer Count"
        
        if value is None:
            new_col = self[cnt_col].apply(self._count_fnc, axis=1).rename(cnt_col_name)
        elif isinstance(value, int):
            new_col = self[cnt_col].apply(lambda row: row.isin([value]).sum(), axis=1).rename(cnt_col_name)
        elif isinstance(value, list):
            new_col = self[cnt_col].apply(lambda row: row.isin(value).sum(), axis=1).rename(cnt_col_name)
        
        with self.preserve_display_msg():
            result = self.assign(**{cnt_col_name: new_col})
            self.__dict__.update(result.__dict__)
        
        
        show_title = ""
        if value is None:
            show_title += f"""<span class="check-bold check-warn">{cnt_col_name}</span> : Answer count"""
        else:
            show_title += f"""<span class="check-bold check-warn">{cnt_col_name}</span> : Value count ({value})"""

        desc = self[cnt_col_name].describe().round(1)
        self.display_descriptin(alt, show_title, desc)
        

    def sum_col(self, sum_col_name: str, cols: Union[List[str], Tuple[str], str]) -> None:
        if not self.col_name_check(*cols):
            return

        sum_col = []
        alt =""
        if isinstance(cols, (tuple, list)):
            sum_col = self.ma_return(cols)
            alt = f"{sum_col[0]}-{sum_col[-1]} : Sum"

        elif isinstance(cols, str):
            sum_col = [cols]
            alt = f"{cols} : Sum"

        new_col = self[sum_col].sum(axis=1).rename(sum_col_name)

        with self.preserve_display_msg():
            result = self.assign(**{sum_col_name: new_col})
            self.__dict__.update(result.__dict__)

        show_title = f"""<div>📊 <span class="check-bold check-warn">{sum_col_name}</span> : Sum of values</div>"""

        desc = self[sum_col_name].describe().round(1)
        self.display_descriptin(alt, show_title, desc)

    def _update_self(self, new_data):
        # self의 내부 데이터를 new_data로 업데이트
        self.__dict__.update(new_data.__dict__)
     

    def ma_to_list(self, list_col_name: str, cols: Union[List[str], Tuple[str], str]) -> None:
        if not self.col_name_check(*cols):
            return

        ma_col = []
        if isinstance(cols, (tuple, list)):
            ma_col = self.ma_return(cols)
        elif isinstance(cols, str):
            ma_col = [cols]

        new_col = self[ma_col].apply(lambda_ma_to_list, axis=1, qids=ma_col).rename(list_col_name)
        result = self.assign(**{list_col_name: new_col})
        self.__dict__.update(result.__dict__)

        print_result = f"""<div>📊 <span class="check-bold check-warn">{list_col_name}</span> : MA Variable to List</div>"""

        if self.attrs['display_msg'] ==  'all' :
            display(HTML(print_result))

    def ma_check(self, 
                ma: Union[List[str], Tuple[str]],
                len_chk: bool = True) -> bool:
        """`ma`가 리스트나 튜플인지, 그리고 다른 조건들을 만족하는지 확인"""
        fail = """<div class="check-bold check-fail">❌ [ERROR] {warn_text}</div>"""
        example_text = """<div class="example-text">Example) {ex_text}</div>"""
        print_str = ""
        if not ma:
            print_str += fail.format(warn_text="Please check variable names")
            print_str += example_text.format(ex_text="['Q1r1', 'Q1r2', 'Q1r3'] / ('Q1r1', 'Q1r3')")
            display(HTML(print_str))
            return True

        if not isinstance(ma, (list, tuple)):
            print_str += fail.format(warn_text="Type of variable must be list or tuple")
            display(HTML(print_str))
            return True

        if len_chk and len(ma) < 2:
            print_str += fail.format(warn_text="Variable must be 2 length or more")
            display(HTML(print_str))
            return True

        if isinstance(ma, tuple) and len(ma) != 2:
            print_str += fail.format(warn_text="The variable must include 2 arguments")
            display(HTML(print_str))
            return True

        if isinstance(ma, tuple):
            cols = list(self.columns)
            first_index = cols.index(ma[0])
            last_index = cols.index(ma[1])
            if first_index > last_index:
                print_str += fail.format(warn_text=f"Please check the column index / current index ({first_index}-{last_index})")
                display(HTML(print_str))
                return True
        return False

    def ma_return(self,
                  ma: Union[List[str], Tuple[str]]) -> List[str]:
        """`ma`에 지정된 열을 반환"""
        if isinstance(ma, tuple):
            cols = list(self.columns)
            first_index = cols.index(ma[0])
            last_index = cols.index(ma[1]) + 1
            return cols[first_index:last_index]
        elif isinstance(ma, list):
            return ma

    def show_message(self, 
                     export_df: ErrorDataFrame) -> None :
        if self.attrs['display_msg'] ==  'all' :
            display(HTML(export_df.chk_msg))    
        elif self.attrs['display_msg'] ==  'error' :
            if len(export_df.err()) > 1 :
                display(HTML(export_df.chk_msg))
        elif self._display_msg is None :
            return


    def safreq(self, 
           qid: Optional[str] = None, 
           cond: Optional[pd.Series] = None, 
           only: Optional[Union[range, List[Union[int, float, str]], int, float, str]] = None,
           disabled: Optional[Union[range, List[Union[int, float, str]], int, float, str]] = None,
           alt: Optional[str]=None) -> 'ErrorDataFrame':
        """
        단수 응답(단일 변수) 데이터 체크 메서드
        """
        
        chk_df = self[self.attrs['default_filter']].copy()
        if not self.col_name_check(qid) : return

        show_cols = [qid]
        
        err_list = []

        # Answer Base Check
        warnings = []

        ms_err = 'BASE'
        filt = (chk_df[qid].isna())  # Default
        if cond is not None:
            filt = (filt) & (cond)
            if len(chk_df[cond.reindex(chk_df.index, fill_value=False)]) == 0:
                warnings.append("No response to this condition")

        chk_df.loc[filt, ms_err] = 1

        err_list.append(ms_err)

        # Cases responded to other than base
        if cond is not None :
            ans_err = 'ADD'
            chk_df.loc[(~chk_df[qid].isna()) & (~cond), ans_err] = 1
            err_list.append(ans_err)

        # ONLY ANSWER CHECK
        if only is not None:
            if isinstance(only, range):
                only = list(only) + [only[-1] + 1]
            elif isinstance(only, (int, float, str)):
                only = [only]

            only_cond = (~chk_df[qid].isin(only))
            if cond is not None:
                only_cond = (only_cond) & (cond)
            
            only_err = 'ONLY'
            chk_df.loc[only_cond, only_err] = 1
            err_list.append(only_err)
        
        # DONT ANSWER CHECK
        if disabled is not None:
            if isinstance(disabled, range):
                disabled = list(disabled) + [disabled[-1] + 1]
            elif isinstance(disabled, (int, float, str)):
                disabled = [disabled]
            
            disabled_cond = (chk_df[qid].isin(disabled))
            if cond is not None:
                disabled_cond = (disabled_cond) & (cond)
            
            disable_err = 'DISABLE'
            chk_df.loc[disabled_cond, disable_err] = 1
            err_list.append(disable_err)
        
        chk_df = chk_df if cond is None else chk_df[cond.reindex(chk_df.index, fill_value=False)]
        
        edf = ErrorDataFrame(qid, 'SA', show_cols, chk_df, err_list, warnings, alt)
        self.show_message(edf)
        self.result_html_update(alt=self.result_alt(qid, alt), result_html=edf.chk_msg, dataframe=edf.err()[show_cols+edf.extra_cols].to_json())
        return edf

    def mafreq(self, 
            qid: Union[List[str], Tuple[str, ...]], 
            cond: Optional[pd.Series] = None, 
            atleast: Optional[int] = None, 
            atmost: Optional[int] = None, 
            exactly: Optional[int] = None,
            alt: Optional[str]=None) -> 'ErrorDataFrame':
        """
        복수 응답(다중 변수) 데이터 체크 메서드
        """
        if (self.ma_check(qid)) :
            return
        
        chk_df = self[self.attrs['default_filter']].copy()
        show_cols = self.ma_return(qid)
        if not self.col_name_check(*show_cols) : return

        cnt = 'ANS_CNT'
        chk_df[cnt] = chk_df[show_cols].apply(lambda x: x.count() - (x==0).sum(), axis=1)

        err_list = []

        # Answer Base Check
        warnings = []

        ms_err = 'BASE'
        filt = (chk_df[cnt]==0)  # Default
        if cond is not None:
            filt = (filt) & (cond)
            if len(chk_df[cond.reindex(chk_df.index, fill_value=False)]) == 0:
                warnings.append("No response to this condition")

        chk_df.loc[filt, ms_err] = 1

        err_list.append(ms_err)

        # Cases responded to other than base
        if cond is not None :
            ans_err = 'ADD'
            chk_df.loc[(chk_df[cnt]>=1) & (~cond), ans_err] = 1
            err_list.append(ans_err)

        # Generalized Answer Check Function
        def check_answer(condition, operator, err_label):
            if condition is not None:
                if operator == '==':
                    cond_err = (chk_df[cnt] != condition)
                elif operator == '<':
                    cond_err = (chk_df[cnt] < condition)
                elif operator == '>':
                    cond_err = (chk_df[cnt] > condition)
                if cond is not None:
                    cond_err = (cond_err) & (cond)
                chk_df.loc[cond_err, err_label] = 1
                err_list.append(err_label)

        # AT LEAST, AT MOST, EXACTLY Answer Checks
        check_answer(atleast, '<', 'ATLEAST')
        check_answer(atmost, '>', 'ATMOST')
        check_answer(exactly, '==', 'EXACTLY')

        show_cols = [cnt] + show_cols
        
        chk_df = chk_df if cond is None else chk_df[cond.reindex(chk_df.index, fill_value=False)]

        edf = ErrorDataFrame(qid, 'MA', show_cols, chk_df, err_list, warnings, alt)
        self.show_message(edf)
        self.result_html_update(alt=self.result_alt(qid, alt), result_html=edf.chk_msg, dataframe=edf.err()[show_cols+edf.extra_cols].to_json())
        return edf


    def logchk(self, 
               base_cond: Optional[pd.Series] = None, 
               answer_cond: pd.Series = None,
               alt: Optional[str]=None) -> 'ErrorDataFrame':
        """
        특정 로직에 대한 응답 체크
        (`base_cond`가 `True`일 때, `answer_cond`도 `True`)

        `base_cond` (pd.Series): 베이스 조건.
        `answer_cond` (pd.Series): 베이스 조건이 True일 때 응답 조건.
        """
        chk_df = self[self.attrs['default_filter']].copy()
        if answer_cond is None :
            display(HTML("""<div class="check-bold check-fail">❌ [ERROR]  answer_cond cannot be None</div>"""))
            return 
        err_list = []

        # Base Condition Answer Check
        warnings = []
        if len(chk_df[base_cond]) == 0:
            warnings.append("No response to this condition")
        
        # Base Filter
        base_col = 'BASE_COND'
        answer_col = 'ANSWER_COND'
        err_list += [base_col, answer_col]
        chk_df.loc[base_cond, base_col] = 1
        chk_df.loc[answer_cond, answer_col] = 1

        # Logic Check
        lg_err = 'LOGIC'
        base = self.comp() if base_cond is None else base_cond
        chk_df.loc[(base) & (~answer_cond), lg_err] = 1
        err_list.append(lg_err)


        chk_df = chk_df[base.reindex(chk_df.index, fill_value=False)]
        
        qid = 'LOGIC CHECK'
        edf = ErrorDataFrame('LOGIC CHECK', 'LOGIC', [], chk_df, err_list, warnings, alt)
        self.show_message(edf)
        
        if alt is not None :
            self.result_html_update(alt=alt, result_html=edf.chk_msg, dataframe=edf.err()[edf.extra_cols].to_json())
        return edf

    def dupchk(self, 
           qid: Union[List[str], Tuple[str, ...]], 
           okUnique: Optional[Union[List[Any], range, int, str]] = None,
           alt: Optional[str]=None) -> 'ErrorDataFrame' :
        """
        중복 응답 데이터 체크 메서드 (순위 응답)        
        `qid` (Union[List[str], Tuple[str]]): 중복을 체크할 열들.
        `okUnique` (Union[List, range, int, str], optional): 무시할 특정 값(들). 기본값은 None.
        """
        if (self.ma_check(qid)) :
            return
        
        chk_df = self[self.attrs['default_filter']].copy()
        show_cols = self.ma_return(qid)
        if not self.col_name_check(*show_cols): return

        warnings = []
        err_list = []

        if okUnique is not None:
            if not isinstance(okUnique, (list, range, int, str)):
                display(HTML("""<div class="check-bold check-fail">❌ [ERROR] Type of okUnique must be list, range, int, or str</div>"""))
                return
            if isinstance(okUnique, range):
                okUnique = list(okUnique)
                okUnique.append(okUnique[-1] + 1)
            elif isinstance(okUnique, (int, str)):
                okUnique = [okUnique]
            
            warnings.append(f"""Allow Duplicates : {okUnique}""")
        else:
            okUnique = []

        dup_err = 'DUP'
        def check_duplicates(row):
            row_values = row.tolist()
            filtered_values = [value for value in row_values if value not in okUnique]
            return 1 if len(filtered_values) != len(set(filtered_values)) else None
        
        chk_df[dup_err] = chk_df[show_cols].apply(check_duplicates, axis=1)
        err_list.append(dup_err)


        edf = ErrorDataFrame(show_cols, 'DUP', show_cols, chk_df, err_list, warnings, alt)
        self.show_message(edf)
        self.result_html_update(alt=self.result_alt(qid, alt), result_html=edf.chk_msg, dataframe=edf.err()[show_cols+edf.extra_cols].to_json())
        return edf

    def masa(self, 
             ma_qid: Union[List[str], Tuple[str]], 
             sa_qid: str, 
             cond: Optional[pd.Series] = None, 
             diff_value: Optional[Union[List[Any], range, int, str]] = None,
             alt: Optional[str]=None) -> 'ErrorDataFrame' :
        """
        `복수 응답`을 베이스로 하는 `단수 응답` 로직 체크.
        `ma_qid` (Union[List[str], Tuple[str]]): 복수 응답 열 목록.
        `sa_qid` (str): 단수 응답 열.
        `cond` (pd.Series, optional): 조건을 적용할 시리즈. 기본값은 None.
        `diff_value` (Union[List, int], optional): 무시할 특정 값(들). 기본값은 None.
        """
        if (self.ma_check(ma_qid)) :
            return
        warnings = []
        err_list = []
         
        chk_df = self[self.attrs['default_filter']].copy()
        base_qid = self.ma_return(ma_qid)
        if not self.col_name_check(*base_qid): return
        if not self.col_name_check(sa_qid): return

        qid_key = get_key_id(base_qid)
        if qid_key is None: return

        ma = base_qid
        sa = sa_qid

        show_cols = [sa] + ma

        filt = ~chk_df[sa].isna()

        base_col = 'BASE_COND'
        if cond is not None :
            chk_df.loc[cond, base_col] = 1
            err_list.append(base_col)
            filt = (filt) & (cond)

        err_col = 'LOGIC'
        # MA Base SA
        if len(chk_df[filt]) == 0 :
            display(HTML("""<div class="check-bold check-warn">⚠️ No data available for verification</div>"""))
            return

        dv = []
        if diff_value is not None:
            if not isinstance(diff_value, (list, range, int, str)):
                display(HTML("""<div class="check-bold check-fail">❌ [ERROR] Type of diff_value must be list, range, int, or str</div>"""))
                return
            if isinstance(diff_value, (int, str)) :
                dv = [diff_value]
            if isinstance(diff_value, list) :
                dv = diff_value
            if isinstance(diff_value, range):
                dv = list(diff_value)
                dv.append(dv[-1] + 1)
            warnings.append(f"""Do not check the code : {dv}""")
        
        def ma_base_check(x) :
            sa_ans = x[sa]
            ma_var = f'{qid_key}{sa_ans}'
            ma_ans = x[ma_var]
            if sa_ans in dv :
                return np.nan

            return 1 if pd.isna(ma_ans) or ma_ans == 0 else np.nan

        chk_df[err_col] = chk_df[filt].apply(ma_base_check, axis=1)

        ma_ans = 'BASE_ANS'
        chk_df[ma_ans] = chk_df[filt].apply(lambda_ma_to_list, axis=1, qids=ma)

        err_list += [err_col, ma_ans]

        chk_df = chk_df if cond is None else chk_df[cond.reindex(chk_df.index, fill_value=False)]
        
        edf = ErrorDataFrame(f"""{sa}(SA) in {ma[0]}-{ma[-1]}(MA)""", 'MASA', show_cols, chk_df, err_list, warnings, alt)
        self.show_message(edf)
        self.result_html_update(alt=self.result_alt(qid, alt), result_html=edf.chk_msg, dataframe=edf.err()[show_cols+edf.extra_cols].to_json())
        return edf

    def mama(self,
             base_ma: Union[List[str], Tuple[str]], 
             chk_ma: Union[List[str], Tuple[str]], 
             cond: Optional[pd.Series] = None, 
             diff_value: Optional[Union[List[Any], range, int, str]] = None,
             alt: Optional[str]=None) -> 'ErrorDataFrame' :
        """
        `복수 응답`을 베이스로 하는 `복수 응답` 로직 체크.
        `base_ma` (Union[List[str], Tuple[str]]): 기준이 되는 복수 응답 열 목록.
        `chk_ma` (Union[List[str], Tuple[str]]): 체크할 복수 응답 열 목록.
        `cond` (pd.Series, optional): 조건을 적용할 시리즈. 기본값은 None.
        `diff_value` (Union[List, int], optional): 무시할 특정 값(들). 기본값은 None.
        """
        if (self.ma_check(base_ma)) or (self.ma_check(chk_ma)) :
            return
        warnings = []
        err_list = []
         
        chk_df = self[self.attrs['default_filter']].copy()
        base = self.ma_return(base_ma)
        chkm = self.ma_return(chk_ma)
        if not self.col_name_check(*base): return
        if not self.col_name_check(*chkm): return

        qid_key = get_key_id(base)
        ans_key = get_key_id(chkm)
        if qid_key is None: return

        zip_cols = [list(x) for x in zip(base, chkm)]
        show_cols = sum(zip_cols, [])

        chk_cnt = 'CHK_CNT'
        chk_df[chk_cnt] = chk_df[chkm].apply(lambda x: x.count() - (x==0).sum(), axis=1)
        filt = chk_df[chk_cnt]>=1

        base_col = 'BASE_COND'
        if cond is not None :
            chk_df.loc[cond, base_col] = 1
            err_list.append(base_col)
            filt = (filt) & (cond)
        
        err_col = 'LOGIC'
        # MA Base MA
        if len(chk_df[filt]) == 0 :
            display(HTML("""<div class="check-bold check-warn">⚠️ No data available for verification</div>"""))
            return

        dv = []
        diff_qids = []
        if diff_value is not None:
            if not isinstance(diff_value, (list, range, int, str)):
                display(HTML("""<div class="check-bold check-fail">❌ [ERROR] Type of diff_value must be list, range, int, or str</div>"""))
                return
            if isinstance(diff_value, (int, str)) :
                dv = [diff_value]
            if isinstance(diff_value, list) :
                dv = diff_value
            if isinstance(diff_value, range):
                dv = list(diff_value)
                dv.append(dv[-1] + 1)
            
            warnings.append(f"""Do not check the code : {dv}""")
            diff_qids = [f'{qid_key}{x}' for x in dv]

        def ma_base_check(x) :
            def flag(b, a) :
                if pd.isna(b) or b == 0 :
                    if not (pd.isna(a) or a == 0) :
                        return True
                
                return False
            return 1 if any(flag(x[base], x[ans]) for base, ans in zip_cols if not base in diff_qids) else np.nan
            
        chk_df[err_col] = chk_df[filt].apply(ma_base_check, axis=1)


        def diff_ans_update(row, cols) :
            return [int(base.replace(qid_key, '')) for base, ans in cols if (pd.isna(row[base]) or row[base] == 0) and not (pd.isna(row[ans]) or row[ans] == 0)]

        base_ans = 'BASE_ANS'
        chk_ans = 'CHK_ANS'
        diff_ans = 'DIFF_ANS'
        chk_df[base_ans] = chk_df[filt].apply(lambda_ma_to_list, axis=1, qids=base)
        chk_df[chk_ans] = chk_df[filt].apply(lambda_ma_to_list, axis=1, qids=chkm)
        chk_df[diff_ans] = chk_df[filt].apply(diff_ans_update, axis=1, cols=zip_cols)
        
        
        err_list += [err_col, base_ans, chk_ans, diff_ans]

        chk_df = chk_df if cond is None else chk_df[cond.reindex(chk_df.index, fill_value=False)]
        
        edf = ErrorDataFrame(f"""{chkm[0]}-{chkm[-1]}(MA) in {base[0]}-{base[-1]}(MA)""", 'MAMA', show_cols, chk_df, err_list, warnings, alt)
        self.show_message(edf)
        self.result_html_update(alt=self.result_alt(qid, alt), result_html=edf.chk_msg, dataframe=edf.err()[show_cols+edf.extra_cols].to_json())
        return edf

    def mark(self,
            base_qid: Union[List[str], Tuple[str]], 
            rank_qid: Union[List[str], Tuple[str]], 
            cond: Optional[pd.Series] = None, 
            diff_value: Optional[Union[List[Any], range, int, str]] = None,
            alt: Optional[str]=None) -> 'ErrorDataFrame' :
        """
        `복수 응답`을 베이스로 하는 `순위 응답` 로직 체크.
        `base_qid` (Union[List[str], Tuple[str]]): 기준이 되는 복수 응답 열 목록.
        `cond` (pd.Series, optional): 조건을 적용할 시리즈. 기본값은 None.
        `rank_qid` (Union[List[str], Tuple[str]]): 체크할 순위 응답 열 목록.
        """
        if (self.ma_check(base_qid)) or (self.ma_check(rank_qid)) :
            return
        warnings = []
        err_list = []
         
        chk_df = self[self.attrs['default_filter']].copy()
        base = self.ma_return(base_qid)
        rank = self.ma_return(rank_qid)
        max_rank = len(rank)
        if not self.col_name_check(*base): return
        if not self.col_name_check(*rank): return

        qid_key = get_key_id(base)

        show_cols = rank

        dv = []
        if diff_value is not None:
            if not isinstance(diff_value, (list, range, int, str)):
                display(HTML("""<div class="check-bold check-fail">❌ [ERROR] Type of diff_value must be list, range, int, or str</div>"""))
                return
            if isinstance(diff_value, (int, str)) :
                dv = [diff_value]
            if isinstance(diff_value, list) :
                dv = diff_value
            if isinstance(diff_value, range):
                dv = list(diff_value)
                dv.append(dv[-1] + 1)
            
            warnings.append(f"""Do not check the code : {dv}""")
            base = [x for x in base if not x in [f'{qid_key}{d}' for d in dv]]

        base_cnt = 'BASE_CNT'
        chk_df[base_cnt] = chk_df[base].apply(lambda x: x.count() - (x==0).sum(), axis=1)

        filt = chk_df[base_cnt]>=1

        base_col = 'BASE_COND'
        if cond is not None :
            chk_df.loc[cond, base_col] = 1
            err_list.append(base_col)
            filt = (filt) & (cond)

        err_col = 'LOGIC'
        # MA Base MA
        if len(chk_df[filt]) == 0 :
            display(HTML("""<div class="check-bold check-warn">⚠️ No data available for verification</div>"""))
            return

        def base_ans_update(row) :
            return [x for x in base if not (pd.isna(row[x]) or row[x] == 0)]


        def ma_base_rank_check(x) :
            able_ans = max_rank if x[base_cnt] > max_rank else x[base_cnt]
            chk_rank = rank[:able_ans]
            return 1 if any(pd.isna(x[rk]) for rk in chk_rank) else np.nan

        chk_df[err_col] = chk_df[filt].apply(ma_base_rank_check, axis=1)

        base_ans = 'BASE_ANS'
        chk_df[base_ans] = chk_df[filt][base].apply(base_ans_update, axis=1)

        # Each Rank masa
        rank_err_list = []
        for rk in rank :
            def ma_base_check(x) :
                sa_ans = x[rk]
                ma_var = f'{qid_key}{sa_ans}'
                ma_ans = x[ma_var]
                if sa_ans in dv :
                    return np.nan

                return 1 if pd.isna(ma_ans) or ma_ans == 0 else np.nan
            rk_err = f'{rk}_ERR'
            chk_df[rk_err] = chk_df[(filt) & (~chk_df[rk].isna())].apply(ma_base_check, axis=1)
            rank_err_list.append(rk_err)

        masa_err = 'MA_BASE_ERR'
        def masa_rank_err(x) :
            if any(x[err]==1 for err in rank_err_list) :
                return [cnt for cnt, rank in enumerate(rank_err_list, 1) if x[rank]==1]
            else :
                return np.nan

        chk_df[masa_err] = chk_df[filt].apply(masa_rank_err, axis=1)
        chk_df.loc[~chk_df[masa_err].isna(), err_col] = 1
        
        show_cols = [base_cnt, masa_err] + rank_err_list + rank + base
        err_list += [err_col, base_ans]

        chk_df = chk_df if cond is None else chk_df[cond.reindex(chk_df.index, fill_value=False)]
        
        edf = ErrorDataFrame(f"""{rank[0]}-{rank[-1]}(RANK) in {base[0]}-{base[-1]}(MA)""", 'MARK', show_cols, chk_df, err_list, warnings, alt)
        self.show_message(edf)
        self.result_html_update(alt=self.result_alt(qid, alt), result_html=edf.chk_msg, dataframe=edf.err()[show_cols+edf.extra_cols].to_json())
        return edf


    def rate_rank(self,
                  rate_qid: Union[List[str], Tuple[str]], 
                  rank_qid: Union[List[str], Tuple[str]],
                  cond: Optional[pd.Series] = None,
                  alt: Optional[str]=None)  -> 'ErrorDataFrame' :
        """
        `척도 응답`을 베이스로 하는 `순위 응답` 로직 체크.
        ()`척도 응답`의 점수 기준으로 `순위 응답`이 순서대로 응답되어야 하는 경우)
        `rate_qid` (Union[List[str], Tuple[str]]): 기준이 되는 복수 응답 열 목록.
        `rank_qid` (Union[List[str], Tuple[str]]): 체크할 순위 응답 열 목록.
        `cond` (pd.Series, optional): 조건을 적용할 시리즈. 기본값은 None.
        """
        if (self.ma_check(rate_qid)) or (self.ma_check(rank_qid)) :
            return
        warnings = []
        err_list = []
         
        chk_df = self[self.attrs['default_filter']].copy()
        rate = self.ma_return(rate_qid)
        rank = self.ma_return(rank_qid)
        if not self.col_name_check(*rate): return
        if not self.col_name_check(*rank): return

        qid_key = get_key_id(rate_qid)

        base_col = 'BASE_COND'
        filt = (~chk_df[rank].isna()).any(axis=1)
        if cond is not None :
            chk_df.loc[cond, base_col] = 1
            err_list.append(base_col)
            filt = (filt) & (cond)

        err_col = 'LOGIC'
        def rate_rank_validate(row, rate_base, rank_base):
            scores = {int(x.replace(qid_key, '')): row[x] for x in rate_base}
            result = {}
            for key, value in scores.items():
                if value not in result:
                    result[value] = []
                result[value].append(key)
            
            sort_score = [[key, result[key]] for key in sorted(result.keys(), reverse=True)]
            
            rk = rank_base.copy()
            is_valid = False
            for sc, able in sort_score :
                if not rk : break
                for idx in range(0, len(able)) :
                    if idx < len(rank_base) :
                        chk = rank_base[idx]
                        if not chk in rk : break
                        
                        rk.remove(chk)
                        if not row[chk] in able :
                            is_valid = True
                            break

            return 1 if is_valid else np.nan


        chk_df[err_col] = chk_df[filt].apply(rate_rank_validate, axis=1, rate_base=rate, rank_base=rank)


        def rate_rank_able_attrs(row, rate_base):
            scores = {int(x.replace(qid_key, '')): row[x] for x in rate_base}
            result = {}
            for key, value in scores.items():
                if value not in result:
                    result[value] = []
                result[value].append(key)
            
            sort_score = {int(key): result[key] for key in sorted(result.keys(), reverse=True)}
            
            return sort_score
        
        able_col = 'SCORE_ATTR'
        chk_df[able_col] = chk_df[filt].apply(rate_rank_able_attrs, axis=1, rate_base=rate)


        err_list.append(err_col)
        show_cols = [able_col] + rank + rate

        chk_df = chk_df if cond is None else chk_df[cond.reindex(chk_df.index, fill_value=False)]
        
        edf = ErrorDataFrame(f"""{rank[0]}-{rank[-1]}(RANK) / {rate[0]}-{rate[-1]}(RATE)""", 'RATERANK', show_cols, chk_df, err_list, warnings, alt)
        self.show_message(edf)
        self.result_html_update(alt=self.result_alt(qid, alt), result_html=edf.chk_msg, dataframe=edf.err()[show_cols+edf.extra_cols].to_json())
        return edf
    
    def lp(self, print_word: str) -> None:
        """
        별도 표시를 위한 메서드
        """
        if self.attrs['display_msg'] ==  'all' :
            display(HTML(f"""
                         <div class="datacheck-print-mw">
                            <div class="datacheck-logic-print">{print_word}</div>
                         </div>
                         """))

    def lchk(self) -> None:
        """
        LIVE 상태에서 검토해야하는 부분 표기
        """
        if self.attrs['display_msg'] ==  'all' :
            display(HTML(f"""
                        <div class="datacheck-print-mw">
                            <div class="datacheck-live-check">LIVE CHECK</div>
                        </div>
                         """))

    def qset(self, qid: str, code: Union[range, List]) -> List :
        """
        `qid`와 `code`를 기준으로 `DataFrame`에서 변수명 추출
        `(startswith(qid) and endswith(each code))`
        `qid` (str) : 문자열로 된 기준 변수명 ('SQ1', 'SQ2')
        `code` (range, list) : 각 변수의 속성 코드 (`[1, 2, 3, 4]`)
        example) qid='SQ7' / code=[1, 3, 5]
        return `['SQ7r1', 'SQ7r3', 'SQ7r5']`
        """
        cols = self.columns
        if not isinstance(code, (range, list)) :
            display(HTML(f"""<div class="check-bold check-fail">❌ The argument code can only be a list or range</div>"""))
            return []

        if any(not isinstance(c, int) for c in code) :
            display(HTML(f"""<div class="check-bold check-fail">❌ The argument code can only be numeric</div>"""))
            return []
        
        chk_code = code
        if isinstance(code, range) :
            chk_code.append(chk_code[-1]+1)

        filt = [col for col in cols if re.match(rf'^{qid}(?!\d)', col) and any(col.endswith(str(c)) for c in chk_code)]
        if not filt :
            display(HTML("""<div class="check-bold check-warn">⚠️ The variable does not exist in the dataframe</div>"""))
        return filt



#### Decipher Ready
def unzip_and_delete(zip_path, extract_to='.'):
    """
    Function to unzip a file and delete the zip file

    Parameters:
    zip_path (str): Path to the zip file
    extract_to (str): Directory path where the contents will be extracted (default: current directory)
    """
    try:
        # Open the zip file
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            # Extract all the contents of the zip file to the specified directory
            zip_ref.extractall(extract_to)
        
        # Delete the zip file
        os.remove(zip_path)
    
    except FileNotFoundError:
        print(f"File not found: {zip_path}")
    
    except zipfile.BadZipFile:
        print(f"Invalid zip file: {zip_path}")
    
    except Exception as e:
        print(f"An error occurred: {e}")

def ensure_directory_exists(directory_path: str) -> None:
    if not os.path.exists(directory_path):
        os.makedirs(directory_path)

def not_empty_cell(cell, row: int) -> bool:
    """셀 값이 비어 있지 않은지 확인합니다."""
    a = cell.cell(row, 1).value
    b = cell.cell(row, 2).value
    c = cell.cell(row, 3).value

    return bool(a or b or c)

def re_big(txt: str) -> Optional[str]:
    """대괄호 안의 내용을 추출합니다."""
    re_chk = re.search(r'\[(.*?)\]', txt)
    if re_chk:
        return re_chk.group(1).strip()
    return None

def colon_split(txt: str, num: int) -> Optional[str]:
    """콜론으로 텍스트를 나누고 지정된 부분을 반환합니다."""
    re_chk = txt.split(":")
    if re_chk:
        return re_chk[num].strip()
    return None

def DecipherSetting(pid: str, 
            mode: str = 'auto', 
            cond: Optional[str] = None,
            use_variable: bool = False,
            key: str = api_key, 
            server: str = api_server, 
            json_export: bool = True, 
            data_layout: bool = False, 
            base_layout: str = 'DoNotDelete',
            datamap_name: str = 'Datamap',
            mkdir: bool = False,
            dir_name: Optional[str] = None) -> None:

    """
    데이터 체크 노트북 파일 및 데이터 세팅
    
    Args:
        `pid` (str): 프로젝트 ID.
        `mode` (str, optional): 모드 ('auto' 또는 'file'). 기본값은 'auto'.
        `cond` (str, optional): 데이터 필터링 조건. 기본값은 None.
        `use_variable` (bool, optional): 변수 파일 사용 여부. 기본값은 False.
        `key` (str, optional): API 키. 기본값은 api_key.
        `server` (str, optional): API 서버. 기본값은 api_server.
        `json_export` (bool, optional): JSON 내보내기 여부. 기본값은 True.
        `data_layout` (bool, optional): 데이터 레이아웃 내보내기 여부. 기본값은 False.
        `base_layout` (str, optional): 기본 레이아웃 이름. 기본값은 'DoNotDelete'.
        `datamap_name` (str, optional): 데이터 맵 이름. 기본값은 'Datamap'.
        `mkdir` (bool, optional): 디렉토리 생성 여부. 기본값은 False.
        `dir_name` (str, optional): 디렉토리 이름. 기본값은 None.
    """

    #pd.io.formats.excel.ExcelFormatter.header_style = None
    excel.ExcelFormatter.header_style = None
    
    if pid == '' or not pid :
        print('❌ Please enter pid')
        return

    if not mode in ['auto', 'file'] :
        print('❌ Please check the mode argument (auto or file)')
        return

    parent_path = os.getcwd()
    if mkdir :
        folder_name = pid
        if dir_name != None :
            folder_name = dir_name
        parent_path =  os.path.join(parent_path, folder_name)
        chk_mkdir(parent_path)

    if mode == 'file' :
        file_name = f'{pid}.xlsx'
        xl = openpyxl.load_workbook(file_name)
        map_sheet = datamap_name
        data_map = xl[map_sheet]
        print('📢 Read excel file (xlsx)')

    if mode == 'auto' :
        file_name = f'{pid}.csv'
        if cond != None :
            if cond.isdigit() :
                print('❌ [ERROR] : The cond argument can only be a string')
                return
        delivery_cond = 'qualified' if cond == None else f'qualified and {cond}'
        try :
            api.login(key, server)
        except :
            print('❌ Error : Decipher api login failed')
            return

        path = f'surveys/selfserve/548/{pid}'
        # get csv data
        try :
            csv_data = api.get(f'{path}/data', format='csv', cond=delivery_cond)
            sav_data = api.get(f'{path}/data', format='spss16', cond=delivery_cond)
        except :
            print('❌ Error : Please check the cond argument')
            return

        csv_binary = f'binary_{pid}.csv'
        ensure_directory_exists('data')
        data_path = os.path.join(parent_path, 'data')
        create_binary_file(data_path, csv_binary, csv_data)
        create_ascii_file(data_path, csv_binary, f'{pid}.csv')
        
        sav_zip = f'{pid}_sav.zip'
        create_binary_file(data_path, sav_zip, sav_data)
        unzip_and_delete(os.path.join(data_path, sav_zip), data_path)
        time.sleep(3)

        # get datamap xlsx

        map_xlsx = api.get(f'{path}/datamap', format='xlsx')
        
        ensure_directory_exists('map')
        map_path = os.path.join(parent_path, 'map')
        create_binary_file(map_path, f'mapsheet_{pid}.xlsx', map_xlsx)

        xl = openpyxl.load_workbook(os.path.join(map_path, f'mapsheet_{pid}.xlsx'))
        map_sheet = 'datamap'
        data_map = xl[map_sheet]

        print('📢 Using Decipher REST API')

    mx_row = data_map.max_row
    mx_col = data_map.max_column

    key_ids = key_vars
    diff_vars = sys_vars
    all_diff = key_ids + diff_vars

    rank_chk = ['1순위', '2순위', '1st', '2nd']

    na = 'noanswer'
    eltxt = 'element'
    col_name = ["a", "b", "c"]
    curr_var = {col:[] for col in col_name }

    variables = []
    
    #print("  ❌ DataCheck Setting Start")
    for row in range(1, mx_row+1) :
        if not_empty_cell(data_map, row) :
            for idx, col in enumerate(range(1, mx_col+1)) :
                curr_col = col_name[idx]
                curr_dict = curr_var[curr_col]
                cell = data_map.cell(row, col)
                if cell.value or cell.value == 0: 
                    curr_dict.append(cell.value)
                    curr_var[curr_col] = curr_dict
                if cell.value == None or cell.value == "" :
                    curr_dict.append("")
                    curr_var[curr_col] = curr_dict
        else :
            variables.append(curr_var)
            curr_var = {col:[] for col in col_name }


    qids = OrderedDict()
    for variable in variables :
        # qid, type summary
        a_cell = variable['a']
        a_cell = [a for a in a_cell if a != '' and a != None]
        b_cell = variable['b']
        #b_cell = [b for b in b_cell if b != '' and b != None]
        c_cell = variable['c']
        #c_cell = [c for c in c_cell if c != '' and c != None]
        qid = a_cell[0] # qid

        # print(qid)
        # print(b_cell)
        # print(c_cell)

        type_chk = a_cell[1] if len(a_cell) >= 2 else None # type check

        # attribute
        main_qlabel = None
        qtype = None
        qelements = []
        qvalue = None
        qtitle = None

        # labels setting
        qlabes = {
            'values' : {b:c_cell[idx] for idx, b in enumerate(b_cell) if type(b) == int}  if c_cell else {},
            'texts' : {re_big(b):c_cell[idx] for idx, b in enumerate(b_cell) if type(b) == str or b == None} if c_cell else {}
        }
        #  find name in []

        # main q label check
        find_qname = re_big(qid.split(':')[0])
        if find_qname :
            main_qlabel = find_qname
            qelements.append(main_qlabel)
            qtitle = colon_split(qid, 1)
        else :
            main_qlabel = colon_split(qid, 0)
            qtitle = colon_split(qid, 1)

        # type check
        if type_chk :
            open_text = 'OPEN'
            if open_text.upper() in type_chk.upper() :
                qtype = 'OE'
            else :
                qtype = 'CE'

            # value check
            value_text = 'VALUES'
            if value_text.upper() in type_chk.upper() :
                qvalue = colon_split(type_chk, 1)
            else :
                qvalue = None


        else :
            qtype = 'OTHER'

        # other oe check
        oe_chk = 'oe'
        if oe_chk in main_qlabel :
            qtype = 'OTHER_OE'

        # elements setting
        for b in b_cell :
            b_chk = re_big(str(b))
            if b_chk :
                qelements.append(b_chk)

        if not 'OE' in qtype:
            # ma check
            unchk = 'Unchecked'
            c_chk = [c.upper() for c in c_cell if type(c) == str]
            if unchk.upper() in c_chk : 
                qtype = 'MA'
            else :
                # radio/number check
                int_chk = [b for b in b_cell if type(b) == int]
                if int_chk :
                    qtype = 'SA'
                else :
                    qtype = 'NUM'

        if len(qelements) >= 2 :
            if main_qlabel in qelements :
                qelements.remove(main_qlabel)

        el_labels = {key:value for key, value in qlabes['texts'].items() if key}
        
        qids[main_qlabel] = {
            'element' : qelements,
            'title' : qtitle,
            'type' : qtype,
            'value' : qvalue,
            'value_label' : qlabes['values'],
            'element_label' : el_labels,
        }

    if na in list(qids.keys()) :
        nas = qids[na]
        els = nas[eltxt]
        for el in els :
            na_name = el.split('_')[:-1]
            na_el = '_'.join(na_name).replace(na, '')
            if not na_el == '' :
                if qids[na_el] :
                    qel = qids[na_el][eltxt]
                    qel.append(el)
                    qids[na_el][eltxt] = qel

    # print(qids)

    # default setting
    nb = nbf.v4.new_notebook()
    ipynb_file_name = f'DataCheck_{pid}.ipynb'
    order_qid = list(qids.items())

    # json export
    if json_export :
        with open(os.path.join(map_path, f'map_{pid}.json'), 'w', encoding='utf-8') as f :
            json.dump(qids, f, ensure_ascii=False, indent=4)

    # print(qids)
    # data layout export
    if data_layout :
        try :
            api.login(key, server)
            survey = f'selfserve/548/{pid}'
            url = f'surveys/{survey}/layouts'
            map = api.get(url)
        except :
            print('❌ Error : Decipher API failed')
        
        maps = [m for m in map if m['description'] == base_layout ]
        if not maps :
            print(f'❌ Error : The base layout({base_layout}) is null')
            return 
        base_map = maps[0]

        variables = base_map['variables']
        # print(variables)
        exactly_diff_vars = key_vars + sys_vars
        ce_vars = ['radio', 'checkbox', 'number', 'float', 'select']
        oe_vars = ['text', 'textarea']
        diff_label_names = ['vqtable', 'voqtable', 'dummy', 'DUMMY', 'Dummmy']
        
        ensure_directory_exists('layout')
        layout_path = os.path.join(parent_path, 'layout')
        ce = open(os.path.join(layout_path, f'CE_{pid}.txt'), 'w')
        oe = open(os.path.join(layout_path, f'OE_{pid}.txt'), 'w')

        for label, width in [ ('record', 7), ('uuid', 16) ]:
            write_text = f'{label},{label},{width}\n'
            ce.write(write_text)
            oe.write(write_text)

        resp_chk = [v for v in variables if v['label'] == 'RespStatus']
        if resp_chk :
            ce.write(f'RespStatus,RespStatus,8\n')

        for var in variables :
            label = var['label']
            qlabel = var['qlabel']
            qtype = var['qtype']
            fwidth = var['fwidth']
            altlabel = var['altlabel']
            shown = var['shown']
            if not shown :
                continue

            write_text = f'{label},{altlabel},{fwidth}\n'
            if (not label in exactly_diff_vars and not qlabel in exactly_diff_vars) :
                if [dl for dl in diff_label_names if (dl in label) or (dl in qlabel)] :
                    continue
                if qtype in ce_vars :
                    if qtype in ['number', 'float'] :
                        verify_check = [attr['value'].split('-')[1] for ql, attr in list(qids.items()) if (ql == qlabel) or (ql == label)]
                        if verify_check :
                            max_width = len(verify_check[0])
                                # print(label, verify_check, max_width)
                            if qtype == 'float' :
                                max_width += 4
                            write_text = f'{label},{altlabel},{max_width}\n'
                    ce.write(write_text)
                if qtype in oe_vars :
                    oe.write(write_text)

        oe.write(f'decLang,decLang,60\n')
        # if resp_chk :
        #     oe.write(f'RespStatus,RespStatus,8\n')

        ce.close()
        oe.close()

    # variable py file create
    variable_py_name = f'variables_{pid}.py'
    py_file = open(os.path.join(map_path, variable_py_name), 'w')
    py_file.write(f'# {pid} variables\n')

    ipynb_cell = []

    # set_file_name = 'pd.read_excel(file_name)' if mode == 'file' else 'pd.read_csv(file_name, low_memory=False)'

    default = f'''import pandas as pd
import pyreadstat
import numpy as np
from map.variables_{pid} import * 
from decipherAutomatic.dataProcessing import *
css_apply() # CSS APPLY

# Use Excel
# file_name = "data/{pid}.xlsx"
# df = DataCheck(pd.read_excel(file_name, engine="openpyxl"), keyid="record")

# Use SPSS
file_name = "data/{pid}.sav"
df, meta = pyreadstat.read_sav(file_name)
df = DataCheck(df, keyid="record", spssmeta=meta)
'''
    
    ipynb_cell.append(nbf.v4.new_code_cell(default))
    ipynb_cell.append(nbf.v4.new_code_cell("""# df.display_msg = 'error'"""))

    for idx, attrs in enumerate(order_qid) :
        qid = attrs[0]
        els = attrs[1]
        if not qid in all_diff :
            qels = els['element']
            qtype = els['type']
            qval = els['value']
            qtitle = els['title']
            val_label = els['value_label']
            el_label = els['element_label']

            cell_texts = []
            cell_texts.append(f'# {qid} : {qtype}')
            # sa check #
            if qtype == 'SA' or (qtype == 'MA' and len(qels) == 1):
                if qtype == 'SA' :
                    if qval :
                        val_chk = f"# value : {qval}"
                        cell_texts.append(val_chk)

                    if len(qels) >= 2 :
                        diff_na = [q for q in qels if not na in q]
                        py_file.write(f"{qid} = {diff_na}\n")

                    for qel in qels :
                        if na in qel :
                            cell_texts.append(f'# The {qid} contains {qel}')
                        else :
                            safreq = f"df.safreq('{qel}')"
                            if use_variable : safreq = f"df.safreq({qel})"

                            py_file.write(f"{qel} = '{qel}'\n")
                            cell_texts.append(safreq)

                    if val_label :
                        values = [v for v in val_label.keys() if not int(v) == 0]
                        py_file.write(f'{qid}_value = {values}\n')
                        #values_txt = f'{qid}_values = {values}'
                        #cell_texts.append(values_txt)

                    # rank check
                    if len(qels) >= 2 :
                        labels = list(el_label.values())
                        rk = []
                        for rk_txt in rank_chk :
                            for label in labels :
                                mu_rk = rk_txt.strip().replace(' ','').upper()
                                mu_label = label.strip().replace(' ','').upper()
                                if mu_rk in mu_label :
                                    rk.append(label)
                        if len(rk) >= 2 :
                            dup_diff_na = [q for q in qels if not na in q]
                            set_qid = f"('{dup_diff_na[0]}', '{dup_diff_na[-1]}')"

                            py_file.write(f"{qid} = {dup_diff_na}\n")
                            #cell_texts.append(f'{qid} = {set_qid}')
                            dupchk = f"df.dupchk({set_qid})"
                            if use_variable : dupchk = f"df.dupchk({qid})"

                            cell_texts.append(dupchk)
                else :
                    if qval :
                        val_chk = f"# value : {qval}"
                        py_file.write(f"{qid} = '{qid}'\n")
                        py_file.write(f'{qid}_value = [0, 1]\n')
                        
                        cell_texts.append(val_chk)
                        safreq = f"df.safreq('{qels[0]}')"
                        if use_variable : safreq = f"df.safreq({qels[0]})"
                        cell_texts.append(safreq)
            ### sa end ###

            # ma check #
            elif qtype == 'MA' :
                if len(qels) > 1 :
                    diff_na = [q for q in qels if not na in q]
                    if not diff_na :
                        continue
                    nas = [q for q in qels if na in q]
                    first_el = diff_na[0]
                    last_el = diff_na[-1]
                    set_qid = f"('{first_el}', '{last_el}')"

                    for q in diff_na :
                        py_file.write(f"{q} = '{q}'\n")

                    py_file.write(f"{qid} = {diff_na}\n")

                    if val_label :
                        values = [v for v in val_label.keys() if not int(v) == 0]
                        if not values == [1] :
                            py_file.write(f'{qid}_value = {values}\n')
                            #values_txt = f'{qid}_values = {values}'
                            #cell_texts.append(values_txt)
                        else :
                            py_file.write(f'{qid}_value = [0, 1]\n')
                    # cell_texts.append(f'{qid} = {set_qid}')

                    mafreq = f"df.mafreq({set_qid})"
                    if use_variable : mafreq = f"df.mafreq({qid})"

                    cell_texts.append(mafreq)

                    if nas :
                        cell_texts.append(f'# The {qid} contains {nas}')
            ### ma end ###


            # num check #
            elif qtype == 'NUM' :
                range_set = None

                if len(qels) >=2 :
                    diff_na = [q for q in qels if not na in q]
                    py_file.write(f"{qid} = {diff_na}\n")

                if qval :
                    values = qval.split('-')
                    range_set = f"only=range({values[0]}, {values[1]})"

                for qel in qels :
                    if na in qel :
                        cell_texts.append(f'# The {qid} contains {qel}')
                    else :
                        if range_set :
                            safreq = f"df.safreq('{qel}', {range_set})"
                            if use_variable : safreq = f"df.safreq({qel}, {range_set})"
                        else :
                            safreq = f"df.safreq('{qel}')"
                            if use_variable : safreq = f"df.safreq({qel})"

                        py_file.write(f"{qel} = '{qel}'\n")
                        cell_texts.append(safreq)

            ### num end ###

            # text check #
            elif qtype == 'OE' :
                if len(qels) >=2 :
                    diff_na = [q for q in qels if not na in q]
                    py_file.write(f"{qid} = {diff_na}\n")

                for qel in qels :
                    if na in qel :
                        cell_texts.append(f'# The {qid} contains {qel}')
                    else :
                        safreq = f"df.safreq('{qel}')"
                        if use_variable : safreq = f"df.safreq({qel})"

                        py_file.write(f"{qel} = '{qel}'\n")
                        cell_texts.append(safreq)
            ### text end ###

            # other open check #
            elif qtype == 'OTHER_OE' :
                for qel in qels :
                    safreq = f"df.safreq('{qel}')"
                    if use_variable : safreq = f"df.safreq({qel})"

                    py_file.write(f"{qel} = '{qel}'\n")
                    cell_texts.append(safreq)
            ### other open end ###


            if cell_texts :
                cell = '\n'.join(cell_texts)
                ipynb_cell.append(nbf.v4.new_code_cell(cell))
            else :
                mark = f'The {qid} not cotains elements'
                ipynb_cell.append(nbf.v4.new_markdown_cell(mark))

            py_file.write(f'\n')

    py_file.close()
    #ipynb_cell
    nb['cells'] = ipynb_cell
    #print(nb)
    ipynb_file_path = os.path.join(parent_path, ipynb_file_name)
    if not os.path.isfile(ipynb_file_path) :
        with open(ipynb_file_path, 'w') as f:
            nbf.write(nb, f)
    else :
        print('❗ The DataCheck ipynb file already exists')
    
    print("✅ DataCheck Setting Complete")