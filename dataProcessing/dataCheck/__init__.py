import pandas as pd
from pandas.api.types import is_numeric_dtype
from IPython.display import display, HTML
from typing import Union, List, Tuple, Dict, Optional, Literal, Callable, Any, TypedDict
import numpy as np
from dataclasses import dataclass, field
import contextlib
import os
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import re
import nbformat as nbf
from collections import OrderedDict, defaultdict
import json
from decipher.beacon import api
import time
from decipherAutomatic.key import api_key, api_server
from decipherAutomatic.getFiles import *
from decipherAutomatic.utils import *
from decipherAutomatic.dataProcessing.table import *
from pandas.io.formats import excel
import zipfile
from matplotlib.colors import to_hex
from matplotlib.colors import LinearSegmentedColormap
from pprint import pprint
import tempfile
from datetime import datetime

VarWithHeader = Tuple[str, Union[str, List[str]]]
ColumnsWithHeader = List[VarWithHeader]
IndexWithTypes = Union[str, List[str], Tuple[str], Dict]
Qtypes = Literal['single', 'rating', 'rank', 'ranksort', 'multiple', 'number', 'float', 'text']

def check_print(variables: Union[List[str], Tuple[str, ...], str], 
                error_type: Literal['SA', 'MA', 'LOGIC', 'MASA', 'MAMA', 'MARK', 'RATERANK', 'DUP'], 
                df: pd.DataFrame, 
                warnings: Optional[List[str]] = None,
                alt: Optional[str] = None) -> str:
    qid = None
    
    if isinstance(variables, str): 
        qid = variables

    if isinstance(variables, list) or isinstance(variables, tuple) :
        list_vars = list(variables)
        if len(list_vars) == 1 :
            qid = list_vars[0]
        if len(list_vars) >= 2 :
            qid = f'{list_vars[0]} - {list_vars[-1]}'
    
    error_type_msg = {
        'SA': 'SA Variable Check',
        'MA': 'MA Variable Check',
        'LOGIC': 'If Base cond is True, then Answer cond is also True',
        'DUP': 'Answer Duplicate Check',
        'MASA': 'Multi Variable Base Single Variable Check',
        'MAMA': 'Multi Variable Base Multi Variable Check',
        'MARK': 'Multi Variable Base Rank Variable Check',
        'RATERANK' : 'Check if Rank is answered in order of Rate question score'
    }


    print_str = ''
    print_str += f"""<div class="datcheck-title">📢 <span class="title-type">{error_type}</span> <span class="title-msg">({error_type_msg[error_type]})</span></div>""" # Error Text Title
    print_str += f"""👨‍👩‍👧‍👦 <span class="print-comment">Check Sample : <span class="check-bold">{len(df)}'s</span></span>"""

    # Result HTML
    correct = """<div class="datacheck-head check-correct">✅ {html_title}</div>"""
    fail    = """<div class="datacheck-head check-fail">❌ {html_title} : Error {err_cnt}'s</div>"""
    check   = """<div class="datacheck-check">📌 <span class="print-comment">{check_title}</span></div>"""
    warning  = """<div class="datacheck-warning check-warn">⚠️ {warn_title}</div>"""

    # Base Check
    err_cols = df.columns


    if warnings is not None :
        for warn in warnings :
            print_str += warning.format(warn_title=warn)

    ms_err = 'DC_BASE'
    if ms_err in err_cols :
        err_cnt = len(df[df[ms_err]==1])
        html_title = f"""Answer Base Check"""
        if err_cnt == 0 :
            print_str += correct.format(html_title=html_title)
        else :
            print_str += fail.format(html_title=html_title, err_cnt=err_cnt)

    # Cases responded to other than base 
    add_err = 'DC_NO_BASE'
    if add_err in err_cols :
        err_cnt = len(df[df[add_err]==1])
        html_title = "Other than Base Check"
        if err_cnt >= 1:
            print_str += fail.format(html_title=html_title, err_cnt=err_cnt)

    # Answer Able Value Check  / SA 
    only_err = 'ONLY_ANS'
    if only_err in err_cols :
        err_cnt = len(df[df[only_err]==1])
        html_title = "Answer Able Value Check"
        if err_cnt == 0 :
            print_str += correct.format(html_title=html_title)
        else :
            print_str += fail.format(html_title=html_title, err_cnt=err_cnt)
        
        err_answer = list(df[df[only_err]==1][variables].values)
        err_answer = ['NA' if pd.isna(x) else x for x in err_answer]
        err_answer = sorted(set(err_answer))
        if err_answer :
            print_str += f"""<div class="print-padding-left">🗒️ <span class="print-comment">Invalid response</span> : {list(err_answer)}</div>"""

    # Disable Value Check  / SA 
    isnot_err = 'ISNOT_ANS'
    if isnot_err in err_cols :
        err_cnt = len(df[df[isnot_err]==1])
        html_title = "Answer Is Not Value Check"
        if err_cnt == 0 :
            print_str += correct.format(html_title=html_title)
        else :
            print_str += fail.format(html_title=html_title, err_cnt=err_cnt)
        
        err_answer = list(set(list(df[df[isnot_err]==1][variables].values)))
        err_answer = sorted(set(err_answer))
        if err_answer :
            print_str += f"""<div class="print-padding-left">🗒️ <span class="print-comment">Invalid response</span> : {list(err_answer)}</div>"""

    # MA Variable Answer Count Check
    if (error_type in ['MA']) :
        for lg in ['DC_ATLEAST', 'DC_ATMOST', 'DC_EXACTLY'] :
            if not lg in err_cols :
                continue
            err_cnt = len(df[df[lg]==1])
            html_title = f"{lg} Check"
            if err_cnt == 0 :
                print_str += correct.format(html_title=html_title)
            else :
                print_str += fail.format(html_title=html_title, err_cnt=err_cnt)
        
        for isx in ['MA_ISIN', 'MA_ISALL', 'MA_ISNOT'] :
            if isx in list(df.columns) :
                err_cnt = len(df[df[isx]==1])
                ma, istype = isx.split('_')
                html_title = f"{ma} {istype.capitalize()} Answer Check"
                if err_cnt == 0 :
                    print_str += correct.format(html_title=html_title)
                else :
                    print_str += fail.format(html_title=html_title, err_cnt=err_cnt)

    # Answer Description Print
    desc_table = None
    if (error_type in ['SA']) and (pd.api.types.is_numeric_dtype(df[qid])) :
        desc = df[qid].describe().round(1)
        if not pd.isna(desc.loc['mean']) :
            desc_table = """
        <div class="datacheck-desc">📋 {qid} Describe</div>
        <table class="print-padding-left"">
            <tr><td><b>Count</b></td><td>{cnt}</td></tr>
            <tr><td><b>Mean</b></td><td>{mean}</td></tr>
            <tr><td><b>Min</b></td><td>{minv}</td></tr>
            <tr><td><b>Max</b></td><td>{maxv}</td></tr>
        </table>""".format(qid=qid, cnt=desc.loc['count'], mean=desc.loc['mean'], minv=desc.loc['min'], maxv=desc.loc['max'])

    if (error_type in ['MA']) :
        if 'ANSWER_CNT' in df.columns :
            desc = df['ANSWER_CNT'].describe().round(1)
            desc_table = """
        <div class="datacheck-desc">📋 {qid} Answer Count Describe</div>
        <table class="print-padding-left">
            <tr><td><b>Mean</b></td><td>{mean}</td></tr>
            <tr><td><b>Min</b></td><td>{minv}</td></tr>
            <tr><td><b>Max</b></td><td>{maxv}</td></tr>
        </table>""".format(qid=qid, mean=desc.loc['mean'], minv=desc.loc['min'], maxv=desc.loc['max'])

    if desc_table is not None :
        print_str += desc_table


    # Logic Check
    if (error_type in ['LOGIC', 'MASA', 'MAMA', 'MARK', 'RATERANK']) :
        if 'DC_LOGIC' in df.columns :
            err_cnt = len(df[df['DC_LOGIC']==1])
            base_cond = 'BASE_COND'
            if base_cond in list(df.columns) :
                base_cnt = len(df[df[base_cond]==1])
                # print_str += check.format(check_title=f"Base Cond Answer Count : <b>{base_cnt}'s</b>")
            if err_cnt == 0 :
                print_str += correct.format(html_title=f"Logic Correct")
            else :
                print_str += fail.format(html_title="Logic has Error", err_cnt=err_cnt)

    # Duplicate Check
    if (error_type in ['DUP']) :
        if 'DC_DUP' in df.columns :
            err_cnt = len(df[df['DC_DUP']==1])
            if err_cnt == 0 :
                print_str += correct.format(html_title="No Duplicate")
            else :
                dup_rows = df[df['DC_DUP'] == 1]
                summary = []
                
                for index, row in dup_rows.iterrows():
                    row_values = row[variables]
                    duplicates = row_values[row_values.duplicated(keep=False)]
                    summary.extend(duplicates.unique().tolist())
                
                summary = list(set(summary))
                print_str += fail.format(html_title=f"Duplicate Answer", err_cnt=err_cnt)
                print_str += f"""<div class="print-padding-left">🗒️ <span class="print-comment">Invalid response</span> : {summary}</div>"""


    print_type = "alt-main"
    if "check-fail" in print_str :
        print_type = "alt-fail"

    final_print = f"""
    <div class="datacheck-print {print_type}">
        <div class="datacheck-alt">{alt if alt is not None else qid}</div>
        <div class="datacheck-result">
            {print_str}
        </div>
    </div>
    """

    return final_print

def classify_variables(variable_list: List[str]):
    # 딕셔너리 생성
    classified_vars = defaultdict(list)
    # 변수명의 규칙을 추출하는 정규식
    pattern = re.compile(r"([A-Za-z]+\d*[_]?[A-Za-z]*)")
    
    for variable in variable_list:
        match = pattern.match(variable)
        if match:
            key = match.group(1)
            classified_vars[key].append(variable)
    
    return dict(classified_vars)

def get_key_id(base: List[str]) -> Union[None, str]:
    """`base`가 `qid`를 포함하는지 확인합니다."""
    qid = base[0]
    qids = list(qid)
    qids.reverse()
    for q in qids:
        if not q.isdigit():
            break
        else:
            qid = qid[:-1]

    for ma in base:
        if qid not in ma:
            qid = None
            return classify_variables(base)
        
    return qid


def lambda_ma_to_list(row, qids) :
    qid_key = get_key_id(qids)

    def return_int_or_str(txt: str) :
        rp = txt.replace(qid_key, '')
        if rp.isdigit() :
            return int(rp)
        else :
            return rp
    
    return [return_int_or_str(x) for x in qids if not (pd.isna(row[x]) or row[x] == 0)]

def calculate_bg_color(value):
    cmap = LinearSegmentedColormap.from_list("custom_blue", ["#ffffff", "#2d6df6"])
    normalized_value = value / 100  # Normalize the value between 0 and 1
    return to_hex(cmap(normalized_value))

def check_duplicate_meta(original_dict: List[Dict[str, str]]) -> List[Dict[str, str]]:
    # 각 value의 발생 횟수를 계산
    value_count = {}
    for value in original_dict.values():
        if value in value_count:
            value_count[value] += 1
        else:
            value_count[value] = 1

    new_dict = {}
    value_rename_count = {}

    for key, value in original_dict.items():
        # 중복된 값이 있는 경우에만 리네임 진행
        if value_count[value] > 1:
            if value in value_rename_count:
                value_rename_count[value] += 1
            else:
                value_rename_count[value] = 1

            # "1_value", "2_value" 형식으로 리네임
            new_value = f"{value_rename_count[value]}_{value}"
        else:
            new_value = value

        new_dict[key] = new_value

    return new_dict


def extract_order_and_labels(metadata: Union[dict], front_variable: Optional[list] = None, back_variable: Optional[list] = None):
    """
    Extracts the order and labels from the provided metadata.
    
    Parameters:
        metadata (list of dict): The metadata to extract order and labels from.
        front_variable, back_variable : All / Total
    
    Returns:
        order (list): The extracted order of keys.
        labels (list): The extracted labels for the keys.
    """
    order = [d for d in metadata.keys()]
    if front_variable is not None :
        order = front_variable + order
    
    if back_variable is not None :
        order = order + back_variable
    
    labels = [d for d in metadata.values()]
    if front_variable is not None :
        labels = front_variable + labels
    
    if back_variable is not None :
        labels = labels + back_variable
    
    return order, labels

def add_missing_indices(df, order):
    """
    Adds missing indices with zero values to the DataFrame.
    
    Parameters:
        df (pd.DataFrame): The DataFrame to add missing indices to.
        order (list): The list of indices to ensure exist in the DataFrame.
    
    Returns:
        pd.DataFrame: The DataFrame with missing indices added.
    """
    
    for idx in order:
        if idx not in df.index :
            df.loc[idx] = 0
    return df

def reorder_and_relabel(df, order, labels, axis, name):
    """
    Reorders and relabels the DataFrame based on the provided order and labels.
    
    Parameters:
        df (pd.DataFrame): The DataFrame to reorder and relabel.
        order (list): The order to apply.
        labels (list): The labels to apply.
        axis (int): The axis to apply the reorder and relabel (0 for index, 1 for columns).
        name (str): The name to assign to the index or columns.
    
    Returns:
        pd.DataFrame: The DataFrame with reordered and relabeled indices/columns.
    """
    if axis == 0:
        df = df.loc[order]
        if name is not None and name is not False :
            df.index = pd.Index(labels, name=name)
        else :
            df.index = pd.Index(labels)
    else:
        df = df[order]
        if name is not None and name is not False :
            df.columns = pd.Index(labels, name=name)
        else :
            df.columns = pd.Index(labels)
    return df


@dataclass
class PrintDataFrame:
    show_cols: List[str]
    df: pd.DataFrame

    def __call__(self, extra: Optional[List[str]] = None):
        if extra is None:
            return self.df[self.show_cols]
        return self.df[self.show_cols + extra] if extra else self.df

@dataclass
class ErrorDataFrame:
    chk_id: str
    qid_type: str
    show_cols: List[str]
    df: pd.DataFrame
    err_list: List[str]
    warnings: List[str] = field(default_factory=list)
    alt: Optional[str] = None

    def __post_init__(self):
        self.show_col_with_err = self.err_list + self.show_cols
        self.err_base = [x for x in self.err_list if x not in ['BASE_COND', 'ANSWER_COND']]
        self.df[self.err_list] = self.df[self.err_list].where(self.df[self.err_list].notna(), 0).astype(int)
        err_df = self.df[(self.df[self.err_base]==1).any(axis=1)]
        self.err = PrintDataFrame(self.show_col_with_err, err_df)
        self.full = PrintDataFrame(self.show_col_with_err, self.df)
        self.chk_msg = check_print(self.chk_id, self.qid_type, self.full(), self.warnings, self.alt)
        self.extra_cols = []

    def __getitem__(self, key):
        extra_cols = [key] if isinstance(key, str) else key
        self.extra_cols = extra_cols
        return self.err()[self.base + extra_cols]

    def __repr__(self):
        return ''

class DataCheck(pd.DataFrame):
    _metadata = ['_keyid', '_css', '_meta_origin', '_meta', '_title', '_default_top', '_default_bottom', '_default_medium', '_default_with_value', '_default_chat_lang']
    def __init__(self, *args, **kwargs):
        self._keyid = kwargs.pop('keyid', None)
        self._css = kwargs.pop('css', None)
        self._meta = kwargs.pop('meta', None)
        self._title = kwargs.pop('title', None)
        self._default_top = kwargs.pop('default_top', None)
        self._default_bottom = kwargs.pop('default_bottom', None)
        self._default_medium = kwargs.pop('default_medium', None)
        self._default_with_value = kwargs.pop('default_with_value', None)
        self._default_chat_lang = kwargs.pop('default_chat_lang', None)
        
        super().__init__(*args, **kwargs)
        if self._keyid is not None:
            self[self._keyid] = self[self._keyid].astype(int)
            self.set_index(self._keyid, inplace=True)

        pd.set_option('display.max_columns', None)
        pd.set_option('display.max_rows', None)

        self._count_fnc: Callable[[pd.Series], int] = lambda x: x.count() - (x==0).sum()
        
        self.attrs['display_msg'] = 'all'
        self.attrs['default_filter'] = pd.Series([True] * len(self), index=self.index)
        self.attrs['result_html'] = []
        self.attrs['css'] = self._css
        self.attrs['meta_origin'] = self._meta if self._meta is not None else {}
        self.attrs['meta'] = self._meta if self._meta is not None else {}
        self.attrs['title_origin'] = self._title if self._title is not None else {}
        self.attrs['title'] = self._title if self._title is not None else {}
        self.attrs['default_top'] = 2 if self._default_top is None else self._default_top
        self.attrs['default_bottom'] = 2 if self._default_bottom is None else self._default_bottom
        self.attrs['default_medium'] = True if self._default_medium is None else self._default_medium
        self.attrs['default_with_value'] = False if self._default_with_value is None else self._default_with_value
        self.attrs['chat_lang'] = 'korean' if self._default_chat_lang is None else self._default_chat_lang
        self.attrs['nets'] = {}
        self.attrs['banner'] = None
        self.attrs['proc_result'] = OrderedDict()



    @property
    def _constructor(self) -> Callable[..., 'DataCheck']:
        return DataCheck

    def any(self, *args, **kwargs) -> pd.Series:
        """
        DataFrame의 any 메서드를 확장하여, 기본 축(axis)을 1로 설정
        """
        if 'axis' not in kwargs:
            kwargs['axis'] = 1
        return super().any(*args, **kwargs)

    def all(self, *args, **kwargs) -> pd.Series:
        """
        DataFrame의 all 메서드를 확장하여, 기본 축(axis)을 1로 설정
        """
        if 'axis' not in kwargs:
            kwargs['axis'] = 1
        return super().all(*args, **kwargs)

    @property
    def keyid(self) -> Optional[str]:
        """
        DataCheck 클래스의 keyid 속성을 반환
        """
        return self._keyid

    @keyid.setter
    def keyid(self, value: Optional[str]) -> None:
        """
        DataCheck 클래스의 keyid 속성을 설정
        """
        self._keyid = value

    @property
    def display_msg(self) -> Optional[Literal['all', 'error', None]]:
        """
        DataCheck 클래스의 display_msg 속성을 반환
        """
        return self.attrs['display_msg']

    @display_msg.setter
    def display_msg(self, option: Optional[Literal['all', 'error', None]]) -> None:
        """
        DataCheck 클래스의 display_msg 속성을 설정
        """
        if not option in ['all', 'error', None] :
            display(HTML(f"""<div class="check-bold check-fail">❌ The argument option can only be a 'all', 'error', None</div>"""))
            return
        self.attrs['display_msg'] = option

    @property
    def default_filter(self) -> pd.Series :
        """
        DataCheck 클래스의 기본 필터 조건을 반환
        """
        return self.attrs['default_filter']

    @property
    def count_fnc(self) -> Callable[[pd.Series], int]:
        """
        DataCheck 클래스의 count_fnc 속성을 반환
        """
        return self._count_fnc

    @count_fnc.setter
    def count_fnc(self, fnc: Callable[[pd.Series], int]) -> None:
        """
        DataCheck 클래스의 count_fnc 속성을 설정
        """
        if not callable(fnc):
            raise ValueError("The value must be a callable.")
        self._count_fnc = fnc

    def result_alt(self, qid: Union[str, List], alt: Optional[str]=None) -> str :
        """
        qid와 alt 값을 사용하여 결과 대체 텍스트를 생성하는 정적 메서드
        """
        alt_qid = qid
        if isinstance(qid, list) :
            alt_qid = f'{qid[0]}-{qid[-1]}'
        
        result_alt = alt_qid

        if self.attrs.get('title') is not None :
            match_qid = qid
            if isinstance(qid, list) :
                match_qid = qid[0]
            
            title_dict = self.attrs.get('title')
            if match_qid in title_dict.keys() :
                vgroup = title_dict[match_qid]['vgroup']
                if vgroup in title_dict :
                    title = title_dict[vgroup]['title']
                    qtype = title_dict[vgroup]['type']
                else :
                    title = title_dict[match_qid]['title']
                    qtype = title_dict[match_qid]['type']

                if not qtype in ['multiple'] :
                    sub_title = title_dict[match_qid]['sub_title']

                    if sub_title is not None :
                        result_alt = f'{alt_qid}: {title}_{sub_title}'
                    else :
                        result_alt = f'{alt_qid}: {title}'
                else :
                    result_alt = f'{alt_qid}: {title}'

        if alt is not None :
            result_alt = f'{alt_qid}: {alt}'

        return result_alt

    def result_html_update(self, **kwargs) :
        """
        결과 HTML을 업데이트하는 메서드로, 제공된 키워드 인수를 사용하여 기존 HTML 결과를 업데이트하거나 새로 추가
        """
        result_html = self.attrs['result_html'].copy()
        key = 'alt'
        updated = False
        if key in kwargs :
            chk_alt = {idx: result[key].strip().replace(' ', '') for idx, result in enumerate(result_html) if key in result and isinstance(result[key], str)}
            curr = kwargs[key].strip().replace(' ', '')
            for idx, value in chk_alt.items() :
                if curr == value :
                    result_html[idx] = kwargs
                    updated = True
        if not updated :
            result_html.append(kwargs)
        self.attrs['result_html'] = result_html

    def comp(self) :
        """
        전체 데이터 기준 `Series`를 Return
        """
        return pd.Series([True] * len(self), index=self.index)

    def set_filter(self, filter_cond: Optional[pd.Series] = None) -> None:
        """
        데이터 체크 기본 필터를 변경
        `filter_cond` (pd.Series or None) : `None`이면 전체 샘플 기준으로 변경
        """
        if filter_cond is None :
            self.attrs['default_filter'] = self.comp()
            if self.attrs['display_msg'] == 'all' :
                display(HTML(f"""🛠️ <span class="check-bold">Data Filter <span class="check-warn">Reset</span> : {len(self)}'s</span>"""))
        else :
            self.attrs['default_filter'] = filter_cond
            filt_data = self[filter_cond]
            if self.attrs['display_msg'] == 'all' :
                display(HTML(f"""🛠️ <span class="check-bold">Data Filter <span class="check-warn">Setting</span> : {len(filt_data)}'s</span>"""))

    def col_name_check(self, *variables: str) -> bool:
        """`qid`에 지정된 열이 데이터프레임에 있는지 확인"""
        chk_qid = [qid for qid in variables if not qid in list(self.columns)]
        if chk_qid :
            display(HTML(f"""<div class="check-bold check-fail">❌ The variable {chk_qid} is not in the data frame</div>"""))
            return False
        return True

    @contextlib.contextmanager
    def preserve_display_msg(self):
        """
        display_msg 속성을 임시로 변경하고, 코드 블록이 끝나면 원래 값으로 복원하는 컨텍스트 관리자
        """
        original_display_msg = self.attrs['display_msg']
        try:
            yield
        finally:
            self.attrs['display_msg'] = original_display_msg

    def display_description(self, alt: str, title: str, desc: pd.DataFrame=None) -> None :
        print_result = f"""
<div class="datacheck-apply alt-sub">
    <div class="datacheck-alt">{alt}</div>
    <div class="apply-title">✅ {title}</div>
</div>"""

        if desc is not None :
            desc_table = """
<table>
    <tr><td><b>Mean</b></td><td>{mean}</td></tr>
    <tr><td><b>Min</b></td><td>{minv}</td></tr>
    <tr><td><b>Max</b></td><td>{maxv}</td></tr>
</table>""".format(mean=desc.loc['mean'], minv=desc.loc['min'], maxv=desc.loc['max'])
            
            print_result = f"""
<div class="datacheck-apply alt-sub">
    <div class="datacheck-alt">{alt}</div>
    <div class="apply-title">📊 {title}</div>
    <div class="print-padding-left">{desc_table}</div>
</div>
    """
        if self.attrs['display_msg'] ==  'all' :
            display(HTML(print_result))

    def count_col(self, cnt_col_name: str, cols: Union[List[str], Tuple[str], str], value: Optional[Union[int, List[int], range]] = None) -> None:
        """
        주어진 열의 응답을 세어 새로운 열을 추가하는 메서드  
        (`nan` / `0` 이 아닌 컬럼 카운트)  
        결과를 요약하여 출력  
        """
        
        cnt_col = []
        alt = ""
        if isinstance(cols, str) :
            cnt_col = [cols]
            alt = f"{cols} Answer Count"
        elif isinstance(cols, tuple) or isinstance(cols, list):
            if not self.col_name_check(*cols) : return
            cnt_col = self.ma_return(cols)
            if len(cnt_col) == 1 :
                alt = f"{cols[0]} Answer Count"
            else :
                alt = f"{cnt_col[0]}-{cnt_col[-1]} Answer Count"
        
        if value is None:
            new_col = self[cnt_col].apply(self._count_fnc, axis=1).rename(cnt_col_name)
        elif isinstance(value, int):
            new_col = self[cnt_col].apply(lambda row: row.isin([value]).sum(), axis=1).rename(cnt_col_name)
        elif isinstance(value, list):
            new_col = self[cnt_col].apply(lambda row: row.isin(value).sum(), axis=1).rename(cnt_col_name)
        elif isinstance(value, range):
            value = list(value) + [value[-1] + 1]
            new_col = self[cnt_col].apply(lambda row: row.isin(value).sum(), axis=1).rename(cnt_col_name)

        with self.preserve_display_msg():
            result = self.assign(**{cnt_col_name: new_col})
            self.__dict__.update(result.__dict__)
        
        
        show_title = ""
        if value is None:
            show_title += f"""<span class="check-bold check-warn">{cnt_col_name}</span> : Answer count"""
        else:
            show_title += f"""<span class="check-bold check-warn">{cnt_col_name}</span> : Value count ({value})"""

        desc = self[cnt_col_name].describe().round(1)
        self.display_description(alt, show_title, desc)
        

    def sum_col(self, sum_col_name: str, cols: Union[List[str], Tuple[str], str]) -> None:
        """
        주어진 열의 값을 합산하여 새로운 열을 추가하는 메서드  
        결과를 요약하여 출력
        """
        if not self.col_name_check(*cols):
            return

        sum_col = []
        alt =""
        if isinstance(cols, (tuple, list)):
            sum_col = self.ma_return(cols)
            alt = f"{sum_col[0]}-{sum_col[-1]} : Sum"

        elif isinstance(cols, str):
            sum_col = [cols]
            alt = f"{cols} : Sum"

        new_col = self[sum_col].sum(axis=1).rename(sum_col_name)

        with self.preserve_display_msg():
            result = self.assign(**{sum_col_name: new_col})
            self.__dict__.update(result.__dict__)

        show_title = f"""<span class="check-bold check-warn">{sum_col_name}</span> : Sum of values"""

        desc = self[sum_col_name].describe().round(1)
        self.display_description(alt, show_title, desc)

    def ma_to_list(self, list_col_name: str, cols: Union[List[str], Tuple[str], str]) -> None:
        if not self.col_name_check(*cols):
            return

        ma_col = []
        if isinstance(cols, (tuple, list)):
            ma_col = self.ma_return(cols)
        elif isinstance(cols, str):
            ma_col = [cols]

        alt = f"{cols[0]}-{cols[-1]} : MA List" if len(cols) > 1 else f"{cols[0]} : MA List"
        new_col = self[ma_col].apply(lambda_ma_to_list, axis=1, qids=ma_col).rename(list_col_name)
        result = self.assign(**{list_col_name: new_col})
        self.__dict__.update(result.__dict__)

        show_title = f"""<span class="check-bold check-warn">{list_col_name}</span> : MA Variable to List"""

        self.display_description(alt, show_title)


    def _update_self(self, new_data):
        """
        self의 내부 데이터를 `new_data`로 업데이트
        """
        self.__dict__.update(new_data.__dict__)
     

    def ma_check(self, 
                ma: Union[List[str], Tuple[str]],
                len_chk: bool = True) -> bool:
        """
        `ma`가 리스트나 튜플인지, 그리고 다른 조건들을 만족하는지 확인
        """
        fail = """<div class="check-bold check-fail">❌ [ERROR] {warn_text}</div>"""
        example_text = """<div class="example-text">Example) {ex_text}</div>"""
        print_str = ""
        if not ma:
            print_str += fail.format(warn_text="Please check variable names")
            print_str += example_text.format(ex_text="['Q1r1', 'Q1r2', 'Q1r3'] / ('Q1r1', 'Q1r3')")
            display(HTML(print_str))
            return True

        if not isinstance(ma, (list, tuple)):
            print_str += fail.format(warn_text="Type of variable must be list or tuple")
            display(HTML(print_str))
            return True

        if len_chk and len(ma) < 2:
            print_str += fail.format(warn_text="Variable must be 2 length or more")
            display(HTML(print_str))
            return True

        if isinstance(ma, tuple) and len(ma) != 2:
            print_str += fail.format(warn_text="The variable must include 2 arguments")
            display(HTML(print_str))
            return True

        if isinstance(ma, tuple):
            cols = list(self.columns)
            first_index = cols.index(ma[0])
            last_index = cols.index(ma[1])
            if first_index > last_index:
                print_str += fail.format(warn_text=f"Please check the column index / current index ({first_index}-{last_index})")
                display(HTML(print_str))
                return True
        return False

    def ma_return(self,
                  ma: Union[List[str], Tuple[str]]) -> List[str]:
        """
        `ma`에 지정된 열을 반환
        """
        if isinstance(ma, tuple):
            cols = list(self.columns)
            first_index = cols.index(ma[0])
            last_index = cols.index(ma[1]) + 1
            return cols[first_index:last_index]
        elif isinstance(ma, list):
            return ma

    def show_message(self, 
                     export_df: ErrorDataFrame) -> None :
        """
        ErrorDataFrame 객체의 메시지를 표시하는 메서드  
        display_msg 속성에 따라 메시지를 출력
        """
        css = self.attrs['css']
        msg = export_df.chk_msg
        display_msg = """%s<br/>%s"""%(css, msg)
        if self.attrs['display_msg'] ==  'all' :
            display(HTML(display_msg))
        elif self.attrs['display_msg'] ==  'error' :
            if len(export_df.err()) > 1 :
                display(HTML(display_msg))
        elif self.attrs['display_msg'] is None :
            return


    def safreq(self, 
           qid: Optional[str] = None, 
           cond: Optional[pd.Series] = None, 
           only: Optional[Union[range, List[Union[int, float, str]], int, float, str]] = None,
           isnot: Optional[Union[range, List[Union[int, float, str]], int, float, str]] = None,
           alt: Optional[str]=None) -> 'ErrorDataFrame':
        """
        단수 응답(단일 변수) 데이터 체크 메서드
        """
        
        if not self.col_name_check(qid) : return

        show_cols = [qid]
        
        err_list = []

        # Answer Base Check
        warnings = []
        cond = (self.attrs['default_filter']) if cond is None else (self.attrs['default_filter']) & (cond)
        chk_df = self[cond].copy()

        def no_base_check() :
            if cond is not None :
                ans_err = 'DC_NO_BASE'
                add_df = self[(self.attrs['default_filter']) & ~(cond)].copy()
                add_df = add_df[~add_df[qid].isna()].copy()
                if len(add_df) > 0 :
                    add_df[ans_err] = 1
                    err_list = [ans_err]
                    chk_df = add_df
                    return [chk_df, err_list]
                else :
                    return None
            else :
                return None

        if len(chk_df) == 0 :
            warnings.append("No response to this condition")
            no_base = no_base_check()
            if no_base is not None :
                chk_df, err_list = no_base
        else :
            ms_err = 'DC_BASE'
            filt = (chk_df[qid].isna())  # Default
            chk_df.loc[filt, ms_err] = 1

            err_list.append(ms_err)

            # ONLY ANSWER CHECK
            if only is not None:
                warnings.append(f"Only value : {only}")
                if isinstance(only, range):
                    only = list(only) + [only[-1] + 1]
                elif isinstance(only, (int, float, str)):
                    only = [only]

                only_cond = (~chk_df[qid].isin(only))
                
                only_err = 'ONLY_ANS'
                chk_df.loc[only_cond, only_err] = 1
                err_list.append(only_err)
            
            # DONT ANSWER CHECK
            if isnot is not None:
                warnings.append(f"Disable value : {isnot}")
                if isinstance(isnot, range):
                    isnot = list(isnot) + [isnot[-1] + 1]
                elif isinstance(isnot, (int, float, str)):
                    isnot = [isnot]
                
                isnot_cond = (chk_df[qid].isin(isnot))
                
                isnot_err = 'ISNOT_ANS'
                chk_df.loc[isnot_cond, isnot_err] = 1
                err_list.append(isnot_err)

            # Cases responded to other than base
            no_base = no_base_check()
            if no_base is not None :
                chk_df, err_list = no_base
        
        set_alt = self.result_alt(qid, alt)
        edf = ErrorDataFrame(qid, 'SA', show_cols, chk_df, err_list, warnings, set_alt)
        self.show_message(edf)
        self.result_html_update(alt=set_alt, result_html=edf.chk_msg, dataframe=edf.err()[show_cols+edf.extra_cols].to_json())
        return edf

    def key_var_setting(self, cols: List, key_var: Optional[str]) -> str :
        def set_code() :
            get_qid = get_key_id(cols)
            if not isinstance(get_qid, str) :
                return get_qid
            return "%s{code}"%(get_qid)

        if key_var is None :
            return set_code()
        else :
            if not "{code}" in key_var :
                display(HTML("""<div class="datacheck-head check-fail">The `key_var` does not contain {code}.</div>"""))
                return set_code()
            return key_var

    def mafreq(self, 
            qid: Union[List[str], Tuple[str, ...]], 
            cond: Optional[pd.Series] = None, 
            atleast: Optional[int] = None, 
            atmost: Optional[int] = None, 
            exactly: Optional[int] = None,
            isin: Optional[Union[range, List[Union[int, str]], int, str]] = None,
            isall: Optional[Union[range, List[Union[int, str]], int, str]] = None,
            isnot: Optional[Union[range, List[Union[int, str]], int, str]] = None,
            no_base: bool = True,
            alt: Optional[str]=None,
            key_var: Optional[str]=None) -> 'ErrorDataFrame':
        """
        복수 응답(다중 변수) 데이터 체크 메서드
        """
        if (self.ma_check(qid)) :
            return
        
        err_list = []

        # Answer Base Check
        warnings = []
        show_cols = self.ma_return(qid)
        
        if not self.col_name_check(*show_cols) : return

        qid_key = self.key_var_setting(cols=show_cols, key_var=key_var)

        cond = (self.attrs['default_filter']) if cond is None else (self.attrs['default_filter']) & (cond)
        chk_df = self[cond].copy()

        cnt = 'ANSWER_CNT'

        def no_base_check() :
            if cond is not None and no_base :
                ans_err = 'DC_NO_BASE'
                add_df = self[self.attrs['default_filter'] & ~(cond)].copy()
                add_df[cnt] = add_df[show_cols].apply(lambda x: x.count() - (x==0).sum(), axis=1)
                add_filt = (add_df[show_cols].isna() | (add_df[show_cols] == 0)).all(axis=1)
                add_df = add_df[~add_filt].copy()
                if len(add_df) > 0 :
                    add_df[ans_err] = 1
                    err_list = [ans_err]
                    chk_df = add_df
                    return [chk_df, err_list]
                else :
                    return None
            else :
                return None

        if len(chk_df) == 0 :
            warnings.append("No response to this condition")
            no_base = no_base_check()
            if no_base is not None :
                chk_df, err_list = no_base
        else :            
            chk_df[cnt] = chk_df[show_cols].apply(lambda x: x.count() - (x==0).sum(), axis=1)

            ms_err = 'DC_BASE'
            # filt = (chk_df[cnt]==0)  # Default
            filt = (chk_df[show_cols].isna() | (chk_df[show_cols] == 0)).all(axis=1)
            chk_df.loc[filt, ms_err] = 1

            err_list.append(ms_err)

            # Generalized Answer Check Function
            def check_answer(condition, operator, err_label):
                if condition is not None:
                    if operator == '==':
                        cond_err = (chk_df[cnt] != condition)
                        warnings.append(f"Exactly : {condition}")
                    elif operator == '<':
                        cond_err = (chk_df[cnt] < condition)
                        warnings.append(f"Atleast : {condition}")
                    elif operator == '>':
                        cond_err = (chk_df[cnt] > condition)
                        warnings.append(f"Atmost : {condition}")
                    
                    chk_df.loc[cond_err, err_label] = 1
                    err_list.append(err_label)

            # AT LEAST, AT MOST, EXACTLY Answer Checks
            check_answer(atleast, '<', 'DC_ATLEAST')
            check_answer(atmost, '>', 'DC_ATMOST')
            check_answer(exactly, '==', 'DC_EXACTLY')

            def process_check(check_type, check_value, check_func, err_label):
                warnings.append(f"{check_type.capitalize()} value : {check_value}")
                if isinstance(check_value, range):
                    check_list = list(check_value) + [check_value[-1] + 1]
                elif isinstance(check_value, (int, str)):
                    check_list = [check_value]
                elif isinstance(check_value, list):
                    check_list = check_value

                chk_cols = [qid_key.format(code=m) for m in check_list]

                def apply_func(row):
                    return 1 if check_func(row, chk_cols) else np.nan

                chk_df[err_label] = chk_df.apply(apply_func, axis=1)

                err_list.append(err_label)

            # Check Functions
            def ma_isin_check(row, cols):
                return not any(not (pd.isna(row[c]) or row[c] == 0) for c in cols)

            def ma_isall_check(row, cols):
                return any(pd.isna(row[c]) or row[c] == 0 for c in cols)

            def ma_isnot_check(row, cols) :
                return any(not (pd.isna(row[c]) or row[c] == 0) for c in cols)

            # Is In Check
            if not isinstance(qid_key, str) :
                warnings.append("A variable structure for which the isin/isall/isnot methods are not available")
            else :
                if isin is not None:
                    process_check('isin', isin, ma_isin_check, 'MA_ISIN')

                # Is All Check
                if isall is not None:
                    process_check('isall', isall, ma_isall_check, 'MA_ISALL')

                # Is Not Check
                if isnot is not None:
                    process_check('isnot', isnot, ma_isnot_check, 'MA_ISNOT')

            # Cases responded to other than base
            if not no_base : 
                warnings.append('No Base Check does not run')
            
            if cond is not None and no_base :
                no_base = no_base_check()
                if no_base is not None :
                    chk_df, err_list = no_base


            show_cols = [cnt] + show_cols
        
        set_alt = self.result_alt(qid, alt)
        edf = ErrorDataFrame(qid, 'MA', show_cols, chk_df, err_list, warnings, set_alt)
        self.show_message(edf)
        self.result_html_update(alt=set_alt, result_html=edf.chk_msg, dataframe=edf.err()[show_cols+edf.extra_cols].to_json())
        return edf


    def logchk(self, 
               base: Optional[pd.Series] = None, 
               ans: pd.Series = None,
               alt: Optional[str]=None) -> 'ErrorDataFrame':
        """
        특정 로직에 대한 응답 체크
        (`base`가 `True`일 때, `ans`도 `True`)

        `base` (pd.Series): 베이스 조건.
        `ans` (pd.Series): 베이스 조건이 True일 때 응답 조건.
        """

        if ans is None :
            display(HTML("""<div class="check-bold check-fail">❌ [ERROR]  answer_cond cannot be None</div>"""))
            return 
        err_list = []

        # Base Condition Answer Check
        warnings = []
        base_cond = self.comp() if base is None else base
        base_cond = (self.attrs['default_filter']) & (base_cond)
        ans_cond  = (self.attrs['default_filter']) & (ans)
        chk_df = self[base_cond].copy()

        if len(chk_df) == 0:
            warnings.append("No response to this condition")
        else :
            # Base Filter
            base_col = 'BASE_COND'
            answer_col = 'ANSWER_COND'
            err_list += [base_col, answer_col]
            chk_df.loc[base_cond, base_col] = 1
            chk_df.loc[ans_cond, answer_col] = 1

            # Logic Check
            lg_err = 'DC_LOGIC'
            chk_df.loc[(base_cond) & (~ans_cond), lg_err] = 1
            err_list.append(lg_err)

            chk_df = chk_df[base_cond.reindex(chk_df.index, fill_value=False)]
            
        edf = ErrorDataFrame('LOGIC CHECK', 'LOGIC', [], chk_df, err_list, warnings, alt)
        self.show_message(edf)
        
        if alt is not None :
            self.result_html_update(alt=alt, result_html=edf.chk_msg, dataframe=edf.err()[edf.extra_cols].to_json())
        return edf

    def dupchk(self, 
           qid: Union[List[str], Tuple[str, ...]], 
           cond: Optional[pd.Series] = None, 
           okUnique: Optional[Union[List[Any], range, int, str]] = None,
           alt: Optional[str]=None) -> 'ErrorDataFrame' :
        """
        중복 응답 데이터 체크 메서드 (순위 응답)        
        `qid` (Union[List[str], Tuple[str]]): 중복을 체크할 열들.
        `okUnique` (Union[List, range, int, str], optional): 무시할 특정 값(들). 기본값은 None.
        """
        if (self.ma_check(qid)) :
            return
        
        show_cols = self.ma_return(qid)
        if not self.col_name_check(*show_cols): return

        warnings = []
        err_list = []

        cond = (self.attrs['default_filter']) if cond is None else (self.attrs['default_filter']) & (cond)
        chk_df = self[cond].copy()
        chk_df = chk_df[(~chk_df[show_cols].isna()).any(axis=1)]

        dup_err = 'DC_DUP'
        err_list.append(dup_err)

        if len(chk_df) == 0 :
            warnings.append("No response to this condition")

        if okUnique is not None:
            if not isinstance(okUnique, (list, range, int, str)):
                display(HTML("""<div class="check-bold check-fail">❌ [ERROR] Type of okUnique must be list, range, int, or str</div>"""))
                return
            if isinstance(okUnique, range):
                okUnique = list(okUnique)
                okUnique.append(okUnique[-1] + 1)
            elif isinstance(okUnique, (int, str)):
                okUnique = [okUnique]
            
            warnings.append(f"""Allow Duplicates : {okUnique}""")
        else:
            okUnique = []

        def check_duplicates(row):
            row_values = row.tolist()
            filtered_values = [value for value in row_values if value not in okUnique and not pd.isna(value)]
            return 1 if len(filtered_values) != len(set(filtered_values)) else None
        
        chk_df[dup_err] = chk_df[show_cols].apply(check_duplicates, axis=1)

        rk = show_cols
        alt = f"""{rk[0]}-{rk[-1]} (DUP)"""
        edf = ErrorDataFrame(qid, 'DUP', show_cols, chk_df, err_list, warnings, alt)
        self.show_message(edf)
        self.result_html_update(alt=self.result_alt(qid, alt), result_html=edf.chk_msg, dataframe=edf.err()[show_cols+edf.extra_cols].to_json())
        return edf

    def display_key_var_error(self, arg_name:str, qid_list: List) -> None :
            print_text = f"""<div class="check-bold check-fail">❌ [ERROR] Please check multi question variable names : `{arg_name}`</div>"""
            for key, var_list in qid_list.items() :
                print_text += f"""<div class="check-bold check-fail">[{key}] : {var_list}</div>"""
            display(HTML(print_text))

    def masa(self, 
             ma_qid: Union[List[str], Tuple[str, ...]], 
             sa_qid: str, 
             cond: Optional[pd.Series] = None, 
             diff_value: Optional[Union[List[Any], range, int, str]] = None,
             alt: Optional[str]=None,
             key_var: Optional[str]=None) -> 'ErrorDataFrame' :
        """
        `복수 응답`을 베이스로 하는 `단수 응답` 로직 체크.
        `ma_qid` (Union[List[str], Tuple[str]]): 복수 응답 열 목록.
        `sa_qid` (str): 단수 응답 열.
        `cond` (pd.Series, optional): 조건을 적용할 시리즈. 기본값은 None.
        `diff_value` (Union[List, int], optional): 무시할 특정 값(들). 기본값은 None.
        """
        if (self.ma_check(ma_qid)) :
            return
        warnings = []
        err_list = []
         
        base_qid = self.ma_return(ma_qid)
        if not self.col_name_check(*base_qid): return
        if not self.col_name_check(sa_qid): return

        qid_key = self.key_var_setting(cols=base_qid, key_var=key_var)
        if not isinstance(qid_key, str): 
            self.display_key_var_error('ma_qid', qid_key)
            return

        cond = (self.attrs['default_filter']) if cond is None else (self.attrs['default_filter']) & (cond)
        chk_df = self[cond].copy()

        ma = base_qid
        sa = sa_qid

        show_cols = [sa] + ma

        if len(chk_df) == 0 :
            warnings.append("No response to this condition")
        else :
            filt = ~chk_df[sa].isna()

            err_col = 'DC_LOGIC'
            # MA Base SA
            if len(chk_df[filt]) == 0 :
                warnings.append("No response to this condition")

            dv = []
            if diff_value is not None:
                if not isinstance(diff_value, (list, range, int, str)):
                    display(HTML("""<div class="check-bold check-fail">❌ [ERROR] Type of diff_value must be list, range, int, or str</div>"""))
                    return
                if isinstance(diff_value, (int, str)) :
                    dv = [diff_value]
                if isinstance(diff_value, list) :
                    dv = diff_value
                if isinstance(diff_value, range):
                    dv = list(diff_value)
                    dv.append(dv[-1] + 1)
                warnings.append(f"""Do not check the code : {dv}""")
            
            def ma_base_check(x) :
                sa_ans = int(x[sa])
                if sa_ans in dv :
                    return np.nan
                
                ma_var = qid_key.format(code=sa_ans)
                ma_ans = x[ma_var]


                return 1 if pd.isna(ma_ans) or ma_ans == 0 else np.nan

            chk_df[err_col] = chk_df[filt].apply(ma_base_check, axis=1)
            err_list.append(err_col)

            ma_ans = 'BASE_MA'
            chk_df[ma_ans] = chk_df[filt].apply(lambda_ma_to_list, axis=1, qids=ma)

            show_cols = [ma_ans] + show_cols

            chk_df = chk_df if cond is None else chk_df[cond.reindex(chk_df.index, fill_value=False)]
        
        qid = f"""{sa}(SA) in {ma[0]}-{ma[-1]}(MA)"""
        edf = ErrorDataFrame(qid, 'MASA', show_cols, chk_df, err_list, warnings, alt)
        self.show_message(edf)
        self.result_html_update(alt=self.result_alt(qid, alt), result_html=edf.chk_msg, dataframe=edf.err()[show_cols+edf.extra_cols].to_json())
        return edf

    def mama(self,
             base_ma: Union[List[str], Tuple[str, ...]], 
             chk_ma: Union[List[str], Tuple[str, ...]], 
             cond: Optional[pd.Series] = None, 
             diff_value: Optional[Union[List[Any], range, int, str]] = None,
             alt: Optional[str]=None,
             base_key_var: Optional[str]=None,
             chk_key_var: Optional[str]=None,) -> 'ErrorDataFrame' :
        """
        `복수 응답`을 베이스로 하는 `복수 응답` 로직 체크.
        `base_ma` (Union[List[str], Tuple[str]]): 기준이 되는 복수 응답 열 목록.
        `chk_ma` (Union[List[str], Tuple[str]]): 체크할 복수 응답 열 목록.
        `cond` (pd.Series, optional): 조건을 적용할 시리즈. 기본값은 None.
        `diff_value` (Union[List, int], optional): 무시할 특정 값(들). 기본값은 None.
        """
        if (self.ma_check(base_ma)) or (self.ma_check(chk_ma)) :
            return
        warnings = []
        err_list = []
         
        base = self.ma_return(base_ma)
        chkm = self.ma_return(chk_ma)
        if not self.col_name_check(*base): return
        if not self.col_name_check(*chkm): return

        qid_key = self.key_var_setting(cols=base, key_var=base_key_var)
        ans_key = self.key_var_setting(cols=chkm, key_var=chk_key_var)

        if any(x is None for x in [qid_key, ans_key]) :
            if not isinstance(qid_key, str) : 
                self.display_key_var_error('base_ma', qid_key)

            if ans_key is None: 
                self.display_key_var_error('chk_ma', ans_key)

            return


        zip_cols = [list(x) for x in zip(base, chkm)]
        show_cols = sum(zip_cols, [])

        cond = (self.attrs['default_filter']) if cond is None else (self.attrs['default_filter']) & (cond)
        chk_df = self[cond].copy()

        if len(chk_df) == 0 :
            warnings.append("No response to this condition")
        else :
            chk_cnt = 'CHK_CNT'
            chk_df[chk_cnt] = chk_df[chkm].apply(lambda x: x.count() - (x==0).sum(), axis=1)
            filt = chk_df[chk_cnt]>=1
            
            err_col = 'DC_LOGIC'
            # MA Base MA
            if len(chk_df[filt]) == 0 :
                warnings.append("No response to this condition")

            dv = []
            diff_qids = []
            diff_ans  = []
            if diff_value is not None:
                if not isinstance(diff_value, (list, range, int, str)):
                    display(HTML("""<div class="check-bold check-fail">❌ [ERROR] Type of diff_value must be list, range, int, or str</div>"""))
                    return
                if isinstance(diff_value, (int, str)) :
                    dv = [diff_value]
                if isinstance(diff_value, list) :
                    dv = diff_value
                if isinstance(diff_value, range):
                    dv = list(diff_value)
                    dv.append(dv[-1] + 1)
                
                warnings.append(f"""Do not check the code : {dv}""")
                diff_qids = [qid_key.format(code=x) for x in dv]
                diff_ans  = [ans_key.format(code=x) for x in dv]

            def ma_base_check(x) :
                def flag(b, a) :

                    if pd.isna(b) or b == 0 :
                        if not (pd.isna(a) or a == 0) :
                            return True
                    
                    return False
                return 1 if any(flag(x[base], x[ans]) for base, ans in zip_cols if (not base in diff_qids) and (not ans in diff_ans)) else np.nan
                
            chk_df[err_col] = chk_df[filt].apply(ma_base_check, axis=1)


            def diff_ans_update(row, cols) :
                def return_int_or_str(txt: str) :
                        rp = txt.replace(qid_key.format(code=''), '')
                        if rp.isdigit() :
                            return int(rp)
                        else :
                            return rp
                return [return_int_or_str(base) for base, ans in cols if (pd.isna(row[base]) or row[base] == 0) and not (pd.isna(row[ans]) or row[ans] == 0)]

            base_ans = 'BASE_MA'
            chk_ans = 'CHECK_MA'
            diff_ans = 'DIFF_ANS'
            chk_df[base_ans] = chk_df[filt].apply(lambda_ma_to_list, axis=1, qids=base)
            chk_df[chk_ans] = chk_df[filt].apply(lambda_ma_to_list, axis=1, qids=chkm)
            chk_df[diff_ans] = chk_df[filt].apply(diff_ans_update, axis=1, cols=zip_cols)
            
            err_list.append(err_col)
            show_cols = [base_ans, chk_ans, diff_ans] + show_cols
            chk_df = chk_df if cond is None else chk_df[cond.reindex(chk_df.index, fill_value=False)]
        
        qid = f"""{chkm[0]}-{chkm[-1]}(MA) in {base[0]}-{base[-1]}(MA)"""
        edf = ErrorDataFrame(qid, 'MAMA', show_cols, chk_df, err_list, warnings, alt)
        self.show_message(edf)
        self.result_html_update(alt=self.result_alt(qid, alt), result_html=edf.chk_msg, dataframe=edf.err()[show_cols+edf.extra_cols].to_json())
        return edf

    def mark(self,
            base_qid: Union[List[str], Tuple[str, ...]], 
            rank_qid: Union[List[str], Tuple[str, ...]], 
            cond: Optional[pd.Series] = None, 
            diff_value: Optional[Union[List[Any], range, int, str]] = None,
            alt: Optional[str]=None,
            key_var: Optional[str]=None) -> 'ErrorDataFrame' :
        """
        `복수 응답`을 베이스로 하는 `순위 응답` 로직 체크.
        `base_qid` (Union[List[str], Tuple[str]]): 기준이 되는 복수 응답 열 목록.
        `cond` (pd.Series, optional): 조건을 적용할 시리즈. 기본값은 None.
        `rank_qid` (Union[List[str], Tuple[str]]): 체크할 순위 응답 열 목록.
        """
        if (self.ma_check(base_qid)) or (self.ma_check(rank_qid)) :
            return
        warnings = []
        err_list = []
         
        base = self.ma_return(base_qid)
        rank = self.ma_return(rank_qid)
        max_rank = len(rank)
        if not self.col_name_check(*base): return
        if not self.col_name_check(*rank): return

        qid_key = self.key_var_setting(cols=base, key_var=key_var)
        if not isinstance(qid_key, str) :
            self.display_key_var_error('base_qid', qid_key)
            return

        show_cols = rank

        cond = (self.attrs['default_filter']) if cond is None else (self.attrs['default_filter']) & (cond)
        chk_df = self[cond].copy()

        if len(chk_df) == 0 :
            warnings.append("No response to this condition")
        else :
            dv = []
            if diff_value is not None:
                if not isinstance(diff_value, (list, range, int, str)):
                    display(HTML("""<div class="check-bold check-fail">❌ [ERROR] Type of diff_value must be list, range, int, or str</div>"""))
                    return
                if isinstance(diff_value, (int, str)) :
                    dv = [diff_value]
                if isinstance(diff_value, list) :
                    dv = diff_value
                if isinstance(diff_value, range):
                    dv = list(diff_value)
                    dv.append(dv[-1] + 1)
                
                warnings.append(f"""Do not check the code : {dv}""")
                # base = [x for x in base if not x in [f'{qid_key}{d}' for d in dv]]

            base_cnt = 'BASE_COUNT'
            chk_df[base_cnt] = chk_df[base].apply(lambda x: x.count() - (x==0).sum(), axis=1)

            filt = chk_df[base_cnt]>=1

            err_col = 'DC_LOGIC'
            # MA Base MA
            if len(chk_df[filt]) == 0 :
                warnings.append("No response to this condition")

            def ma_base_rank_check(x) :
                able_ans = max_rank if x[base_cnt] > max_rank else x[base_cnt]
                chk_rank = rank[:able_ans]
                return 1 if any(pd.isna(x[rk]) for rk in chk_rank) else np.nan


            # ma_base_cond = (~chk_df[rank].isna()).any(axis=1)
            ma_base_cond = chk_df[base_cnt]>=1
            
            chk_df[err_col] = chk_df[ma_base_cond].apply(ma_base_rank_check, axis=1)

            base_ans = 'BASE_MA'
            chk_df[base_ans] = chk_df[base].apply(lambda_ma_to_list, axis=1, qids=base)


            def ma_base_check(x, rank_qid) :
                sa_ans = int(x[rank_qid])
                if sa_ans in dv :
                    return np.nan
                
                ma_var = qid_key.format(code=sa_ans)
                ma_ans = x[ma_var]


                return 1 if pd.isna(ma_ans) or ma_ans == 0 else np.nan
            # Each Rank masa
            rank_err_list = []
            for rk in rank :
                rk_err = f'{rk}_ERR'
                sa_base_cond = ~chk_df[rk].isna()
                chk_df[rk_err] = chk_df[sa_base_cond].apply(ma_base_check, axis=1, rank_qid=rk)
                rank_err_list.append(rk_err)

            masa_err = 'ERR_RK'
            def masa_rank_err(x) :
                if any(x[err]==1 for err in rank_err_list) :
                    return [cnt for cnt, rank in enumerate(rank_err_list, 1) if x[rank]==1]
                else :
                    return np.nan

            chk_df[masa_err] = chk_df[(chk_df[rank_err_list]==1).any(axis=1)].apply(masa_rank_err, axis=1)
            chk_df.loc[~chk_df[masa_err].isna(), err_col] = 1
            
            show_cols = [base_cnt, base_ans, masa_err] + rank + base
            err_list += [err_col]
            err_list += rank_err_list

            chk_df = chk_df if cond is None else chk_df[cond.reindex(chk_df.index, fill_value=False)]
        
        qid = f"""{rank[0]}-{rank[-1]}(RANK) in {base[0]}-{base[-1]}(MA)"""
        edf = ErrorDataFrame(qid, 'MARK', show_cols, chk_df, err_list, warnings, alt)
        self.show_message(edf)
        self.result_html_update(alt=self.result_alt(qid, alt), result_html=edf.chk_msg, dataframe=edf.err()[show_cols+edf.extra_cols].to_json())
        return edf


    def rate_rank(self,
                  rate_qid: Union[List[str], Tuple[str, ...]], 
                  rank_qid: Union[List[str], Tuple[str, ...]],
                  cond: Optional[pd.Series] = None,
                  alt: Optional[str]=None,
                  key_var: Optional[str]=None)  -> 'ErrorDataFrame' :
        """
        `척도 응답`을 베이스로 하는 `순위 응답` 로직 체크.
        ()`척도 응답`의 점수 기준으로 `순위 응답`이 순서대로 응답되어야 하는 경우)
        `rate_qid` (Union[List[str], Tuple[str]]): 기준이 되는 복수 응답 열 목록.
        `rank_qid` (Union[List[str], Tuple[str]]): 체크할 순위 응답 열 목록.
        `cond` (pd.Series, optional): 조건을 적용할 시리즈. 기본값은 None.
        """
        if (self.ma_check(rate_qid)) or (self.ma_check(rank_qid, len_chk=False)) :
            return
        warnings = []
        err_list = []
         
        rate = self.ma_return(rate_qid)
        rank = self.ma_return(rank_qid)
        if not self.col_name_check(*rate): return
        if not self.col_name_check(*rank): return

        qid_key = self.key_var_setting(cols=rate_qid, key_var=key_var)
        if not isinstance(qid_key, str) :
            self.display_key_var_error('rate_qid', qid_key)
            return

        cond = (self.attrs['default_filter']) if cond is None else (self.attrs['default_filter']) & (cond)
        chk_df = self[cond].copy()

        if len(chk_df) == 0 :
            warnings.append("No response to this condition")
        else :
            filt = (~chk_df[rank].isna()).any(axis=1)

            err_col = 'DC_LOGIC'
            def rate_rank_validate(row, rate_base, rank_base):
                scores = {int(x.replace(qid_key.format(code=''), '')): row[x] for x in rate_base if not pd.isna(row[x])}
                result = {}
                for key, value in scores.items():
                    if value not in result:
                        result[value] = []
                    result[value].append(key)
                
                sort_score = [[key, result[key]] for key in sorted(result.keys(), reverse=True)]
                
                rk = rank_base.copy()
                is_valid = False
                for sc, able in sort_score :
                    albe_rk = rk[:len(able)]
                    if not rk :
                        break
                    
                    for ar in albe_rk :
                        if not row[ar] in able :
                            is_valid = True
                        
                        rk.remove(ar)

                return 1 if is_valid else np.nan


            chk_df[err_col] = chk_df[filt].apply(rate_rank_validate, axis=1, rate_base=rate, rank_base=rank)


            def rate_rank_able_attrs(row, rate_base):
                scores = {int(x.replace(qid_key.format(code=''), '')): row[x] for x in rate_base if not pd.isna(row[x])}
                result = {}
                for key, value in scores.items():
                    if value not in result:
                        result[value] = []
                    result[value].append(key)
                
                sort_score = {int(key): result[key] for key in sorted(result.keys(), reverse=True)}
                
                return sort_score
            
            able_col = 'SCORE_ATTR'
            chk_df[able_col] = chk_df[filt].apply(rate_rank_able_attrs, axis=1, rate_base=rate)


            err_list.append(err_col)
            show_cols = [able_col] + rank + rate

            chk_df = chk_df if cond is None else chk_df[cond.reindex(chk_df.index, fill_value=False)]
        
        qid = f"""{rank[0]}-{rank[-1]}(RANK) / {rate[0]}-{rate[-1]}(RATE)"""
        edf = ErrorDataFrame(qid, 'RATERANK', show_cols, chk_df, err_list, warnings, alt)
        self.show_message(edf)
        self.result_html_update(alt=self.result_alt(qid, alt), result_html=edf.chk_msg, dataframe=edf.err()[show_cols+edf.extra_cols].to_json())
        return edf
    
    def note(self, print_word: str) -> None:
        """
        별도 표시를 위한 메서드
        """
        if self.attrs['display_msg'] ==  'all' :
            display(HTML(f"""
                         <div class="datacheck-print-mw">
                            <div class="datacheck-note-print">
                                <div class="note-title">📝 NOTE</div>
                                <div class="note-desc">{print_word}</div>
                            </div>
                         </div>
                         """))

    def live_only(self) -> None:
        """
        LIVE 상태에서 검토해야하는 부분 표기
        """
        if self.attrs['display_msg'] ==  'all' :
            display(HTML(f"""
                        <div class="datacheck-print-mw">
                            <div class="datacheck-live-check">LIVE CHECK</div>
                        </div>
                         """))

    def qset(self, qid: str, code: Union[range, List]) -> List :
        """
        `qid`와 `code`를 기준으로 `DataFrame`에서 변수명 추출
        `(startswith(qid) and endswith(each code))`
        `qid` (str) : 문자열로 된 기준 변수명 ('SQ1', 'SQ2')
        `code` (range, list) : 각 변수의 속성 코드 (`[1, 2, 3, 4]`)
        example) qid='SQ7' / code=[1, 3, 5]
        return `['SQ7r1', 'SQ7r3', 'SQ7r5']`
        """
        cols = self.columns
        if not isinstance(code, (range, list)) :
            display(HTML(f"""<div class="check-bold check-fail">❌ The argument code can only be a list or range</div>"""))
            return []

        if any(not isinstance(c, int) for c in code) :
            display(HTML(f"""<div class="check-bold check-fail">❌ The argument code can only be numeric</div>"""))
            return []
        
        chk_code = code
        if isinstance(code, range) :
            chk_code.append(chk_code[-1]+1)

        filt = [col for col in cols if col.startswith(qid) and any(str(c) in re.findall(r'\d+$', col.replace(qid, '')) for c in chk_code)]
        
        if not filt :
            display(HTML("""<div class="check-bold check-warn">⚠️ The variable does not exist in the dataframe</div>"""))
        return filt

    # DataProcessing
    def setting_meta(self, meta, variable, dup_chk=True, check_title=True) :
        if variable is None :
            return None

        if meta is False :
            return None

        return_meta = None
        if meta is None :
            meta = self.attrs.get('meta')
            titles = self.attrs.get('title')
            if meta :
                if isinstance(variable, str) :
                    if variable in meta.keys() :
                        return_meta = meta[variable]
                
                if isinstance(variable, list) :
                    qtypes = [titles[v]['type'] for v in variable if v in titles.keys()]
                    
                    return_meta = {v: meta[v] if v in meta.keys() else v for v in variable}
                    if qtypes :
                        qtype = list(set(qtypes))[0]
                        if qtype in ['rank', 'single', 'rating'] :
                            vgroup = list(set([titles[v]['vgroup'] for v in variable]))
                            if len(vgroup) > 1 :
                                print(f"⚠️ duplicate variable group detected : {vgroup} / Gets the meta of the first variable reference")
                            
                            return_meta = meta[vgroup[0]]
            else :
                return None
        else :
            return_meta = meta
        
        if return_meta is not None and dup_chk :
            return_meta = check_duplicate_meta(return_meta)

        return return_meta

    def setting_title(self, title, variable) :
        if variable is None :
            return None

        if title is False :
            return None

        return_title = None
        if title is None :
            title_attr = self.attrs.get('title')
            if title_attr :
                chk_var = variable
                if isinstance(chk_var, list) :
                    chk_var = variable[0]
            
                if chk_var in title_attr.keys() :
                    set_title = ''
                    qtype = None
                    if 'vgroup' in title_attr.keys() :
                        vgroup = title_attr[chk_var]['vgroup']
                        set_title = title_attr[vgroup]['title']
                        qtype = title_attr[vgroup]['type']
                    else :
                        set_title = title_attr[chk_var]['title']
                        qtype = title_attr[chk_var]['type']
                    
                    set_title = set_title.replace('(HIDDEN)', '').strip()

                    if not qtype in ['multiple'] :
                        sub_title = title_attr[chk_var]['sub_title']

                        if sub_title is not None :
                            set_title = f'{set_title}_{sub_title}'

                    return_title = set_title
            else :
                return None
        else :
            return_title = title

        return return_title

    def wordcloud(self, index: str,
                        columns: Optional[Union[str, List[str]]] = None,
                        cond: Optional[pd.Series] = None,
                        font_path='malgun.ttf', 
                        image_path='WordCloud',
                        width=800, 
                        height=500,
                        base_desc: Optional[str] = None) :

        if isinstance(index, list) :
            raise ValueError('index must be string')

        titles = self.attrs.get('title')
        metas = self.attrs.get('meta')

        qtype = 'text'
        if isinstance(index, str) :
            if index in titles.keys() :
                qtype = titles[index]['type']


        if not qtype in ['text'] :
            raise ValueError('index type is not text')


        cond = (self.attrs['default_filter']) if cond is None else (self.attrs['default_filter']) & (cond)
        df = self[cond].copy()

        filt_variables = []

        if isinstance(columns, tuple) :
            columns = self.ma_return(columns)
            if not self.col_name_check(*columns) : return

        # Table Header
        varable_text = []
        index_cond = None

        varable_text.append(index)
        index_cond = ~df[index].isna()

        if isinstance(columns, list) :
            varable_text.append(f'{columns[0]}-{columns[-1]}')
        else :
            if columns is not None :
                varable_text.append(columns)

        # Index
        filt_variables.append(index)

        # Columns
        if isinstance(columns, list) :
            filt_variables += columns
        else :
            if columns is not None :
                filt_variables.append(columns)

        filt_variables = list(set(filt_variables))

        
        df = df[index_cond][filt_variables].copy()
        
        sample_count = len(self.index.to_list())
        all_count = len(df)
        
        if base_desc is None :
            if sample_count == all_count :
                base_desc = 'All Base'
            else :
                sample_ratio = round(all_count/sample_count, 2) * 100
                base_desc = f'Not All Base ({sample_ratio:.0f}%)'
        
        #### Save Folder ####
        if not os.path.exists(image_path) :
            os.makedirs(image_path)

        ####  Opened Ended Data Only ####
        result = None

        question_title = index
        if index in titles.keys() :
            question_title = titles[index]['title']

        if columns is None :
                result = wordcloud_table(df, 
                                         index, 
                                         font_path=font_path,
                                         image_path=image_path,
                                         width=width,
                                         height=height)

        else :
            # by_columns = f'{index}'
            wc_path = os.path.join(image_path, index)
            if not os.path.exists(wc_path) :
                os.makedirs(wc_path)

            meta_wc = []
            only_index_result = wordcloud_table(df, 
                                                index, 
                                                font_path=font_path,
                                                image_path=wc_path,
                                                width=width,
                                                height=height)

            meta_wc.append(('Total', only_index_result))

            result = wordcloud_table(df, 
                                     index, 
                                     columns,
                                     font_path=font_path,
                                     image_path=wc_path,
                                     width=width,
                                     height=height)

            if result :
                for r in result :
                    base, wc = r
                    col_title = base
                    if isinstance(columns, list) :
                        if base in metas.keys() :
                            col_title = metas[base]

                    if isinstance(columns, str) :
                        if columns in metas.keys() :
                            col_meta = [v for m, v in metas[columns].items() if m == str(base)]
                            if col_meta :
                                col_title = col_meta[0]
                    
                    meta_wc.append((col_title, wc))
                
                result = meta_wc
        
        
        return WordCloudHandler(question_title, result, index, base_desc, sample_size=all_count)
        #### ======================= ####

    def banner_wordcloud(self, 
                         index: str,
                         columns: Optional[Union[str, List[str], dict]] = None,
                         cond: Optional[pd.Series] = None,
                         base_desc: Optional[str] = None,
                         **options) :

            if columns is None :
                banner = self.attrs['banner']
                if banner is not None :
                    columns = banner

            cond = (self.attrs['default_filter']) if cond is None else (self.attrs['default_filter']) & (cond)
            df = self[cond].copy()

            titles = self.attrs.get('title')
            clouds = []

            total_flag = True

            if columns is not None :
                if isinstance(columns, dict) :
                    columns = list(columns.items())
                
                if isinstance(columns, str) :
                    columns = [(titles.get(columns, {}).get('title', columns), columns)]

                for col_head, col in columns :
                    header = col_head
                    if col_head in titles.keys() :
                        header = titles[col_head]['title']
                    
                    cloud_list = self.wordcloud(index, col, cond=cond, **options)
                    title = cloud_list.title
                    _list = cloud_list.cloud_list
                    sample_size = cloud_list.sample_size
                    if base_desc is None :
                        base_desc = cloud_list.base_desc
                    if total_flag :
                        total_flag = False
                        clouds.append(('', WordCloudHandler('Total', _list[:1], index, base_desc, sample_size))) # Total
                        clouds.append((header, WordCloudHandler(title, _list[1:], index, base_desc, sample_size)))
                    else :
                        cloud_list = _list[1:]
                        clouds.append((header, WordCloudHandler(title, cloud_list, index, base_desc, sample_size)))
                
                return BannerWordCloud(clouds, index, base_desc, sample_size=len(df))
            else :
                return self.wordcloud(index, cond=cond, **options)
            

    def table(self, index: Union[str, List[str]],
                    columns: Optional[Union[str, List[str]]] = None,
                    cond: Optional[pd.Series] = None,
                    index_meta: Optional[List[Dict[str, str]]] = None,
                    columns_meta: Optional[List[Dict[str, str]]] = None,
                    index_filter: Optional[List[Union[str, int]]] = None,
                    columns_filter: Optional[List[Union[str, int]]] = None,
                    index_sort: Optional[Literal['asc', 'desc']]=None,
                    columns_sort: Optional[Literal['asc', 'desc']]=None,
                    fill: bool = True,
                    qtype: Optional[Qtypes] = None,
                    score: Optional[int] = None,
                    conversion: bool = True,
                    top: Optional[int] = None,
                    medium: Optional[Union[int, List[int], bool]] = None,
                    bottom: Optional[int] = None,
                    reverse_rating: Optional[bool]=False,
                    min_score : Optional[int] = None,
                    net: Optional[Dict] = None,
                    aggfunc: Optional[list] = None,
                    with_value: bool = None,
                    group_name: Optional[str] = None,
                    base_desc: Optional[str] = None) -> pd.DataFrame :

            titles = self.attrs.get('title')
            metas = self.attrs.get('meta')
            
            if qtype is None :
                if isinstance(index, str) :
                    if index in titles.keys() :
                        qtype = titles[index]['type']

                if isinstance(index, list) :
                    # if index[0] in titles.keys() :
                    #     qtype = titles[index[0]]['type']
                    if all(idx in titles.keys() for idx in index) :
                        chk_type = list(set([titles[idx]['type'] for idx in index]))
                        if len(chk_type) >= 2 :
                            raise ValueError('index type is not same')
                        else :
                            qtype = chk_type[0]

            if qtype is None :
                raise ValueError('qtype is not defined')

            if net is not None :
                if not isinstance(net, dict) :
                    raise ValueError('net must be dict')
                
                for value in net.values() :
                    if not isinstance(value, list) :
                        raise ValueError(f'net value must be list / Error : {value}')
                        

            cond = (self.attrs['default_filter']) if cond is None else (self.attrs['default_filter']) & (cond)
            df = self[cond].copy()

            with_value = self.attrs['default_with_value'] if with_value is None else with_value

            filt_variables = []

            if isinstance(index, tuple) :
                index = self.ma_return(index)
                if not self.col_name_check(*index) : return

            if isinstance(columns, tuple) :
                columns = self.ma_return(columns)
                if not self.col_name_check(*columns) : return


            if index_filter is not None :
                if not isinstance(index_filter, list) :
                    raise ValueError("index_filter must be list")
            
            if columns_filter is not None :
                if not isinstance(columns_filter, list) :
                    raise ValueError("columns_filter must be list")

            # Table Header
            varable_text = []
            index_cond = None
            if isinstance(index, list) :
                if qtype == 'multiple' :
                    if index_filter is not None :
                        index_cond = ((~df[index_filter].isna()).any(axis=1)) & ((df[index_filter] != 0).any(axis=1))
                        varable_text.append(f'{index_filter[0]}-{index_filter[-1]}')
                    else :
                        index_cond = ((~df[index].isna()).any(axis=1)) & ((df[index] != 0).any(axis=1))
                        varable_text.append(f'{index[0]}-{index[-1]}')
                else :
                    index_cond = (~df[index].isna()).any(axis=1)
            else :
                varable_text.append(index)
                index_cond = ~df[index].isna()
                if index_filter is not None :
                    index_cond = (index_cond) & (df[index].isin(index_filter))

            if isinstance(columns, list) :
                if columns_filter is not None :
                    varable_text.append(f'{columns_filter[0]}-{columns_filter[-1]}')
                    columns = columns_filter
                else :
                    varable_text.append(f'{columns[0]}-{columns[-1]}')
            else :
                if columns is not None :
                    varable_text.append(columns)

                    if columns_filter is not None :
                        index_cond = (index_cond) & (df[columns].isin(columns_filter))

            # Index
            if isinstance(index, (list)) :
                filt_variables += index
            else :
                filt_variables.append(index)

            # Columns
            if isinstance(columns, (list)) :
                filt_variables += columns
            else :
                if columns is not None :
                    filt_variables.append(columns)

            filt_variables = list(set(filt_variables))

            
            df = df[index_cond][filt_variables].copy()
            if isinstance(index, list) :
                df[index] = df[index].replace(0, np.nan)
            
            if isinstance(columns, list) :
                df[columns] = df[columns].replace(0, np.nan)
            
            original_index_meta = index_meta
            original_columns_meta = columns_meta

            index_meta = self.setting_meta(original_index_meta, index, not qtype in ['number', 'float', 'text'])
            original_index_meta = index_meta
            if index_filter is not None :
                index_filter = [str(i) for i in index_filter]
                if index_meta is not None :
                    index_meta = {k: v for k, v in index_meta.items() if k in index_filter}
                else :
                    index_meta = {i: i for i in index_filter}
                

            if isinstance(index, str) and isinstance(index_meta, str) :
                index_meta = None

            if index_meta is not None and index_sort is not None :
                if index_sort == 'asc' :
                    index_meta = dict(sorted(index_meta.items()))
                    
                if index_sort == 'desc' :
                    index_meta = dict(sorted(index_meta.items(), reverse=True))

            columns_meta = self.setting_meta(original_columns_meta, columns, dup_chk=False, check_title=False)
            original_columns_meta = columns_meta
            if columns_filter is not None :
                columns_filter = [str(i) for i in columns_filter]
                if columns_meta is not None :
                    columns_meta = {k:v for k, v in columns_meta.items() if k in columns_filter}
                else :
                    columns_meta = {i: i for i in columns_filter}

            if isinstance(columns, str) and isinstance(columns_meta, str) :
                columns_meta = None

            if columns_meta is not None and columns_sort is not None :
                if columns_sort == 'asc' :
                    columns_meta = dict(sorted(columns_meta.items()))
                
                if columns_sort == 'desc' :
                    columns_meta = dict(sorted(columns_meta.items(), reverse=True))

            # Number Type
            if qtype in ['number', 'float'] :
                if aggfunc is None :
                    aggfunc = ['count', 'mean', 'max', 'median', 'min']

            # Rating Type
            if qtype == 'rating' :
                if not isinstance(index, str) :
                    raise TypeError("index must be str")
                
                top = self.attrs['default_top'] if top is None else top
                bottom = self.attrs['default_bottom'] if bottom is None else bottom
                medium = self.attrs['default_medium'] if medium is None else medium
                
                if net is not None :
                    top = None
                    bottom = None
                    medium = None

                if aggfunc is None :
                    aggfunc = ['mean']
                
                if score is None :
                    if index in metas.keys() :
                        chk_meta = metas[index]
                        values = [int(k) for k in chk_meta.keys()]
                        answers = max(values)
                        score = answers
                    else :
                        chk_idx = df[index].value_counts().index.tolist()
                        answers = max(chk_idx)
                        score = answers
                        
            # With Value
            if index_meta is not None and with_value :
                new_index_meta = {}
                for key, v in index_meta.items() :
                    new_index_meta[key] = f'[{key}] {v}'
                    
                index_meta = new_index_meta
            

            ####  Closed Ended Data Only ####
            # Add top and bottom summaries if needed
            if (qtype == 'rating') and (not index_meta) and (score is None) :
                raise ValueError("If qtype is 'rating', score or index_meta must be provided.")

            num_qtype_chk = qtype in ['number', 'float']

            total_label ='Total'

            # CrossTab Result
            result = None 
            

            if num_qtype_chk :
                if aggfunc is not None :
                    if not isinstance(aggfunc, list) :
                        raise ValueError("Aggfunc must be a list of strings.")
                else :
                    aggfunc = []

                if not 'count' in aggfunc :
                    aggfunc = ['count'] + aggfunc
                
                def number_single_table(qid, columns, aggfunc):
                    return number_with_columns(df, qid, columns, aggfunc) if columns is not None else number_total(df, qid, aggfunc)

                if isinstance(index, list):
                    number_tables = pd.concat(
                        [number_single_table(idx, columns, aggfunc).rename(index=lambda x: f'{idx}_{x}') for idx in index]
                    )

                    group_index = [
                        (titles.get(qid, {}).get('sub_title', qid), agg)
                        for qid, agg in (idx.split('_') for idx in number_tables.index)
                    ]

                    number_tables.index = pd.MultiIndex.from_tuples(group_index)
                    result = number_tables
                else:
                    result = number_single_table(index, columns, aggfunc)

                
            else :
                # MA/SA Only

                if isinstance(index, list) and qtype in ['single', 'rank'] :
                    sa_sum_table = [
                        create_crosstab(df, index=idx, columns=columns, total_label=total_label)
                        for idx in index
                    ]
                    result = pd.concat(sa_sum_table).groupby(level=0).sum()

                    result.fillna(0, inplace=True)
                    result.loc[total_label, total_label] = len(df)

                    if isinstance(columns, str):
                        column_sample_counts = df[columns].value_counts()
                        for col in result.columns:
                            if col != total_label:
                                result.loc[total_label, col] = column_sample_counts.get(col, 0)

                    elif isinstance(columns, list):
                        for col in result.columns:
                            if col != total_label:
                                sample_count = df[df[col].notna()].shape[0]
                                result.loc[total_label, col] = sample_count
                else :
                    result = create_crosstab(df,
                                            index=index,
                                            columns=columns,
                                            total_label=total_label)
                    
                
                result.fillna(0, inplace=True)
                result = result.astype(int)

                default_index = [idx for idx in result.index if idx != total_label]

                # Rating Type
                if qtype in ['rating'] :
                    # score_min = min(default_index)
                    if score is None :
                        score_meta = [int(k) for k in index_meta.keys()]
                        score = max(score_meta)
                        # score_min = max(score_meta)
                    
                    min_score = 1 if min_score is None else min_score
                    scores = [i for i in range(min_score, score+1)]
                    default_index = [str(idx) for idx in scores]
                    result = rating_netting(result, 
                                            scores, 
                                            reverse_rating=reverse_rating, 
                                            top=top, 
                                            bottom=top, 
                                            medium=medium)
                    
                
                # Netting
                if net is not None :
                    net_result = var_netting(result, net)
                    result = pd.concat([result, net_result])

                
                if aggfunc is not None : 
                    calc_result = None 
                    if columns is not None :
                        calc_result = number_with_columns(df, index, columns, aggfunc)
                    else :
                        calc_result = number_total(df, index, aggfunc)

                    if calc_result is not None : 
                        if qtype in ['rating'] and (score is not None) and (score >= 4) :
                            if 'mean' in calc_result.index and conversion :
                                conversion_index = '100 point conversion'
                                calc_result.loc[conversion_index, :] = [0 if (pd.isna(i)) or (i == 0) else ((i-1)/(score-1))*100 for i in calc_result.loc['mean', :].values]
                    
                    
                    result = pd.concat([result, calc_result])

            
            if not num_qtype_chk :
                index_order = [total_label] + [i for i in result.index.to_list() if not i == total_label]
                result = result.reindex(index_order)
                
                
                # Process index metadata
                if index_meta :
                    result.index = result.index.map(str)
                    index_order = [total_label] + [m for m in index_meta.keys()]
                    if original_index_meta is not None :
                        index_order = index_order + [idx for idx in result.index if (not idx in index_order) and (not idx in original_index_meta.keys())]
                    else :
                        index_order = index_order + [idx for idx in result.index if (not idx in index_order)]
                    
                    result = add_missing_indices(result, index_order)
                    result = result.loc[index_order, :]
                    result.rename(index=index_meta, inplace=True)
                    
            # Process columns metadata
            if columns and columns_meta:
                result.columns = result.columns.map(str)
                columns_order = [total_label] + [m for m in columns_meta.keys()]
                if original_columns_meta is not None :
                    columns_order = columns_order + [idx for idx in result.columns if (not idx in columns_order) and (not idx in original_columns_meta.keys())]
                else :
                    columns_order = columns_order + [idx for idx in result.columns if (not idx in columns_order)]
                
                result = add_missing_indices(result.T, columns_order).T
                result = result.loc[:, columns_order]
                result.rename(columns=columns_meta, inplace=True)

            if not fill :
                result.fillna(0, inplace=True)
                result = result.loc[(result != 0).any(axis=1), (result != 0).any(axis=0)]

            if base_desc is None :
                sample_count = len(self.index.to_list())
                tot = result.iloc[0, 0]
                tot = 0 if pd.isna(tot) else tot
                all_count = int(tot)
                
                if sample_count == all_count :
                    base_desc = 'All Base'
                else :
                    sample_ratio = round(all_count/sample_count, 2) * 100
                    base_desc = f'Not All Base ({sample_ratio:.2f}%)'
            
            result = CrossTabs(result)
            result.attrs['type'] = qtype
            result.attrs['qid'] = varable_text
            result.attrs['base'] = base_desc
            result.attrs['group_name'] = group_name

            return result
            #### ======================= ####


    def sa_netting(self, 
                name: str,
                code_dict: Dict,
                title: Optional[str] = None,
                sub_title: Optional[str] = None,
                qtype: Literal['single', 'rating', 'rank'] = 'single') :
        if not isinstance(code_dict, dict):
            raise TypeError("code_dict must be a dictionary")

        if not isinstance(name, str):
            raise TypeError("name must be a string")
        
        df = self
        code_meta = {}
        result_series = pd.Series(index=df.index, dtype=int)  # 빈 시리즈 생성
        for code, [label, cond] in code_dict.items():
            result_series.loc[cond] = code  # 조건에 맞는 값 할당
            code_meta[str(code)] = label
        
        result = self.assign(**{name: result_series})
        self.__dict__.update(result.__dict__)
        self[name] = self[name].astype(int)

        self.set_meta(name, code_meta)
        self.set_title(
            qid=name,
            qtype=qtype,
            title=title if title is not None else name,
            sub_title=sub_title,
        )

        result = self.table(name)
        return result
    

    def ma_netting(self, 
                name: str,
                code_dict: Dict,
                title: Optional[str] = None,
                sub_title: Optional[str] = None,
                qtype: Literal['multiple'] = 'multiple') :
        
        if not isinstance(code_dict, dict):
            raise TypeError("code_dict must be a dictionary")

        if not isinstance(name, str):
            raise TypeError("name must be a string")
        
        df = self
        ma_net_var = {}

        self.set_title(
            qid=name,
            qtype=qtype,
            title=title if title is not None else name,
            sub_title=sub_title,
        )

        for code, [label, cond] in code_dict.items():
            net_result = pd.Series(0, index=df.index, dtype=int)  # 빈 시리즈 생성: Default 0
            net_result.loc[cond] = 1  # 조건에 맞는 값 할당
            qlabel = f'{name}#{code}'
            ma_net_var[qlabel] = label

            result = self.assign(**{qlabel: net_result})
            self.__dict__.update(result.__dict__)
            self[qlabel] = self[qlabel].astype(int)

            self.set_meta(qlabel, label)
            self.set_title(
                qid=qlabel,
                qtype=qtype,
                title=title if title is not None else qlabel,
                sub_title=label,
                vgroup=name
            )
        
        self.set_meta(name, ma_net_var)
        result = self.table(list(ma_net_var.keys()))

        return result

        

    def set_banner(self, banner_list: Optional[Union[str, List, Dict]] = None):
        self.attrs['banner'] = banner_list


    def total_row(self,
                  index: Union[str, List[str]],
                  columns: Optional[Union[str, List, Dict]]=None, 
                  cond=None):
        titles = self.attrs.get('title')
        if not isinstance(index, (str, list)) :
            raise ValueError("index must be str or list")

        if columns is None:
            banner = self.attrs.get('banner')
            columns = banner

        if isinstance(columns, dict) :
            new_columns = []

            total_col = self.table(index, cond=cond)
            new_columns.extend(('', col) for col in total_col.columns)
            tables = [
                (titles.get(col_head, {}).get('title', col_head), self.table(index, col, cond=cond))
                for col_head, col in columns.items()
            ]

            result = pd.concat([t.loc[:, t.columns[1:]] for _, t in tables], axis=1)
            for head, table in tables:
                new_columns.extend((head, col) for col in table.columns[1:])  # Total 제외
            
            result = pd.concat([total_col, result], axis=1)
            result.columns = pd.MultiIndex.from_tuples(new_columns)
            result = result.loc[:, ~result.columns.duplicated()]

        else :
            result = self.table(index, columns, cond=cond)
        
        result.fillna(0, inplace=True)
        
        return result.loc[['Total'], :]

    def index_net(self, 
                  total_index: Union[str, List],
                  net_table: Dict,
                  columns: Optional[ColumnsWithHeader] = None, 
                  cond: Optional[pd.Series] = None,
                  base_desc: Optional[str] = None,
                  total_net:bool = True) :
        if not isinstance(total_index, (str, list)) :
            raise ValueError("index must be str or list")
        if not isinstance(net_table, dict) :
            raise ValueError("net_table must be dict")

        if columns is None:
            banner = self.attrs.get('banner')
            columns = banner
        
        index_group = []
        total_row = self.total_row(total_index, columns, cond=cond)
        index_group.extend(('', c) for c in total_row.index)
        titles = self.attrs.get('title')
        total_label = 'Total'
        concat_table = []        

        for net_name, table in net_table.items() :
            if total_net :
                table.rename(index={total_label: '▣ %s'%(titles.get(net_name, {}).get('title', net_name))}, inplace=True)
                concat_table.append(table.loc[:, [t for t in table.columns if t != total_label]])
                index_group.extend((net_name, idx) for idx in table.index)
            else :
                concat_table.append(table.loc[[idx for idx in table.index if idx != total_label], [t for t in table.columns if t != total_label]])
        
        result = pd.concat([total_row, *concat_table])
        
        if total_net :
            result.index = pd.MultiIndex.from_tuples(index_group)

        index_name = total_index
        if isinstance(total_index, list) :
            if len(total_index) == 1 :
                index_name = total_index[0]
            else :
                index_name = f'{total_index[0]}-{total_index[-1]}'

        if base_desc is None:
            sample_count = len(self.index.to_list())
            tot = result.iloc[0, 0] if not pd.isna(result.iloc[0, 0]) else 0
            all_count = int(tot)
            if sample_count == all_count:
                base_desc = 'All Base'
            else:
                sample_ratio = round(all_count / sample_count, 2) * 100
                base_desc = f'Not All Base ({sample_ratio:.0f}%)'

        if total_net :
            result.index.names = pd.Index([index_name, 'Index'])

        result = CrossTabs(result)
        result.attrs['type'] = 'index_net'
        result.attrs['qid'] = index_name
        result.attrs['base'] = base_desc

        return result



    def proc(self, 
            index: IndexWithTypes, 
            columns: Optional[ColumnsWithHeader] = None, 
            cond: Optional[pd.Series] = None,
            fill: bool = True, 
            group_name: Optional[str] = None,
            base_desc: Optional[str] = None,
            **options):
        
        if not isinstance(index, (str, list)):
            raise TypeError("index must be a string or list")

        if columns is None:
            banner = self.attrs.get('banner')
            columns = banner

        titles = self.attrs.get('title')

        if columns is not None :
            if not isinstance(columns, (str, list, dict)):
                raise TypeError("columns must be a string, list, or dictionary")

        index_name = index
        tables = []
        new_index = []
        new_columns = []
        qtypes = []

        if isinstance(index, list) :
            if len(index) == 1 :
                index_name = index[0]
            else :
                index_name = f'{index[0]}-{index[-1]}'


        # By Column Table
        if isinstance(columns, dict) :
            total_col = self.table(index, cond=cond, **options)
            new_columns.extend(('', col) for col in total_col.columns)
            tables = [
                (titles.get(col_head, {}).get('title', col_head), self.table(index, col, cond=cond, **options))
                for col_head, col in columns.items()
            ]
            

            result = pd.concat([t.loc[:, t.columns[1:]] for _, t in tables], axis=1)
            
            
            for head, table in tables:
                # new_columns.append(('', table.columns[0]))
                qtypes.append(table.attrs['type'])
                new_columns.extend((head, col) for col in table.columns[1:])  # Total 제외

            result = pd.concat([total_col, result], axis=1)

        elif isinstance(columns, (str, list)) :
            result = self.table(index, columns, cond=cond, **options)
        
        else :
            result = self.table(index, cond=cond, **options)
            qtypes = result.attrs['type']
        
        if new_index :
            result.index = pd.MultiIndex.from_tuples(new_index)
            result = result.loc[~result.index.duplicated(), :]

        if new_columns :
            result.columns = pd.MultiIndex.from_tuples(new_columns)
            result = result.loc[:, ~result.columns.duplicated()]

        qtype = list(set(qtypes))

        if base_desc is None:
            sample_count = len(self.index.to_list())
            tot = result.iloc[0, 0] if not pd.isna(result.iloc[0, 0]) else 0
            all_count = int(tot)
            if sample_count == all_count:
                base_desc = 'All Base'
            else:
                sample_ratio = round(all_count / sample_count, 2) * 100
                base_desc = f'Not All Base ({sample_ratio:.0f}%)'

        result.index.name = 'Index'

        if not fill:
            result = result.loc[(result != 0).any(axis=1), (result != 0).any(axis=0)]
        
        result = CrossTabs(result)
        result.attrs['type'] = qtype
        result.attrs['qid'] = index_name
        result.attrs['base'] = base_desc
        result.attrs['group_name'] = group_name
        
        return result



    def grid_summary(self, index: Union[List[str], List[List[str]], Tuple[str], CrossTabs],
                    summary_name: str = '',
                    cond: Optional[pd.Series] = None,
                    base_desc: Optional[str] = None,
                    fill: bool = True,
                    **kwargs):
        if not isinstance(index, list):
            raise ValueError(f'index must be list : {index}')

        cond = self.attrs['default_filter'] if cond is None else self.attrs['default_filter'] & cond

        summary_df = []
        multi_col = []
        titles = self.attrs.get('title', {})

        for idx in index:
            if isinstance(idx, CrossTabs):
                table = idx
            else:
                table = self.table(idx, cond=cond, **kwargs)

            index_name = idx[0] if isinstance(idx, list) else idx
            sub_title = titles.get(index_name, {}).get('sub_title', titles.get(index_name, {}).get('title', ''))
            if isinstance(idx, list):
                index_name = idx[0] if len(idx) == 1 else f'{idx[0]}-{idx[-1]}'

            multi_col.append((index_name, sub_title))
            table = table.rename(columns={'Total': index_name})
            summary_df.append(table)

        qtypes = list(set(x.attrs['type'] for x in summary_df))
        summary = pd.concat(summary_df, axis=1)

        summary.index.name = 'Index'
        result = summary

        result.columns = pd.MultiIndex.from_tuples(multi_col)

        if base_desc is None:
            sample_count = len(self.index)
            all_count = summary.iloc[0, :].tolist()
            base_desc = 'Difference Total' if len(set(all_count)) > 1 else (
                'All Base' if sample_count == all_count[0] else f'Not All Base ({round(all_count[0] / sample_count, 2) * 100:.0f}%)'
            )

        if not fill:
            result = result.loc[:, (result != 0).any(axis=0)]

        var_names = [f'{i[0]}-{i[-1]}' if isinstance(i, list) else i for i in index]
        var_names = f'{var_names[0]} to {var_names[-1]}'

        result = CrossTabs(result)
        result.attrs['type'] = qtypes
        result.attrs['qid'] = index_name
        result.attrs['base'] = base_desc
        result.attrs['group_name'] = None

        return result



    def get_title(self, qid: str) :
        title = self.attrs.get('title')
        if title :
            if qid in title.keys():
                qtitle = title[qid]['title']
                sub_title = title[qid]['sub_title']
                if sub_title :
                    return f'{qtitle}_{sub_title}'
                else :
                    return qtitle
            else :
                return qid
        else :
            return qid


    def proc_append(self, 
                    table_id: Union[str, tuple], 
                    table: Union[pd.DataFrame, CrossTabs],
                    ai: bool = False,
                    model: Literal['gpt-4o', 'gpt-4o-mini', 'llama3', 'llama3.1'] = 'gpt-4o-mini',
                    prompt: Optional[str] = None,
                    heatmap: Optional[bool] = True,) :
        if not isinstance(table_id, (str, tuple)) :
            raise ValueError(f'table_id must be str or tuple')
        
        table_name = table_id
        table_desc = None
        proc_result = self.attrs['proc_result']

        if isinstance(table_id, tuple) :
            if len(table_id) != 2 :
                raise ValueError(f'table_id must be tuple with 2 elements')
            else :
                table_name = table_id[0]
                table_desc = table_id[1]
                

        if isinstance(table, (WordCloudHandler, BannerWordCloud)) : 
            # WordCloud Append
            proc_result[table_name] = {
                'desc': table_desc, 
                'table': table,
                'ai': None
            }

            table.show(desc=table_desc)
        
        else :
            # CE Table Append
            if not isinstance(table, (pd.DataFrame, CrossTabs)) :
                raise ValueError(f'table must be pd.DataFrame or CrossTabs')
            
            if not isinstance(table, CrossTabs) :
                table = CrossTabs(table)
            
            
            if table_name in proc_result.keys() :
                print(f'⚠️ result title already exists : {table_name}')
            
            chat_result = None
            
            table_type = table.attrs.get('type', None)
            base = table.attrs.get('base', None)
            qid = table.attrs.get('qid', None)
            group_name = table.attrs.get('group_name', None)
            if ai :
                chat_result = table.chat_ai(model=model, 
                                            prompt=prompt, 
                                            with_table=False,
                                            lang=self.attrs.get('chat_lang', 'korean'),
                                            table_type=table_type,
                                            sub_title=table_desc)

            proc_result[table_name] = {
                'desc': table_desc, 
                'table': table,
                'ai': chat_result,
                'base': base,
                'group_name': group_name,
                'qid': qid
            }

            table_html = None
            if table_type in ['number', 'float', 'text'] :
                table_html = table.to_html(escape=False, index=True, border=0, classes='table table-striped table-hover')
            else :
                table_html = table.ratio(ratio_round=0, heatmap=heatmap).to_html()

            table_desc_html = f"""<div style="font-size: 1rem; padding: 7px; max-width: 600px; margin-bottom: 7px; font-weight: bold;">
                    {f'<span style="color:#2d6df6;">[{group_name}]</span><br/>' if group_name is not None else ''}{table_desc}
            </div>"""

            if base is not None :
                table_desc_html += f"""<div style="font-size: 0.8rem; padding: 7px; max-width: 600px; font-style: italic; margin-bottom: 3px;">
                    📌 {base}
            </div>"""

            table_analysis_html = f"""<div style="font-weight: bold; font-size: 0.8rem; padding: 7px; max-width: 700px; margin-bottom: 7px;border: 1px solid #2d6df6; border-radius: 5px;">
                    {chat_result}
            </div>
            """

            table_id_html = f"""
                <div style="width:fit-content;padding: 7px; font-size:1rem;font-weight:bold; background-color: #2d6df6; border-radius: 5px; color:white; margin-bottom: 7px;">
                    {table_name}
                </div>
                {table_desc_html if table_desc is not None else ''}
                {table_analysis_html if chat_result is not None else ''}
                <div>
                    {table_html}
                </div>
            """
            display(HTML(table_id_html))
    
    
    def load_log(self, pkl_path: str) :
        self.attrs['proc_result'] = load_from_pickle(pkl_path)
        print(f'✅ Load log : {pkl_path}')

    def proc_export_excel(self, file_name: str, heatmap: bool = False) :
        total_label = 'Total'
        proc_result = self.attrs.get('proc_result', None)
        if not proc_result : 
            raise ValueError('No result to export')
        
        # pickle 파일로 저장
        # 현재 날짜와 시간을 형식에 맞게 가져오기
        current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
        save_to_pickle(proc_result, f'proccessing_result_{current_time}.pkl', 'LOG')

        excel.ExcelFormatter.header_style = {
            "font": {"bold": True, "size": 9},
            "borders": {
                "top": "medium",
                "right": "thin",
                "bottom": "medium",
                "left": "thin",
            },
            "alignment": {"horizontal": "center", "vertical": "center", "wrapText": True},
            "fill": {"bgColor": "#DCE6F1"}
        }

        # 엑셀 파일 생성
        # file_name = get_versioned_filename(file_name)

        writer = pd.ExcelWriter(file_name, engine='xlsxwriter')
        workbook = writer.book

        # 목차 시트 추가
        index_sheet = workbook.add_worksheet('Index')
        index_format = workbook.add_format({
            'align': 'center', 
            'bold': True, 
            'underline': True, 
            'font_size': 11,
            'font_color': '#2d6df6', 
            'border': 1})
        
        desc_format = workbook.add_format({
            'align': 'left', 
            'italic': True,
            'font_size': 11,
            'border': 1})

        qid_format = workbook.add_format({
            'align': 'center', 
            'font_size': 11,
            'border': 1})

        head_format = workbook.add_format({
            'align': 'center', 
            'font_size': 12,
            'bold': True, 
            'border': 1, 
            'bg_color': '#DDEBF7'})

        table_head = workbook.add_format({
            'align': 'center',
            'font_size': 9,
            'bold': True,
            'border': 1,
            'top': 2,
            'bottom': 2,
            'bg_color': '#DCE6F1',
        })

        image_cell = workbook.add_format({
            'align': 'center',
            'font_size': 9,
            'bold': True,
            'border': 1,
            'top': 2,
            'bottom': 2,
            'left': 2,
            'right': 2,
        })

        image_var_cell = workbook.add_format({
            'align': 'center',
            'font_size': 9,
            'bold': True,
            'border': 1,
            'top': 2,
            'bottom': 2,
            'bg_color': '#DCE6F1',
            'valign': 'top',
        })

        zero_float_format = workbook.add_format({
            'num_format': '0',
            'align': 'center',
            'border': 1,
            'font_size': 9,
        })

        float_format = workbook.add_format({
            'num_format': '0.00',
            'align': 'center',
            'border': 1,
            'font_size': 9,
        })

        default_format = workbook.add_format({
            'align': 'center',
            'border': 1,
            'font_size': 9,
        })
        
        total_format = workbook.add_format({
            'num_format': '0',
            'align': 'center',
            'border': 1,
            'font_size': 9,
            'bold': True,
        })


        # 데이터프레임을 저장할 시트 추가
        data_sheet = workbook.add_worksheet('Table')
        wc_sheet = workbook.add_worksheet('WordCloud')

        # B열 틀고정
        data_sheet.freeze_panes(0, 2)
        wc_sheet.freeze_panes(0, 2)

        # 목차 시트에 하이퍼링크 추가
        row = 1
        col = 0
        data_start_row = 2
        oe_start_row = 2


        index_sheet.write(0, 0, 'Table Index', head_format)
        index_sheet.write(0, 1, 'Table Description', head_format)
        index_sheet.write(0, 2, 'Variable', head_format)
        index_sheet.write(0, 3, 'Base', head_format)

        # 너비 설정
        index_sheet.set_column('A:A', 20) # Talbe Index
        index_sheet.set_column('B:B', 60) # Table Description
        index_sheet.set_column('C:C', 20) # Variable
        index_sheet.set_column('D:D', 30) # Base

        data_sheet.set_column('A:A', 15)
        data_sheet.set_column('B:B', 30)
        data_sheet.set_column('C:C', 7)

        for key, table_attrs in proc_result.items():
            # try :
                result = table_attrs.get('table', None)
                desc = table_attrs.get('desc', None)
                ai = table_attrs.get('ai', None)

                if isinstance(result, WordCloudHandler) :
                    # WordCloud
                    index_sheet.write_url(row, col, f'internal:WordCloud!A{oe_start_row+1}', string=key, cell_format=index_format)
                    index_sheet.write(row, col + 1, desc, desc_format)
                    base_desc = result.base_desc
                    qid_name = result.variable
                    sample_size = result.sample_size

                    index_sheet.write(row, col + 2, qid_name, qid_format)
                    index_sheet.write(row, col + 3, base_desc, qid_format)

                    row += 1
                    cl_col = 2
                    cloud_list = result.cloud_list

                    wc_sheet.merge_range(oe_start_row-1, col, oe_start_row-1, col+1, key, table_head)
                    wc_sheet.merge_range(oe_start_row, col, oe_start_row, col+1, desc, table_head)
                    
                    if sample_size is not None :
                        qid_name = f'{qid_name} ({sample_size}\'s)'
                    wc_sheet.merge_range(oe_start_row+1, col, oe_start_row+9, col+1, qid_name, image_var_cell)

                    if isinstance(cloud_list, WordCloud) :
                        with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as tmpfile:
                            cloud_list.to_file(tmpfile.name)
                            image_path = tmpfile.name
                        
                        wc_sheet.insert_image(oe_start_row, cl_col, image_path, {'x_scale': 0.4, 'y_scale': 0.4})
                        merge_row = oe_start_row-1
                        wc_sheet.merge_range(merge_row, cl_col, merge_row, cl_col+4, 'Total', table_head)
                    
                    if isinstance(cloud_list, list) :
                        for cl in cloud_list :
                            head = cl[0]
                            wc = cl[1]

                            with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as tmpfile:
                                wc.to_file(tmpfile.name)
                                image_path = tmpfile.name

                            wc_sheet.insert_image(oe_start_row, cl_col, image_path, {'x_scale': 0.4, 'y_scale': 0.4})

                            merge_row = oe_start_row-1
                            wc_sheet.merge_range(merge_row, cl_col, merge_row, cl_col+4, head, table_head)
                            cl_col += 5
                            
                    oe_start_row += 14

                elif isinstance(result, BannerWordCloud) :
                    # WordCloud
                    index_sheet.write_url(row, col, f'internal:WordCloud!A{oe_start_row+1}', string=key, cell_format=index_format)
                    index_sheet.write(row, col + 1, desc, desc_format)
                    base_desc = result.base_desc
                    qid_name = result.variable
                    sample_size = result.sample_size

                    index_sheet.write(row, col + 2, qid_name, qid_format)
                    index_sheet.write(row, col + 3, base_desc, qid_format)
                    
                    group_cloud_list = result.cloud_list

                    row += 1
                    cl_col = 2

                    wc_sheet.merge_range(oe_start_row+1, col, oe_start_row+1, col+1, key, table_head)
                    wc_sheet.merge_range(oe_start_row+2, col, oe_start_row+2, col+1, desc, table_head)
                    
                    if sample_size is not None :
                        qid_name = f'{qid_name} ({sample_size}\'s)'

                    wc_sheet.merge_range(oe_start_row+3, col, oe_start_row+12, col+1, qid_name, image_var_cell)

                    for gcl in group_cloud_list :
                        head = gcl[0]
                        gwc = gcl[1]
                        group_start_col = cl_col
                        group_cl = cl_col
                        for cl in gwc.cloud_list :
                            col_name = cl[0]
                            wc = cl[1]
                            with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as tmpfile:
                                wc.to_file(tmpfile.name)
                                image_path = tmpfile.name

                            wc_sheet.insert_image(oe_start_row+3, cl_col, image_path, {'x_scale': 0.4, 'y_scale': 0.4})

                            wc_sheet.merge_range(oe_start_row+2, cl_col, oe_start_row+2, cl_col+4, col_name, table_head)
                            wc_sheet.merge_range(oe_start_row+3, cl_col, oe_start_row+12, cl_col+4, None, image_cell)
                            cl_col += 5
                            group_cl += 4
                        
                        wc_sheet.merge_range(oe_start_row+1, group_start_col, oe_start_row+1, group_cl+(len(gwc.cloud_list)-1), head, table_head)

                    oe_start_row += 16
                    
                else :
                    # CE TABLE
                    use_qid = table_attrs.get('qid', '')
                    base_text = table_attrs.get('base', '')
                    group_name = table_attrs.get('group_name', '')
                    
                    if not isinstance(result.index, pd.MultiIndex) :
                        result.index = pd.MultiIndex.from_tuples([('' if group_name is None else group_name, i) for i in result.index])
                        result.index.names = pd.Index([use_qid, base_text])
                    
                    if not isinstance(result.columns, pd.MultiIndex) :
                        result.columns = pd.MultiIndex.from_tuples([('', i) for i in result.columns])
                        result.columns.names = pd.Index([use_qid, base_text])

                    index_sheet.write_url(row, col, f'internal:Table!A{data_start_row+1}', string=key, cell_format=index_format)
                    index_sheet.write(row, col + 1, desc, desc_format)

                    index_sheet.write(row, col + 2, use_qid, qid_format)
                    index_sheet.write(row, col + 3, base_text, qid_format)

                    if isinstance(result, CrossTabs) :                        
                        resurt_type = result.attrs.get('type', None)
                        
                        if isinstance(resurt_type, list) :
                            if all(not x in ['number', 'float'] for x in resurt_type) :
                                result = result.ratio(ratio_round=None, heatmap=False)
                            
                        elif not resurt_type in ['number', 'float'] :
                            result = result.ratio(ratio_round=None, heatmap=False)
                        

                    row += 1

                    result.to_excel(writer, 
                                    sheet_name='Table', 
                                    startrow=data_start_row, 
                                    startcol=0, engine='openpyxl')

                    data_sheet.merge_range(data_start_row, col, data_start_row, col+1, key, table_head)
                    data_sheet.merge_range(data_start_row+1, col, data_start_row+1, col+1, desc, table_head)
                    
                    head_row = data_start_row + 1
                    blank_row = data_start_row + 2
                    format_start = data_start_row+3


                    set_group = []
                    for col_idx, col_name in enumerate(result.columns) :
                        set_col = col_idx+2
                        group_col_name = col_name[0]
                        set_col_name = col_name[-1]
                        if col_idx == 0 and set_col_name == 'Total' :
                            data_sheet.write(data_start_row, set_col, None, table_head)
                            data_sheet.write(head_row, set_col, None, table_head)
                            data_sheet.write(blank_row, set_col, set_col_name, table_head)
                        else :
                            data_sheet.write(blank_row, set_col, set_col_name, table_head)
                            
                            if not group_col_name in set_group :
                                if group_col_name != '' and group_col_name is not None :
                                    set_group.append(group_col_name)
                                    merge_col_count = len([x for x in result.columns if x[0] == group_col_name]) - 1
                                    end_col = set_col+merge_col_count
                                    for m in range(set_col, end_col+1) :
                                        data_sheet.write(data_start_row, m, f'#{len(set_group)}', table_head)
                                    
                                    if merge_col_count > 0 :
                                        data_sheet.merge_range(head_row, set_col, head_row, set_col+merge_col_count, group_col_name, table_head)
                                    else :
                                        data_sheet.write(head_row, set_col, group_col_name, table_head)
                    
                    if len(set_group) == 1 :
                        data_sheet.write(data_start_row, 2, None, table_head)

                    last_row = None
                    for df_row, i in enumerate(range(format_start, format_start+len(result))) :
                        for df_col, j in enumerate(range(2, len(result.columns)+2)) :
                            cell_value = result.iloc[df_row, df_col]
                            if pd.isna(cell_value) :
                                cell_value = '-'

                            df_row_name = result.index[df_row]

                            if isinstance(result.index, pd.MultiIndex) :
                                df_row_name = result.index[df_row][-1]

                            if df_row_name == 'Total' :
                                data_sheet.write(i, j, cell_value, total_format)
                            else :
                                if pd.isna(cell_value) :
                                    data_sheet.write(i, j, cell_value, default_format)
                                else :
                                    data_sheet.write(i, j, cell_value, zero_float_format)
                                    if heatmap :
                                        if cell_value != '-' :
                                            bg_color = calculate_bg_color(float(cell_value))
                                            data_sheet.write(i, j, cell_value, workbook.add_format({
                                                'num_format': '0',
                                                'align': 'center',
                                                'border': 1,
                                                'font_size': 9,
                                                'bg_color': bg_color
                                            }))
                                    if df_row_name in ['mean', 'min', 'max', 'std', 'median', '100 point conversion'] :
                                        data_sheet.write(i, j, cell_value, float_format)

                    last_row = i + 1

                    if ai is not None :
                        ai_result_format = workbook.add_format({
                            'font_size': 9,
                            'text_wrap': True,  # 자동 줄바꿈 설정
                            'valign': 'top',
                            'align': 'left',
                        })
                        data_sheet.merge_range(last_row, col+2, last_row, col+12, ai, ai_result_format)
                        data_sheet.set_row(last_row, 150)

                    data_start_row += len(result) + 6  # 3행 간격
            
            # except :
            #     print(f"⚠️ Error in {key}")

            
        writer.close()
        
        wb = load_workbook(file_name)
        ws = wb['Table']
        
        # 열 B의 서식 설정: 오른쪽 정렬
        for cell in ws['B']:
            cell.alignment = Alignment(horizontal='right')

        # Last Column


        wb.save(file_name)

        display(HTML(f"""
<div>💾 {file_name} : Complete</div>
"""))
    
    def meta(self, qid: Union[str, list]) :
        if not isinstance(qid, (str,  list)) :
            raise ValueError("qid must be str or list")

        meta = self.attrs.get('meta')

        if not meta :
            raise ValueError("There is no meta data")
        
        
        if isinstance(qid, str) :
            if not qid in meta.keys() :
                raise ValueError(f"There is no meta data for {qid}")
            return meta[qid]
        elif isinstance(qid, list) :
            qid_chk = [x for x in qid if not x in meta.keys()]

            if qid_chk :
                raise ValueError(f"There is no meta data for {qid_chk}")
            
            return {x: meta[x] for x in qid}

    def title(self, qid: Union[str, list]) :
        if not isinstance(qid, (str,  list)) :
            raise ValueError("qid must be str or list")

        title = self.attrs.get('title')

        if not title :
            raise ValueError("There is no title data")

        if isinstance(qid, str) :
            if not qid in title.keys() :
                raise ValueError(f"There is no title data for {qid}")
            return title[qid]
        elif isinstance(qid, list) :
            qid_chk = [x for x in qid if not x in title.keys()]

            if qid_chk :
                raise ValueError(f"There is no title data for {qid_chk}")
            
            return {x: title[x] for x in qid}
    
    def meta_validation(self, meta_data: Union[str, dict]) :
        if isinstance(meta_data, str) :
            return meta_data
        else :
            if not isinstance(meta_data, dict) :
                raise ValueError("meta_data must be str or dict")

            to_str_meta = {str(k): v for k, v in meta_data.items()}

        return to_str_meta

    def set_meta(self, qid: str, meta_data: Union[List[Union[Dict, Tuple]], str]) :
        if not isinstance(qid, str) :
            raise ValueError("qid must be str")
        
        meta = self.attrs.get('meta')

        if qid in meta.keys() :
            raise ValueError(f"qid already exists")

        self.attrs.get('meta')[qid] = self.meta_validation(meta_data)

    
    def update_meta(self, qid: str, meta_data: Dict) :
        if not isinstance(qid, str) :
            raise ValueError("qid must be str")
        
        meta = self.attrs.get('meta')

        if not qid in meta.keys() :
            raise ValueError(f"qid does not exist")

        self.attrs.get('meta')[qid] = self.meta_validation(meta_data)


    def title_validation(self, qtype: Qtypes, title: str, sub_title: Optional[str] = None, vgroup: Optional[str] = None) :
        title_meta = self.attrs.get('title')

        if not isinstance(title, str) :
            raise ValueError("title must be str")

        if not qtype in ['single', 'rating', 'rank', 'ranksort', 'multiple', 'number', 'float', 'text'] :
            raise ValueError("qtype must be one of'single', 'rating', 'rank', 'ranksort','multiple', 'number', 'float', 'text'")
        
        if sub_title is not None :
            if not isinstance(sub_title, str) :
                raise ValueError("sub_title must be str")
        
        if vgroup is not None :
            if not isinstance(vgroup, str) :
                raise ValueError("vgroup must be str")
            
            if not vgroup in title_meta.keys() :
                raise ValueError(f"vgroup does not exist in title meta (vgroup={vgroup})")
        
        return {
            'type': qtype,
            'title': title,
            'sub_title': sub_title,
            'vgroup': vgroup
        }

    def set_title(self, qid: str, qtype: Qtypes, title: str, sub_title: Optional[str] = None, vgroup: Optional[str] = None) :
        if not isinstance(qid, str) :
            raise ValueError("qid must be str")
        
        title_meta = self.attrs.get('title')

        if qid in title_meta.keys() :
            raise ValueError(f"qid already exists")

        
        self.attrs.get('title')[qid] = self.title_validation(qtype, title, sub_title, vgroup)


    def update_title(self, qid: str, qtype: Qtypes, title: str, sub_title: Optional[str] = None, vgroup: Optional[str] = None) :
        if not isinstance(qid, str) :
            raise ValueError("qid must be str")
        
        title_meta = self.attrs.get('title')

        if not qid in title_meta.keys() :
            raise ValueError(f"qid already exists")

        
        self.attrs.get('title')[qid] = self.title_validation(qtype, title, sub_title, vgroup)



def get_css(path: os.path) -> str:
    css_file_path = os.path.join(path)
    css = None
    try:
        with open(css_file_path, 'r') as file:
            css_content = file.read()
        css = f"""
<style>
{css_content}
</style>
"""
    except Exception as e:
        print(f"Failed to load CSS file: {e}")

    return css

def convert_columns_to_nullable_int(df):
    for col in df.columns:
        if pd.api.types.is_numeric_dtype(df[col]):
            # 소수점이 포함된 데이터가 있는지 확인
            if all(df[col].dropna() == df[col].dropna().astype(int)):
                df[col] = df[col].astype(pd.Int64Dtype())
            else:
                df[col] = df[col].astype(float)
    return df

def SetUpDataCheck(dataframe: pd.DataFrame, **kwargs) :
    module_path = os.path.dirname(__file__)
    css = get_css(os.path.join(module_path, 'styles.css'))
    display(HTML(css))
    df = convert_columns_to_nullable_int(dataframe)
    return DataCheck(df, css=css, **kwargs)


class DefaultArgs(TypedDict, total=False):
    top: Union[int, List[int]]
    medium: Union[bool, int, List[int]]
    bottom: Union[int, List[int]]
    with_value: bool


def DataProcessing(dataframe: pd.DataFrame, 
                   keyid: Optional[str] = None,
                   default_args: Optional[DefaultArgs] = {
                    'top': 2,
                    'medium': True,
                    'bottom': 2,
                    'with_value': False,
                    'chat_lang': 'korean'
                    }):
    module_path = os.path.dirname(__file__)
    css_path = os.path.join(os.path.dirname(module_path), 'dataCheck')
    css = get_css(os.path.join(css_path, 'styles.css'))
    display(HTML(css))
    df = convert_columns_to_nullable_int(dataframe)

    default_args = {f'default_{key}': value for key, value in default_args.items()}

    return DataCheck(df, 
                     css=css, 
                     keyid=keyid,
                     **default_args)


def DecipherDataProcessing(dataframe: pd.DataFrame, 
                           keyid: Optional[str] = "record",
                           map_json: Optional[str] = None,
                           meta_path: Optional[str] = None,
                           title_path: Optional[str] = None,
                           chat_lang: Optional[Literal['korean', 'english']] = 'korean',
                           default_args: Optional[DefaultArgs] = {
                               'top': 2,
                               'medium': True,
                               'bottom': 2,
                               'with_value': False,
                               'chat_lang': 'korean'
                           }) :
    module_path = os.path.dirname(__file__)
    css_path = os.path.join(os.path.dirname(module_path), 'dataCheck')
    css = get_css(os.path.join(css_path, 'styles.css'))
    display(HTML(css))
    df = convert_columns_to_nullable_int(dataframe)

    metadata = None
    title = None

    if map_json is None :
        if meta_path is not None:
            try:
                with open(meta_path, 'r', encoding='utf-8') as meta_file:
                    metadata = json.load(meta_file)
            except FileNotFoundError:
                print(f"File not found: {meta_path}")

        if title_path is not None:
            try:
                with open(title_path, 'r', encoding='utf-8') as title_file:
                    title = json.load(title_file)
            except FileNotFoundError:
                print(f"File not found: {title_path}")
    else :
        try :
            _map = None
            with open(map_json, 'r', encoding='utf-8') as map_file:
                    _map = json.load(map_file)
            metadata = {}
            title = {}

            for m in _map :
                variables = m['variables']
                qtype = m['type']
                meta = m['meta']
                grouping = m['grouping']
                mtitle = m['title']
                qlabel = m['qlabel']
                title[qlabel] = {
                    "type": qtype,
                    "title": mtitle,
                    "sub_title": None,
                    "vgroup": None
                }

        
                for v, dt in variables.items() :
                    qtitle = m['title']

                    if qtype in ['single', 'rating', 'rank', 'other_open'] :
                        metadata[v] = meta
                    else :
                        if grouping == 'rows' :
                            metadata[v] = meta[v]['colTitle']
                        else :
                            metadata[v] = meta[v]['rowTitle']
                    
                    sub_title = None
                    if grouping == 'rows' :
                        sub_title = dt['rowTitle']
                        if sub_title is None :
                            sub_title = dt['colTitle']
                    if grouping == 'cols' :
                        sub_title = dt['colTitle']
                        if sub_title is None :
                            sub_title = dt['rowTitle']

                    title[v] = {
                        "type": qtype,
                        "title": qtitle,
                        "sub_title": sub_title,
                        "vgroup": dt['vgroup']
                    }
                
                if not qlabel in metadata :
                    if grouping == "rows":
                        metadata[qlabel] = {str(k):v if isinstance(v, str) else v['colTitle'] for k, v in meta.items()}
                    else :
                        metadata[qlabel] = {str(k):v if isinstance(v, str) else v['rowTitle'] for k, v in meta.items()}

        except FileNotFoundError :
            print(f"File not found: {title_path}")
    

    default_args = {f'default_{key}': value for key, value in default_args.items()}

    return DataCheck(df, 
                     css=css, 
                     keyid=keyid,
                     meta=metadata, 
                     title=title, 
                     **default_args)
    # return DataProcessing(dc, meta=metadata, title=title, default_top=default_top, default_bottom=default_bottom)


#### Decipher Ready
def unzip_and_delete(zip_path, extract_to='.'):
    """
    Function to unzip a file and delete the zip file

    Parameters:
    zip_path (str): Path to the zip file
    extract_to (str): Directory path where the contents will be extracted (default: current directory)
    """
    try:
        # Open the zip file
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            # Extract all the contents of the zip file to the specified directory
            zip_ref.extractall(extract_to)
        
        # Delete the zip file
        os.remove(zip_path)
    
    except FileNotFoundError:
        print(f"File not found: {zip_path}")
    
    except zipfile.BadZipFile:
        print(f"Invalid zip file: {zip_path}")
    
    except Exception as e:
        print(f"An error occurred: {e}")

def ensure_directory_exists(directory_path: str) -> None:
    if not os.path.exists(directory_path):
        os.makedirs(directory_path)

def not_empty_cell(cell, row: int) -> bool:
    """셀 값이 비어 있지 않은지 확인합니다."""
    a = cell.cell(row, 1).value
    b = cell.cell(row, 2).value
    c = cell.cell(row, 3).value

    return bool(a or b or c)

def re_big(txt: str) -> Optional[str]:
    """대괄호 안의 내용을 추출합니다."""
    re_chk = re.search(r'\[(.*?)\]', txt)
    if re_chk:
        return re_chk.group(1).strip()
    return None

def colon_split(txt: str, num: int) -> Optional[str]:
    """콜론으로 텍스트를 나누고 지정된 부분을 반환합니다."""
    re_chk = txt.split(":")
    if re_chk:
        return re_chk[num].strip()
    return None


def download_decipher_data(pid: str, data_path:str , file_format: Literal['csv', 'spss', 'both'] = 'csv', folder: Optional[Union[str, List[str]]]=None, cond:Optional[str]=None):
    delivery_cond = 'qualified' if cond == None else f'qualified and {cond}'

    if not file_format in ['csv','spss', 'both'] :
        raise ValueError('The file_format must be either csv, spss or both')
    
    
    if folder is not None :
        if not isinstance(folder, (str, list)) :
            raise ValueError('The folder must be either str or list')

        if isinstance(folder, str) :
            data_path = os.path.join(data_path, folder)
        else :
            data_path = os.path.join(data_path, *folder)
        
        if not os.path.exists(data_path) :
            os.makedirs(data_path)
        
        
    if file_format in ['csv', 'both'] :
        csv_data = get_decipher_data(pid, data_format='csv', cond=delivery_cond)
        csv_binary = f'{pid}.csv'
        ensure_directory_exists(data_path)
        create_binary_file(data_path, csv_binary, csv_data)
        create_ascii_file(data_path, csv_binary)
    
    if file_format in ['spss', 'both'] :
        sav_data = get_decipher_data(pid, data_format='spss16', cond=delivery_cond)
        sav_zip = f'{pid}_sav.zip'
        create_binary_file(data_path, sav_zip, sav_data)
        unzip_and_delete(os.path.join(data_path, sav_zip), data_path)

    print("✅ Download Complete")

def DecipherSetting(pid: str, 
            cond: Optional[str] = None,
            use_variable: bool = False,
            key: str = api_key, 
            server: str = api_server, 
            meta: bool = True, 
            data_layout: bool = False, 
            base_layout: str = 'DoNotDelete',
            mkdir: bool = False,
            dir_name: Optional[str] = None,
            mode: Literal['datacheck', 'processing'] = 'datacheck') -> None:

    """
    데이터 체크 노트북 파일 및 데이터 세팅
    
    Args:
        `pid` (str): 프로젝트 ID.
        `cond` (str, optional): 데이터 필터링 조건. 기본값은 None.
        `use_variable` (bool, optional): 변수 파일 사용 여부. 기본값은 False.
        `key` (str, optional): API 키. 기본값은 api_key.
        `server` (str, optional): API 서버. 기본값은 api_server.
        `json_export` (bool, optional): JSON 내보내기 여부. 기본값은 True.
        `data_layout` (bool, optional): 데이터 레이아웃 내보내기 여부. 기본값은 False.
        `base_layout` (str, optional): 기본 레이아웃 이름. 기본값은 'DoNotDelete'.
        `datamap_name` (str, optional): 데이터 맵 이름. 기본값은 'Datamap'.
        `mkdir` (bool, optional): 디렉토리 생성 여부. 기본값은 False.
        `dir_name` (str, optional): 디렉토리 이름. 기본값은 None.
    """

    #pd.io.formats.excel.ExcelFormatter.header_style = None
    excel.ExcelFormatter.header_style = None
    
    if pid == '' or not pid :
        print('❌ Please enter pid')
        return

    parent_path = os.getcwd()
    if mkdir :
        folder_name = pid
        if dir_name != None :
            folder_name = dir_name
        parent_path =  os.path.join(parent_path, folder_name)
        chk_mkdir(parent_path)


    # META DATA
    map_py = decipher_map(pid) # import variable
    if meta :
        meta_path = os.path.join(parent_path, 'meta')
        ensure_directory_exists(meta_path)
        # metadata = decipher_meta(pid) # attr meta
        # title = decipher_title(pid) # title meta

        # with open(os.path.join(meta_path, f'meta_{pid}.json'), 'w', encoding='utf-8') as f :
        #     json.dump(metadata, f, ensure_ascii=False, indent=4)
        
        # with open(os.path.join(meta_path, f'title_{pid}.json'), 'w', encoding='utf-8') as f :
        #     json.dump(title, f, ensure_ascii=False, indent=4)

        with open(os.path.join(meta_path, f'map_{pid}.json'), 'w', encoding='utf-8') as f :
            json.dump(map_py, f, ensure_ascii=False, indent=4)

        with open(os.path.join(meta_path, f'variables_{pid}.py'), 'w', encoding='utf-8') as f :
            for mp in map_py :
                qlabel = mp['qlabel']
                variables = mp['variables']
                variables = [v for v in variables.keys()]
                qtype = mp['type']
                var_text = f"""# {qlabel} : {qtype}\n"""

                if len(variables) >= 2 :
                    for v in variables :
                        var_text += f"""{v} = '{v}'\n"""
                
                if len(variables) == 1 :
                    main_qlabel = variables[0]
                    qlabel = main_qlabel
                    variables = f"""'{main_qlabel}'"""
                
                values = mp['values'] if 'values' in mp.keys() else None
                attrs = mp['attrs'] if 'attrs' in mp.keys() else None

                var_text += f"""{qlabel} = {variables}\n"""
                if values :
                    var_text += f"""{qlabel}_value = {values}\n"""

                if attrs :
                    var_text += f"""{qlabel}_attrs = {attrs}\n"""
                var_text += "\n"
                f.write(var_text)
    #----

    # DATA DOWNLOAD
    if cond != None :
        if cond.isdigit() :
            print('❌ [ERROR] : The cond argument can only be a string')
            return
    delivery_cond = 'qualified' if cond == None else f'qualified and {cond}'
    try :
        api.login(key, server)
    except :
        print('❌ Error : Decipher api login failed')
        return

    data_path = os.path.join(parent_path, 'data')
    ensure_directory_exists(data_path)
    download_decipher_data(pid=pid, data_path=parent_path, file_format='both', folder='data', cond=delivery_cond)
    #----

    # DATA CHECK SETTING
    map_xlsx = get_decipher_datamap(pid, 'xlsx')
    create_binary_file(data_path, f'mapsheet_{pid}.xlsx', map_xlsx)

    xl = openpyxl.load_workbook(os.path.join(data_path, f'mapsheet_{pid}.xlsx'))
    map_sheet = 'datamap'
    data_map = xl[map_sheet]

    mx_row = data_map.max_row
    mx_col = data_map.max_column

    key_ids = key_vars
    diff_vars = sys_vars
    all_diff = key_ids + diff_vars

    rank_chk = ['1순위', '2순위', '1st', '2nd']

    na = 'noanswer'
    eltxt = 'element'
    col_name = ["a", "b", "c"]
    curr_var = {col:[] for col in col_name }

    variables = []

    for row in range(1, mx_row+1) :
        if not_empty_cell(data_map, row) :
            for idx, col in enumerate(range(1, mx_col+1)) :
                curr_col = col_name[idx]
                curr_dict = curr_var[curr_col]
                cell = data_map.cell(row, col)
                if cell.value or cell.value == 0: 
                    curr_dict.append(cell.value)
                    curr_var[curr_col] = curr_dict
                if cell.value == None or cell.value == "" :
                    curr_dict.append("")
                    curr_var[curr_col] = curr_dict
        else :
            variables.append(curr_var)
            curr_var = {col:[] for col in col_name }


    qids = OrderedDict()
    for variable in variables :
        # qid, type summary
        a_cell = variable['a']
        a_cell = [a for a in a_cell if a != '' and a != None]
        b_cell = variable['b']
        c_cell = variable['c']
        qid = a_cell[0] # qid

        type_chk = a_cell[1] if len(a_cell) >= 2 else None # type check

        # attribute
        main_qlabel = None
        qtype = None
        qelements = []
        qvalue = None
        qtitle = None

        # labels setting
        qlabes = {
            'values' : {b:c_cell[idx] for idx, b in enumerate(b_cell) if type(b) == int}  if c_cell else {},
            'texts' : {re_big(b):c_cell[idx] for idx, b in enumerate(b_cell) if type(b) == str or b == None} if c_cell else {}
        }
        #  find name in []

        # main q label check
        find_qname = re_big(qid.split(':')[0])
        if find_qname :
            main_qlabel = find_qname
            qelements.append(main_qlabel)
            qtitle = colon_split(qid, 1)
        else :
            main_qlabel = colon_split(qid, 0)
            qtitle = colon_split(qid, 1)

        # type check
        if type_chk :
            open_text = 'OPEN'
            if open_text.upper() in type_chk.upper() :
                qtype = 'OE'
            else :
                qtype = 'CE'

            # value check
            value_text = 'VALUES'
            if value_text.upper() in type_chk.upper() :
                qvalue = colon_split(type_chk, 1)
            else :
                qvalue = None


        else :
            qtype = 'OTHER'

        # other oe check
        oe_chk = 'oe'
        if oe_chk in main_qlabel :
            qtype = 'OTHER_OE'

        # elements setting
        for b in b_cell :
            b_chk = re_big(str(b))
            if b_chk :
                qelements.append(b_chk)

        if not 'OE' in qtype:
            # ma check
            unchk = 'Unchecked'
            c_chk = [c.upper() for c in c_cell if type(c) == str]
            if unchk.upper() in c_chk : 
                qtype = 'MA'
            else :
                # radio/number check
                int_chk = [b for b in b_cell if type(b) == int]
                if int_chk :
                    qtype = 'SA'
                else :
                    qtype = 'NUM'

        if len(qelements) >= 2 :
            if main_qlabel in qelements :
                qelements.remove(main_qlabel)

        el_labels = {key:value for key, value in qlabes['texts'].items() if key}
        
        qids[main_qlabel] = {
            'element' : qelements,
            'title' : qtitle,
            'type' : qtype,
            'value' : qvalue,
            'value_label' : qlabes['values'],
            'element_label' : el_labels,
        }

    if na in list(qids.keys()) :
        nas = qids[na]
        els = nas[eltxt]
        for el in els :
            na_name = el.split('_')[:-1]
            na_el = '_'.join(na_name).replace(na, '')
            if not na_el == '' :
                if qids[na_el] :
                    qel = qids[na_el][eltxt]
                    qel.append(el)
                    qids[na_el][eltxt] = qel

    if mode == 'datacheck' :
        # DATACHECK NOTEBOOK
        nb = nbf.v4.new_notebook()
        ipynb_file_name = f'DataCheck_{pid}.ipynb'
        order_qid = list(qids.items())

        ipynb_cell = []

        # set_file_name = 'pd.read_csv(file_name)' if mode == 'file' else 'pd.read_csv(file_name, low_memory=False)'

        csv_meta = f'''DecipherDataProcessing(df, map_json=f"meta/map_{{pid}}.json")''' if meta else '''DecipherDataProcessing(df)'''

        default = f'''import pandas as pd
import pyreadstat
import numpy as np
from meta.variables_{pid} import * 
from decipherAutomatic.dataProcessing.dataCheck import *

pid = "{pid}"

# Use SPSS
# file_name = f"data/{{pid}}.sav"
# df, meta = pyreadstat.read_sav(file_name)
# df = DecipherDataProcessing(df)

# Use CSV
file_name = f"data/{{pid}}.csv"
df = pd.read_csv(file_name, low_memory=False)
df = {csv_meta}'''
        
        ipynb_cell.append(nbf.v4.new_code_cell(default))
        ipynb_cell.append(nbf.v4.new_code_cell("""# df.display_msg = 'error'"""))

        for idx, attrs in enumerate(order_qid) :
            qid = attrs[0]
            els = attrs[1]
            if not qid in all_diff :
                qels = els['element']
                qtype = els['type']
                qval = els['value']
                qtitle = els['title']
                val_label = els['value_label']
                el_label = els['element_label']

                cell_texts = []
                cell_texts.append(f'# {qid} : {qtype}')
                # sa check #
                if qtype == 'SA' or (qtype == 'MA' and len(qels) == 1):
                    if qtype == 'SA' :
                        if qval :
                            val_chk = f"# value : {qval}"
                            cell_texts.append(val_chk)

                        if len(qels) >= 2 :
                            diff_na = [q for q in qels if not na in q]

                        for qel in qels :
                            if na in qel :
                                cell_texts.append(f'# The {qid} contains {qel}')
                            else :
                                safreq = f"df.safreq('{qel}')"
                                if use_variable : safreq = f"df.safreq({qel})"

                                cell_texts.append(safreq)

                        if val_label :
                            values = [v for v in val_label.keys() if not int(v) == 0]

                        # rank check
                        if len(qels) >= 2 :
                            labels = list(el_label.values())
                            rk = []
                            for rk_txt in rank_chk :
                                for label in labels :
                                    mu_rk = rk_txt.strip().replace(' ','').upper()
                                    mu_label = label.strip().replace(' ','').upper()
                                    if mu_rk in mu_label :
                                        rk.append(label)
                            if len(rk) >= 2 :
                                dup_diff_na = [q for q in qels if not na in q]
                                set_qid = f"('{dup_diff_na[0]}', '{dup_diff_na[-1]}')"

                                dupchk = f"df.dupchk({set_qid})"
                                if use_variable : dupchk = f"df.dupchk({qid})"

                                cell_texts.append(dupchk)
                    else :
                        if qval :
                            val_chk = f"# value : {qval}"
                            
                            cell_texts.append(val_chk)
                            safreq = f"df.safreq('{qels[0]}')"
                            if use_variable : safreq = f"df.safreq({qels[0]})"
                            cell_texts.append(safreq)
                ### sa end ###

                # ma check #
                elif qtype == 'MA' :
                    if len(qels) > 1 :
                        diff_na = [q for q in qels if not na in q]
                        if not diff_na :
                            continue
                        nas = [q for q in qels if na in q]
                        first_el = diff_na[0]
                        last_el = diff_na[-1]
                        set_qid = f"('{first_el}', '{last_el}')"


                        if val_label :
                            values = [v for v in val_label.keys() if not int(v) == 0]

                        mafreq = f"df.mafreq({set_qid})"
                        if use_variable : mafreq = f"df.mafreq({qid})"

                        cell_texts.append(mafreq)

                        if nas :
                            cell_texts.append(f'# The {qid} contains {nas}')
                ### ma end ###


                # num check #
                elif qtype == 'NUM' :
                    range_set = None

                    if len(qels) >=2 :
                        diff_na = [q for q in qels if not na in q]

                    if qval :
                        values = qval.split('-')
                        range_set = f"only=range({values[0]}, {values[1]})"

                    for qel in qels :
                        if na in qel :
                            cell_texts.append(f'# The {qid} contains {qel}')
                        else :
                            if range_set :
                                safreq = f"df.safreq('{qel}', {range_set})"
                                if use_variable : safreq = f"df.safreq({qel}, {range_set})"
                            else :
                                safreq = f"df.safreq('{qel}')"
                                if use_variable : safreq = f"df.safreq({qel})"

                            cell_texts.append(safreq)

                ### num end ###

                # text check #
                elif qtype == 'OE' :
                    if len(qels) >=2 :
                        diff_na = [q for q in qels if not na in q]

                    for qel in qels :
                        if na in qel :
                            cell_texts.append(f'# The {qid} contains {qel}')
                        else :
                            safreq = f"df.safreq('{qel}')"
                            if use_variable : safreq = f"df.safreq({qel})"

                            cell_texts.append(safreq)
                ### text end ###

                # other open check #
                elif qtype == 'OTHER_OE' :
                    for qel in qels :
                        safreq = f"df.safreq('{qel}')"
                        if use_variable : safreq = f"df.safreq({qel})"

                        cell_texts.append(safreq)
                ### other open end ###


                if cell_texts :
                    cell = '\n'.join(cell_texts)
                    ipynb_cell.append(nbf.v4.new_code_cell(cell))
                else :
                    mark = f'The {qid} not cotains elements'
                    ipynb_cell.append(nbf.v4.new_markdown_cell(mark))

        #ipynb_cell
        nb['cells'] = ipynb_cell
        #print(nb)
        ipynb_file_path = os.path.join(parent_path, ipynb_file_name)
        if not os.path.isfile(ipynb_file_path) :
            with open(ipynb_file_path, 'w', encoding='utf-8') as f:
                nbf.write(nb, f)
        else :
            print('❗ The DataCheck ipynb file already exists')
    #----

    # LAYOUT
    if data_layout :
        layouts = decipher_create_layout(pid, base_layout=base_layout, qids=qids)
        ce_layout = layouts['CE']
        oe_layout = layouts['OE']
        
        layout_path = os.path.join(parent_path, 'layout')
        ensure_directory_exists(layout_path)
        with open(os.path.join(layout_path, f'Close_Ended_{pid}.txt'), 'w', encoding='utf-8') as f :
            f.write(ce_layout)

        with open(os.path.join(layout_path, f'Open_Ended_{pid}.txt'), 'w', encoding='utf-8') as f :
            f.write(oe_layout)
    #----

    #---
    if mode == 'processing' :
        processing_ipynb = f'DataProcessing_{pid}.ipynb'
        nb = nbf.v4.new_notebook()

        ipynb_cell = []

        guide_cell = f"""##### 📌 DecipherDataProcessing Default_args  
```python  
default_args = {{  
    'top': 2,  
    'medium': True,  
    'bottom': 2,  
    'with_value': False,  
    'chat_lang': 'korean'  
}}  
    
# If you want to change the default_args, you can do it like this  
# (If you want to change the top, medium, and bottom default options,)  
df = DecipherDataProcessing([Raw DataFrame], map_json='[Map JSON Path]', default_args = {{  
    'top': [3, 2],  
    'medium' : False, # or 3 or [3, 4, 5], etc ...  
    'bottom': [2, 3],  
}})  
```  """

        ipynb_cell.append(nbf.v4.new_markdown_cell(guide_cell))


        # Python Code
        default = f"""import pandas as pd
import numpy as np
from meta.variables_{pid} import *
import os
from decipherAutomatic.dataProcessing.dataCheck import DecipherDataProcessing, download_decipher_data
from decipherAutomatic.utils import get_versioned_filename
# import pyreadstat

pid = '{pid}'
# download_decipher_data(pid=pid, data_path=os.getcwd(), file_format='csv', folder='data', cond='qualified')

raw_df = pd.read_csv(f'data/{{pid}}.csv', low_memory=False)
df = DecipherDataProcessing(raw_df, map_json=f'meta/map_{{pid}}.json')"""
        
        ipynb_cell.append(nbf.v4.new_code_cell(default))

        banner_cell = f"""# Example
# df.set_banner({{
#     '성별' : 'SQ1', # SA
#     '연령': 'SQ2', # SA
#     '브랜드': ['SQ3r1', 'SQ3r2', 'SQ3r3'] # MA
# }})"""
        ipynb_cell.append(nbf.v4.new_code_cell(banner_cell))

        for idx, var in enumerate(map_py, 1) :
            qid = var['qlabel']
            variables = var['variables']
            if qid in all_diff or qid in ['RespStats']:
                continue

            qtype = var['type']
            if qtype in ['other_open'] :
                continue
            
            meta = var['meta']
            title = var['title']
            title = title.replace('\n', ' ')
            title = title.replace('"', '\\"')
            title = title.replace("'", "\\'")
            title = title.split('?')[0]

            table_id = f"Table_T{idx}"
            cell_text = ""
            cell_text += f"# {qid} ({qtype})\n"
    
            if not variables :
                if meta and qtype in ['number', 'float', 'text', 'rating'] :
                    if len(meta) == 1 :
                        var_name = meta.keys()[0]
                        cell_text += f"""# {qid} = '{var_name}'"""

            elif len(variables) == 1 :
                var_name = list(variables.keys())[0]
                if var_name != qid :
                    qid = var_name
                
                cell_text += f"""# {qid} = '{var_name}'"""
            
            elif len(variables) >= 2 :
                var_list = [v for v in variables.keys()]
                var_list = ', '.join([f"'{v}'" for v in var_list])
                cell_text += f"# {qid} = [{var_list}]"


            cell_text += "\n"
            # print(qtype, variables)
            if isinstance(variables, dict) and len(list(variables.keys())) >= 2 :
                if qtype in ['rating'] :
                    rating_text = cell_text
                    rating_text += f"# {qid} Grid summary\n"

                    grid_var = [v for v in variables.keys()]
                    grid_var = ', '.join(grid_var)
                    rating_text += f"table = df.grid_summary([{grid_var}], fill=False)\n"
                    rating_text += f"df.proc_append(\n\t\t(f'{table_id}_sumamry', '{qid} Grid Summary'), \n\t\ttable, \n\t\tai=False\n\t)"
                    ipynb_cell.append(nbf.v4.new_code_cell(rating_text))

                if qtype in ['number', 'float', 'rating', 'single'] :
                    cell_text += f"""for idx, qid in enumerate({qid}, 1) :\n"""
                    cell_text += f"""\ttitle_dict = df.title(qid)\n"""
                    cell_text += f"""\ttitle = title_dict['title']\n"""
                    cell_text += f"""\tsub_title = title_dict['sub_title']\n"""
                    cell_text += f"""\tsub_title = qid if sub_title is None else sub_title\n"""
                    cell_text += f"""\ttable = df.proc(qid)\n"""
                    cell_text += f"""\tdf.proc_append(\n\t\t(f'{table_id}_{{idx}}', f'{{title}}_{{sub_title}}'), \n\t\ttable, \n\t\tai=False\n\t)"""
                    ipynb_cell.append(nbf.v4.new_code_cell(cell_text))
                
                if qtype in ['rank'] :
                    # 1 to max rank
                    max_rank = len(list(variables.keys()))
                    for i in range(1, max_rank+1) :
                        rank_cnt = '1' if i == 1 else f"1-{i}"
                        rank_cell_text = f"# {qid} ({qtype}) : Rank {rank_cnt}\n"
                        rank_var = [v for v in variables.keys()]
                        filt_var = rank_var[:i]
                        join_qid = ', '.join(filt_var)
                        
                        rank_cell_text += f"""table = df.proc([{join_qid}])\n"""
                        rank_cell_text += f"""df.proc_append(\n\t('{table_id}_{i}', '[Rank {rank_cnt}] {title}'), \n\ttable, \n\tai=False\n)"""
                        ipynb_cell.append(nbf.v4.new_code_cell(rank_cell_text))
                
                if qtype in ['multiple'] :
                    cell_text += f"""table = df.proc({qid})\n"""
                    cell_text += f"""df.proc_append(\n\t('{table_id}', '{title}'), \n\ttable, \n\tai=False\n)"""
                    ipynb_cell.append(nbf.v4.new_code_cell(cell_text))

                if qtype in ['text'] :
                    for idx, vid in enumerate(variables.keys(), 1) :
                        text_cellt_text = f"# {vid} ({qtype})\n"
                        text_cellt_text += f"""table = df.banner_wordcloud({vid})\n"""
                        text_cellt_text += f"""df.proc_append(\n\t('{table_id}_{idx}', '{title}'), \n\ttable\n)"""

                        ipynb_cell.append(nbf.v4.new_code_cell(text_cellt_text))
            else :
                if qtype in ['text'] :
                    cell_text += f"""table = df.banner_wordcloud({qid})\n"""
                else :
                    cell_text += f"""table = df.proc({qid})\n"""
                
                cell_text += f"""df.proc_append(\n\t('{table_id}', '{title}'), \n\ttable, \n\tai=False\n)"""

                ipynb_cell.append(nbf.v4.new_code_cell(cell_text))


        export_cell = f"""
save_file_name = get_versioned_filename(f'{pid}_table_result.xlsx')
df.proc_export_excel(save_file_name, heatmap=True)"""
        ipynb_cell.append(nbf.v4.new_code_cell(export_cell))
        
        nb['cells'] = ipynb_cell

        ipynb_file_path = os.path.join(parent_path, processing_ipynb)
        if not os.path.isfile(ipynb_file_path) :
            with open(ipynb_file_path, 'w', encoding='utf-8') as f:
                nbf.write(nb, f)
        else :
            print('❗ The Processing ipynb file already exists')


    print("✅ Setting Complete")



def shutdown_window(time: int = 0) :
    time = time * 60
    print(f"⏰ Shutdown Window : {time} min")
    time.sleep(time)
    
    print("✅ Shutdown Window")
    os.system("shutdown /s /t 1")

